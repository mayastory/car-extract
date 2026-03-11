# -*- coding: utf-8 -*-
"""ShipingList Viewer (GUI) - PyQt5 / Dark / MySQL(pymysql) - FAST

v30 FIX
- Sp_Data: CAVITY를 1/2로 강제/매핑하지 않고, ShipingList(출하) 기준의 (Tool,Cavity,prod_date) 조합만 생성
- 각 조합에 대해 AOI header(=source_file 날짜 파싱된 meas_date) 매칭 후, 존재하는 지정 FAI 데이터만 출력(데이터 없으면 행 생성 금지)

v28 FIX
- Sp_Data 엑셀: AOI 데이터 없는 행은 생성하지 않음(빈행 금지)
- Sp_Data 엑셀: Data 컬럼은 실제 데이터가 있는 최대 번호까지만 생성(Data5+ 제거)
- 소포장내역(매트릭스) 행 정렬: 차수(REV) → 날짜(prod_date) 순으로 Excel과 동일하게 정렬
- 소포장내역(매트릭스) 5종(IR/X/Y/Z/ZS) 한 줄(가로)로 쭉 배치(줄바꿈 없음)
- 블록 내부 빈공간 제거(마지막 컬럼 stretch 금지 + 표/그룹 고정폭)
- 날짜는 생산일자(prod_date) 기반 유지
- UNIQUE 리스트(수량 옆): 날짜 목록 + "차수 1,2,3,4 /Cav" 형태로 Cav 포함 표시
- 매트릭스 표 클릭 시 선택 음영으로 TRUE/FALSE 색이 날아가지 않게: 표 선택 비활성화(NoSelection)
- UNIQUE 리스트 Tool/Cav 줄맞춤: 고정폭(모노스페이스) + 패딩 정렬
- 소포장내역/메인 표 데이터셀 가운데 정렬(헤더 포함)

v27 ADD
- "특정치수 엑셀출력" 버튼 동작 복구(핸들러/스레드)
- Sp_Data_YYYYMMDD.xlsx 생성 (현재 From~To/납품처/프로젝트 기준 출하분)
- AOI DB(ipqc_aoi_*)에서 지정 치수만 수집:
  * X CARRIER: 68-1, 68-2, 70-1, 70-2
  * Y CARRIER: 94-1(C1), 94-2(C2)
  * Z CARRIER: 114-1, 114-2
- 출하된 모델만 시트 생성(1~3개) / 시트명: X CARRIER, Y CARRIER, Z CARRIER

"""
from __future__ import annotations
import traceback
import sys
import os


import datetime as _dt
import re as _re
import unicodedata as _ud
import time as _time
from dataclasses import dataclass, asdict, is_dataclass
from typing import Any, Dict, List, Optional, Tuple

import pymysql
from PyQt5 import QtCore, QtGui, QtWidgets

# ---- copy/select helpers (matrix tables + unique list) ----
def _copy_table_selection_to_clipboard(tbl: QtWidgets.QTableWidget) -> None:
    try:
        sel = tbl.selectedIndexes()
        if not sel:
            return
        rows = [i.row() for i in sel]
        cols = [i.column() for i in sel]
        r0, r1 = min(rows), max(rows)
        c0, c1 = min(cols), max(cols)
        out_lines = []
        for r in range(r0, r1 + 1):
            row_vals = []
            for c in range(c0, c1 + 1):
                it = tbl.item(r, c)
                row_vals.append("" if it is None else (it.text() or ""))
            out_lines.append("\t".join(row_vals))
        QtWidgets.QApplication.clipboard().setText("\n".join(out_lines))
    except Exception:
        pass


def _copy_list_selection_to_clipboard(lw: QtWidgets.QListWidget) -> None:
    try:
        items = lw.selectedItems()
        if not items:
            return
        txt = "\n".join([i.text() for i in items])
        QtWidgets.QApplication.clipboard().setText(txt)
    except Exception:
        pass


class _CopyKeyFilter(QtCore.QObject):
    def __init__(self, owner: QtWidgets.QWidget, kind: str):
        super().__init__(owner)
        self._owner = owner
        self._kind = kind  # "table" | "list"

    def eventFilter(self, obj, ev):
        try:
            if ev.type() == QtCore.QEvent.KeyPress:
                # Ctrl+A
                if ev.matches(QtGui.QKeySequence.SelectAll):
                    if self._kind == "table":
                        self._owner.selectAll()
                    else:
                        self._owner.selectAll()
                    return True
                # Ctrl+C
                if ev.matches(QtGui.QKeySequence.Copy):
                    if self._kind == "table":
                        _copy_table_selection_to_clipboard(self._owner)
                    else:
                        _copy_list_selection_to_clipboard(self._owner)
                    return True
        except Exception:
            pass
        return False


def _enable_copyable_table(tbl: QtWidgets.QTableWidget) -> None:
    # 드래그 선택 + Ctrl+A/Ctrl+C
    tbl.setSelectionMode(QtWidgets.QAbstractItemView.ExtendedSelection)
    tbl.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectItems)
    tbl.setFocusPolicy(QtCore.Qt.StrongFocus)
    # 선택 시 색 날아가지 않게(최대한 투명)
    try:
        tbl.setStyleSheet((tbl.styleSheet() or "") + "\nQTableWidget::item:selected{background-color: rgba(0,0,0,0);}\n")
    except Exception:
        pass
    f = _CopyKeyFilter(tbl, "table")
    tbl.installEventFilter(f)
    tbl._copy_filter = f  # keep ref


def _enable_copyable_list(lw: QtWidgets.QListWidget) -> None:
    lw.setSelectionMode(QtWidgets.QAbstractItemView.ExtendedSelection)
    lw.setFocusPolicy(QtCore.Qt.StrongFocus)
    f = _CopyKeyFilter(lw, "list")
    lw.installEventFilter(f)
    lw._copy_filter = f  # keep ref
# ---------------------------------------------------------

os.environ["QT_AUTO_SCREEN_SCALE_FACTOR"] = "1"
os.environ["QT_ENABLE_HIGHDPI_SCALING"] = "1"
os.environ["QT_SCALE_FACTOR"] = "1"

# openpyxl (xlsx export)
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter
except Exception:
    Workbook = None
    Font = None
    Alignment = None
    get_column_letter = None



# =========================
# OSX-STYLE TOAST (from Cal_timesheet_minwage_FIXED_v59.py, adapted)
# - No log files
# - Uses QSettings (registry) to remember "shown today"
# =========================

def _viewer_app_icon_path() -> Optional[str]:
    """Find app icon path (for .py run and PyInstaller exe).

    IMPORTANT:
    - We must use user's provided icon assets if present (do NOT invent/rename).
    - Primary expected filenames (as provided by user):
        * 5.25 Inch Floppy Disk.ico
        * 5.25 Inch Floppy Disk.png
    - Fallback: shipinglist/app legacy names.
    """
    bases: List[str] = []
    try:
        b = os.path.dirname(sys.argv[0])
        if b:
            bases.append(b)
    except Exception:
        pass
    try:
        bases.append(os.getcwd())
    except Exception:
        pass
    try:
        bases.append(os.path.dirname(os.path.abspath(__file__)))
    except Exception:
        pass
    b = getattr(sys, "_MEIPASS", None)
    if b:
        bases.append(b)

    seen = set()
    uniq = []
    for b in bases:
        if not b:
            continue
        b = os.path.abspath(b)
        if b in seen:
            continue
        seen.add(b)
        uniq.append(b)

    # User-specified assets (priority)
    primary = (
        "5.25 Inch Floppy Disk.ico",
        "5.25 Inch Floppy Disk.png",
    )
    fallback = (
        "shipinglist.ico",
        "shipinglist.png",
        "app.ico",
        "app.png",
    )

    for base in uniq:
        # 1) exact match first
        for fn in primary + fallback:
            p = os.path.join(base, fn)
            if os.path.exists(p):
                return p

        # 2) if primary not found, try case-insensitive match (Windows zip/unzip quirks)
        try:
            files = {f.lower(): f for f in os.listdir(base)}
            for fn in primary:
                hit = files.get(fn.lower())
                if hit:
                    p = os.path.join(base, hit)
                    if os.path.exists(p):
                        return p
        except Exception:
            pass

    return None


class ToastNotification(QtWidgets.QWidget):
    """OSX 느낌 토스트 (Windows-safe).
    - fixed size
    - overlay close/options on hover
    - shake periodically
    """
    openRequested = QtCore.pyqtSignal(str)     # payload
    activateRequested = QtCore.pyqtSignal()
    closed = QtCore.pyqtSignal(str)            # payload

    def __init__(self, title: str, message: str, payload: str = "", parent=None):
        super().__init__(None)
        self._payload = str(payload or "")
        self.payload = self._payload
        self._closed_emitted = False

        flags = QtCore.Qt.Tool | QtCore.Qt.FramelessWindowHint | QtCore.Qt.WindowStaysOnTopHint
        try:
            flags |= QtCore.Qt.WindowDoesNotAcceptFocus
        except Exception:
            pass
        self.setWindowFlags(flags)
        try:
            self.setAttribute(QtCore.Qt.WA_ShowWithoutActivating, True)
        except Exception:
            pass
        self.setFocusPolicy(QtCore.Qt.NoFocus)

        try:
            self.setAttribute(QtCore.Qt.WA_TranslucentBackground, True)
            self.setAttribute(QtCore.Qt.WA_NoSystemBackground, True)
        except Exception:
            pass
        self.setAutoFillBackground(False)
        self.setStyleSheet("background: transparent;")

        self._card_w = 420
        self._card_h = 112
        self._radius = 14
        self.setFixedSize(self._card_w, self._card_h)

        self.card = QtWidgets.QFrame(self)
        self.card.setObjectName("toast_card")
        self.card.setGeometry(0, 0, self._card_w, self._card_h)
        self.card.setStyleSheet(r"""
            QFrame#toast_card{
                background: #2b2b2b;
                border: 1px solid rgba(255,255,255,40);
                border-radius: 14px;
            }
            QLabel{ color: #f2f2f2; }
        """)

        content = QtWidgets.QWidget(self.card)
        content.setGeometry(12, 12, self._card_w - 24, self._card_h - 24)
        row = QtWidgets.QHBoxLayout(content)
        row.setContentsMargins(22, 0, 0, 0)
        row.setSpacing(12)

        try:
            row.setAlignment(QtCore.Qt.AlignVCenter)
        except Exception:
            pass
        ICON_SIZE = 64

        self.icon_lbl = QtWidgets.QLabel(content)
        self.icon_lbl.setFixedSize(ICON_SIZE, ICON_SIZE)
        self.icon_lbl.setStyleSheet("background:transparent;")
        pm = None
        try:
            ico = _viewer_app_icon_path()
            if ico:
                pm = QtGui.QIcon(ico).pixmap(ICON_SIZE, ICON_SIZE)
        except Exception:
            pm = None
        try:
            if pm is None or pm.isNull():
                pm = self.style().standardIcon(QtWidgets.QStyle.SP_ComputerIcon).pixmap(ICON_SIZE, ICON_SIZE)
        except Exception:
            pass
        try:
            if pm and (not pm.isNull()):
                self.icon_lbl.setPixmap(pm)
        except Exception:
            pass

        row.addWidget(self.icon_lbl, 0, QtCore.Qt.AlignVCenter)

        text_box = QtWidgets.QVBoxLayout()
        text_box.setContentsMargins(0, 0, 0, 0)
        text_box.setSpacing(3)


        try:
            text_box.setAlignment(QtCore.Qt.AlignVCenter)
        except Exception:
            pass
        self.title_lbl = QtWidgets.QLabel(title or "", content)
        self.title_lbl.setStyleSheet("font-weight:800; font-size:13px;")
        self.title_lbl.setFixedHeight(18)

        self.msg_lbl = QtWidgets.QLabel(message or "", content)
        self.msg_lbl.setWordWrap(True)
        self.msg_lbl.setStyleSheet("font-size:12px; color: rgba(255,255,255,215);")
        self.msg_lbl.setFixedHeight(42)

        text_box.addWidget(self.title_lbl)
        text_box.addWidget(self.msg_lbl)
        row.addLayout(text_box, 1)

        # overlay buttons
        self.btn_close = QtWidgets.QToolButton(self.card)
        self.btn_close.setText("×")
        self.btn_close.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.btn_close.setFixedSize(20, 20)
        self.btn_close.setStyleSheet(r"""
            QToolButton{
                background: rgba(0,0,0,0);
                color: rgba(255,255,255,180);
                border: 1px solid rgba(255,255,255,40);
                border-radius: 10px;
                font-weight: 900;
            }
            QToolButton:hover{
                color: rgba(255,255,255,235);
                border: 1px solid rgba(255,255,255,80);
                background: rgba(255,255,255,18);
            }
        """)
        self.btn_close.clicked.connect(self._on_close_clicked)

        self.btn_opts = QtWidgets.QToolButton(self.card)
        self.btn_opts.setText("옵션 ▾")
        self.btn_opts.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.btn_opts.setPopupMode(QtWidgets.QToolButton.InstantPopup)
        self.btn_opts.setStyleSheet(r"""
            QToolButton{
                background: rgba(255,255,255,10);
                color: rgba(255,255,255,215);
                border: 1px solid rgba(255,255,255,35);
                border-radius: 10px;
                padding: 2px 8px;
                font-size: 11px;
            }
            QToolButton:hover{
                background: rgba(255,255,255,16);
                border: 1px solid rgba(255,255,255,60);
            }
            QMenu{
                background: rgba(42,42,42,245);
                color: #f2f2f2;
                border: 1px solid rgba(255,255,255,55);
                border-radius: 12px;
                padding: 6px;
            }
            QMenu::item{
                padding: 7px 14px;
                border-radius: 10px;
                margin: 1px 2px;
                font-size: 12px;
            }
            QMenu::item:selected{
                background: rgba(255,255,255,18);
            }
            QMenu::separator{
                height: 1px;
                background: rgba(255,255,255,22);
                margin: 6px 8px;
            }
        """)

        menu = QtWidgets.QMenu(self.btn_opts)
        act_today = QtWidgets.QAction("오늘 출하 보기", menu)
        act_today.triggered.connect(lambda: self.openRequested.emit(self._payload))
        act_app = QtWidgets.QAction("프로그램 열기", menu)
        act_app.triggered.connect(self.activateRequested.emit)
        act_close = QtWidgets.QAction("닫기", menu)
        act_close.triggered.connect(self._on_close_clicked)
        menu.addAction(act_today)
        menu.addSeparator()
        menu.addAction(act_app)
        menu.addSeparator()
        menu.addAction(act_close)
        self.btn_opts.setMenu(menu)

        self.btn_close.hide()
        self.btn_opts.hide()
        self._position_overlay_buttons()

        # shake timers
        self._nudge_interval_ms = 2600
        self._shake_amp = 7
        self._shake_total_ms = 560
        self._base_pos = None
        self._shake_anim = None

        self._nudge_timer = QtCore.QTimer(self)
        self._nudge_timer.setInterval(self._nudge_interval_ms)
        self._nudge_timer.timeout.connect(self._nudge)

        self.setMouseTracking(True)
        self.card.setMouseTracking(True)

    def _position_overlay_buttons(self):
        try:
            pad = 8
            self.btn_close.move(pad, pad - 2)
            w = self.btn_opts.sizeHint().width()
            w = max(96, min(180, w))
            self.btn_opts.setFixedWidth(w)
            self.btn_opts.setFixedHeight(22)
            self.btn_opts.move(self._card_w - w - pad, self._card_h - self.btn_opts.height() - pad)
            self.btn_close.raise_()
            self.btn_opts.raise_()
        except Exception:
            pass

    def show_top_right(self, margin: int = 18):
        geo = None
        try:
            scr = QtGui.QGuiApplication.screenAt(QtGui.QCursor.pos())
            if scr is None:
                scr = QtGui.QGuiApplication.primaryScreen()
            geo = scr.availableGeometry() if scr else None
        except Exception:
            geo = None
        if geo is None:
            geo = QtWidgets.QApplication.desktop().availableGeometry()

        try:
            x = geo.x() + geo.width() - self.width() - margin
            y = geo.y() + margin
            self.move(int(x), int(y))
        except Exception:
            pass

        self._base_pos = QtCore.QPoint(self.x(), self.y())
        self.show()
        self.raise_()

        self._shake_once(amplitude=self._shake_amp, strong=True)
        self._nudge_timer.start()

    def enterEvent(self, e):
        try:
            self.btn_close.show()
            self.btn_opts.show()
            self._position_overlay_buttons()
        except Exception:
            pass
        return super().enterEvent(e)

    def leaveEvent(self, e):
        try:
            m = self.btn_opts.menu()
            if m is None or (not m.isVisible()):
                self.btn_close.hide()
                self.btn_opts.hide()
        except Exception:
            pass
        return super().leaveEvent(e)

    def mousePressEvent(self, e):
        # click toast -> activate app
        try:
            if e.button() == QtCore.Qt.LeftButton:
                self.activateRequested.emit()
        except Exception:
            pass
        return super().mousePressEvent(e)

    def _emit_closed_once(self):
        if self._closed_emitted:
            return
        self._closed_emitted = True
        try:
            self.closed.emit(self._payload)
        except Exception:
            pass

    def _on_close_clicked(self):
        try:
            self._nudge_timer.stop()
        except Exception:
            pass
        try:
            self.hide()
        except Exception:
            pass
        self._emit_closed_once()
        try:
            self.deleteLater()
        except Exception:
            pass

    def _nudge(self):
        self._shake_once(amplitude=self._shake_amp, strong=False)

    def _shake_once(self, amplitude: int = 7, strong: bool = False):
        try:
            if self._shake_anim is not None and self._shake_anim.state() == QtCore.QAbstractAnimation.Running:
                return
        except Exception:
            pass

        if self._base_pos is None:
            self._base_pos = QtCore.QPoint(self.x(), self.y())

        amp = int(amplitude if amplitude else 7)
        if strong:
            amp = int(max(amp, 8))

        base = QtCore.QPoint(self._base_pos)
        a1 = QtCore.QPoint(base.x() + amp, base.y())
        a2 = QtCore.QPoint(base.x() - amp, base.y())
        a3 = QtCore.QPoint(base.x() + int(amp * 0.6), base.y())
        a4 = QtCore.QPoint(base.x() - int(amp * 0.4), base.y())

        anim = QtCore.QPropertyAnimation(self, b"pos")
        anim.setDuration(int(self._shake_total_ms))
        anim.setEasingCurve(QtCore.QEasingCurve.InOutSine)
        anim.setKeyValueAt(0.00, base)
        anim.setKeyValueAt(0.18, a1)
        anim.setKeyValueAt(0.36, a2)
        anim.setKeyValueAt(0.56, a3)
        anim.setKeyValueAt(0.76, a4)
        anim.setKeyValueAt(1.00, base)

        def _done():
            try:
                self.move(self._base_pos)
            except Exception:
                pass

        try:
            anim.finished.connect(_done)
        except Exception:
            pass
        self._shake_anim = anim
        try:
            anim.start(QtCore.QAbstractAnimation.DeleteWhenStopped)
        except Exception:
            anim.start()


class TodayShipDialog(QtWidgets.QDialog):
    def __init__(self, ymd: str, rows: List[Dict[str, Any]], parent=None):
        super().__init__(parent)
        self.setWindowTitle(f"오늘 출하 ({ymd})")
        self.resize(560, 520)
        lay = QtWidgets.QVBoxLayout(self)

        top = QtWidgets.QLabel(f"오늘({ymd}) 출하된 모델별 합계")
        top.setStyleSheet("font-weight:800;")
        lay.addWidget(top)

        self.table = QtWidgets.QTableWidget(self)
        self.table.setColumnCount(2)
        self.table.setHorizontalHeaderLabels(["모델", "합계"])
        self.table.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
        self.table.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
        self.table.horizontalHeader().setStretchLastSection(True)
        try:
            self.table.horizontalHeader().setDefaultAlignment(QtCore.Qt.AlignCenter)
        except Exception:
            pass
        hide_row_numbers(self.table)
        lay.addWidget(self.table, 1)

        self.btn_open_matrix = QtWidgets.QPushButton("소포장내역 열기", self)
        lay.addWidget(self.btn_open_matrix)

        btn_close = QtWidgets.QPushButton("닫기", self)
        btn_close.clicked.connect(self.accept)
        lay.addWidget(btn_close)

        self.set_rows(rows)

    def set_rows(self, rows: List[Dict[str, Any]]):
        rows = list(rows or [])
        self.table.setRowCount(len(rows))
        for r, row in enumerate(rows):
            pn = str(row.get("part_name", "") or "")
            q = row.get("sum_qty", 0) or 0
            it0 = QtWidgets.QTableWidgetItem(pn)
            it1 = QtWidgets.QTableWidgetItem(f"{int(q):,}")
            it0.setTextAlignment(QtCore.Qt.AlignCenter)
            it1.setTextAlignment(QtCore.Qt.AlignCenter)
            self.table.setItem(r, 0, it0)
            self.table.setItem(r, 1, it1)

# =========================
# DB 접속정보 (여기만 수정)
# =========================
DEFAULT_DB_HOST = "dpams.ddns.net"     # 예: "dpams.ddns.net"
DEFAULT_DB_PORT = 3306
DEFAULT_DB_NAME = "dp"
DEFAULT_DB_USER = "maya"             # <- 여기
DEFAULT_DB_PASS = "##Gmlakd2323"             # <- 여기
DEFAULT_DB_CHARSET = "utf8mb4"


# 오류 로그 (일반 유저에게는 숨기고 파일로만 기록)
ERROR_LOG_PATH = None  # file logging disabled (console only)

def _log_error(tag: str, msg: str):
    # 콘솔(커널)로만 로그 출력. 일반 유저 UI에는 상세 traceback을 숨기고, 콘솔에서만 확인.
    try:
        ts = _dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    except Exception:
        ts = ""
    try:
        print(f"[{ts}] {tag}\n{msg}", file=sys.stderr)
    except Exception:
        pass

def _is_admin(user: Optional[dict]) -> bool:
    try:
        return (user or {}).get("role") == "admin"
    except Exception:
        return False

# 소포장내역(가로 매트릭스) 그룹
DEFAULT_MATRIX_GROUPS: List[Tuple[str, List[int]]] = [
    ("MEM-IR-BASE",   [1, 2, 3, 4]),
    ("MEM-X-CARRIER", [1, 2, 3, 4]),
    ("MEM-Y-CARRIER", [1, 2, 3, 4]),
    ("MEM-Z-CARRIER", [5, 6, 7, 8]),
    ("MEM-Z-STOPPER", [5, 6, 7, 8]),
]

# 소포장내역(매트릭스) 품번명 별칭 -> 5종(IR/X/Y/Z/ZS)으로 묶기
MATRIX_PART_ALIAS: Dict[str, str] = {
    "MEM-Y-CARRIER DAMPERLESS": "MEM-Y-CARRIER",
}

# =========================
# 특정치수(Sp_Data) 엑셀 출력 설정
# =========================
# - ShipingList 출하 기준으로 X/Y/Z CARRIER만 시트 생성
# - AOI 측정(DB: ipqc_aoi_*)에서 지정 FAI만 수집
# - Y CARRIER: FAI 표기는 94-1(C1), 94-2(C2) 이지만,
#   DB상 ipqc_aoi_* 는 (tool, cavity, meas_date) 단위(header_id) 안에
#   94-1(C1) / 94-2(C2) 가 같이 존재할 수 있으므로
#   (94-1(C1)/94-2(C2))는 '몰드 cavity'와 무관한 라벨이므로 cavity 제한/매핑을 하지 않는다.
#   -> 따라서 Y CARRIER도 header_id 안에 존재하는
#      94-1/94-2 측정치만 그대로 모은다.
# 형식: (mapkey, sheet_name, [(db_fai_key, out_label), ...])
SP_EXPORT_SHEETS: List[Tuple[str, str, List[Tuple[str, str]]]] = [
    ("XCARRIER", "X CARRIER", [
        ("68-1", "68-1"),
        ("68-2", "68-2"),
        ("70-1", "70-1"),
        ("70-2", "70-2"),
    ]),
    ("YCARRIER", "Y CARRIER", [
        ("94-1", "94-1(C1)"),
        ("94-2", "94-2(C2)"),
    ]),
    ("ZCARRIER", "Z CARRIER", [
        ("114-1", "114-1"),
        ("114-2", "114-2"),
    ]),
]

SP_EXPORT_ALL_FAIS: List[str] = sorted({dbk for (_mk, _sn, specs) in SP_EXPORT_SHEETS for (dbk, _lbl) in specs})
POPUP_ALLOWED = {
    "pack_date": {"label": "포장일자", "type": "date"},
    "prod_date": {"label": "생산일자", "type": "date"},
    "ann_date":  {"label": "어닐링일자", "type": "date"},
    "pack_no":   {"label": "포장번호", "type": "text"},
}

COLUMNS: List[Tuple[str, str]] = [
    ("warehouse", "창고"),
    ("pack_date", "포장일자"),
    ("prod_date", "생산일자"),
    ("ann_date", "어닐링일자"),
    ("facility", "설비"),
    ("cavity", "CAVITY"),
    ("part_code", "품번코드"),
    ("revision", "차수"),
    ("customer_part_no", "고객사 품번"),
    ("part_name", "품번명"),
    ("model", "모델"),
    ("project", "프로젝트"),
    ("small_pack_no", "소포장 NO"),
    ("tray_no", "Tray NO"),
    ("ship_to", "납품처"),
    ("qty", "출고수량"),
    ("ship_datetime", "출하일자"),
    ("customer_lot_id", "고객 LOTID"),
    ("pack_barcode", "포장바코드"),
    ("pack_no", "포장번호"),
    ("avi", "AVI"),
    ("ret_disp", "RETURNDATE"),
    ("rma_visit_cnt", "RMA"),
]
POPUP_COLS = {"pack_date", "prod_date", "ann_date", "pack_no"}


# ✅ QTableView::item background 제거 (여기가 핵심)
APP_DARK_QSS = r"""
QWidget { background: #1f1f1f; color: #e8eaed; font-size: 12px; }
QLineEdit, QComboBox, QSpinBox, QDateEdit {
  background: #121212; border: 1px solid #444; border-radius: 6px; padding: 6px;
}
QPushButton { background: #2a2a2a; border: 1px solid #444; border-radius: 8px; padding: 6px 10px; }
QPushButton:hover { background: #333; }
QPushButton:pressed { background: #222; }

QTableWidget, QTableView {
  background: #121212;
  alternate-background-color: #161616;
  gridline-color: #2b2b2b;
  border: 1px solid #333;
  selection-background-color: #2d2d2d;
  selection-color: #ffffff;
}

QHeaderView::section { background: #262626; color: #e8eaed; border: 1px solid #333; padding: 6px; }
QTableCornerButton::section { background: #262626; border: 1px solid #333; }

QGroupBox { border: 1px solid #333; border-radius: 10px; margin-top: 10px; }
QGroupBox::title { subcontrol-origin: margin; left: 10px; padding: 0 6px; color: #cfcfcf; }
QLabel#muted { color: #a0a0a0; }
QLabel#matrixTotal { font-weight: 700; padding: 6px 8px; border: 1px solid #3a3a3a; border-radius: 8px; background: #151515; }
QListWidget#matrixUnique { background: #121212; border: 1px solid #333; border-radius: 8px; }
QDateEdit::drop-down { subcontrol-origin: padding; subcontrol-position: top right; width: 18px; border-left: 1px solid #444; }
QComboBox::drop-down { subcontrol-origin: padding; subcontrol-position: top right; width: 18px; border-left: 1px solid #444; }
QComboBox::down-arrow { image: none; border: none; width: 0px; height: 0px; }
QDateEdit::down-arrow { image: none; border: none; width: 0px; height: 0px; }
"""


def valid_ymd(s: str) -> bool:
    s = (s or "").strip()
    if not _re.fullmatch(r"\d{4}-\d{2}-\d{2}", s):
        return False
    try:
        _dt.datetime.strptime(s, "%Y-%m-%d")
        return True
    except ValueError:
        return False


def ymd_to_dt_range(from_ymd: str, to_ymd: str) -> Tuple[str, str]:
    from_dt = from_ymd + " 00:00:00"
    to_plus = (_dt.datetime.strptime(to_ymd, "%Y-%m-%d").date() + _dt.timedelta(days=1)).strftime("%Y-%m-%d")
    to_dt = to_plus + " 00:00:00"
    return from_dt, to_dt


def fmt_ea(n: Any) -> str:
    try:
        v = int(n or 0)
    except Exception:
        v = 0
    return f"{v:,} EA"


def normalize_part_name(v: Any) -> str:
    """품번명(Part Name) 정규화: TRIM + 공백/개행/특수공백 통합 + 유니코드 NFKC."""
    try:
        s = "" if v is None else str(v)
    except Exception:
        s = ""
    try:
        s = _ud.normalize("NFKC", s)
    except Exception:
        pass
    # 특수 공백/제로폭/개행 제거
    s = s.replace("\u00A0", " ").replace("\u200b", " ")
    s = s.replace("\r", " ").replace("\n", " ").replace("\t", " ")
    s = _re.sub(r"\s+", " ", s).strip()
    return s


def ipqc_model_to_mapkey(model: str) -> str:
    """ipqc_view.php와 동일한 모델 키 매핑(느슨한 매칭)."""
    m = (model or "").strip()
    if not m:
        return ""
    if _re.search(r"z\s*[-_ ]?\s*stopper", m, _re.I):
        return "ZSTOPPER"
    if _re.search(r"z\s*[-_ ]?\s*carrier", m, _re.I):
        return "ZCARRIER"
    if _re.search(r"y\s*[-_ ]?\s*carrier", m, _re.I):
        return "YCARRIER"
    if _re.search(r"x\s*[-_ ]?\s*carrier", m, _re.I):
        return "XCARRIER"
    if _re.search(r"ir\s*[-_ ]?\s*base", m, _re.I):
        return "IRBASE"
    return ""


def _parse_int_cav(v: Any) -> Optional[int]:
    if v is None:
        return None
    if isinstance(v, int):
        return v
    try:
        s = str(v).strip()
    except Exception:
        return None
    if not s:
        return None
    m = _re.search(r"(\d+)", s)
    if not m:
        return None
    try:
        return int(m.group(1))
    except Exception:
        return None


def _to_number_or_text(v: Any):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return v
    s = str(v).strip()
    if s == "":
        return None
    # remove comma
    s2 = s.replace(",", "")
    try:
        if _re.fullmatch(r"-?\d+", s2):
            return int(s2)
        if _re.fullmatch(r"-?\d+(\.\d+)?", s2):
            return float(s2)
    except Exception:
        pass
    return s


def _today_yyyymmdd() -> str:
    try:
        return _dt.datetime.now().strftime("%Y%m%d")
    except Exception:
        return ""


@dataclass
class ConnInfo:
    host: str
    port: int
    user: str
    password: str
    db: str
    charset: str = DEFAULT_DB_CHARSET


def conn_info_default() -> ConnInfo:
    return ConnInfo(
        host=DEFAULT_DB_HOST,
        port=int(DEFAULT_DB_PORT),
        user=DEFAULT_DB_USER,
        password=DEFAULT_DB_PASS,
        db=DEFAULT_DB_NAME,
        charset=DEFAULT_DB_CHARSET,
    )


def db_connect(ci: ConnInfo):
    return pymysql.connect(
        host=ci.host,
        port=ci.port,
        user=ci.user,
        password=ci.password,
        database=ci.db,
        charset=ci.charset,
        cursorclass=pymysql.cursors.DictCursor,
        autocommit=True,
        connect_timeout=7,
        read_timeout=30,
        write_timeout=30,
    )


def hide_row_numbers(table: QtWidgets.QTableWidget):
    table.verticalHeader().setVisible(False)
    table.setCornerButtonEnabled(False)


def monospace_font(point_size: Optional[int] = None) -> QtGui.QFont:
    """고정폭 폰트(Consolas/FixedFont) 반환."""
    try:
        f = QtGui.QFontDatabase.systemFont(QtGui.QFontDatabase.FixedFont)
    except Exception:
        f = QtGui.QFont("Consolas")
    if point_size is not None:
        try:
            f.setPointSize(int(point_size))
        except Exception:
            pass
    return f



def build_where_and_params(
    from_date: str,
    to_date: str,
    filter_ship_to: str,
    filter_project: str,
    filter_pack_bc: str,
    filter_small_no: str,
    filter_tray_no: str,
    filter_pname: str,
    force_fallback: bool,
) -> Tuple[str, List[Any], List[str], List[Any]]:
    date_parts: List[str] = []
    date_params: List[Any] = []

    if not force_fallback:
        if from_date and to_date and valid_ymd(from_date) and valid_ymd(to_date):
            from_dt, to_dt = ymd_to_dt_range(from_date, to_date)
            date_parts += ["s.ship_datetime >= %s", "s.ship_datetime < %s"]
            date_params += [from_dt, to_dt]
        else:
            if from_date and valid_ymd(from_date):
                date_parts.append("s.ship_datetime >= %s")
                date_params.append(from_date + " 00:00:00")
            if to_date and valid_ymd(to_date):
                to_plus = (_dt.datetime.strptime(to_date, "%Y-%m-%d").date() + _dt.timedelta(days=1)).strftime("%Y-%m-%d")
                date_parts.append("s.ship_datetime < %s")
                date_params.append(to_plus + " 00:00:00")

    f_parts: List[str] = []
    f_params: List[Any] = []
    if filter_ship_to:
        f_parts.append("s.ship_to LIKE %s")
        f_params.append(f"%{filter_ship_to}%")
    if filter_project:
        # dropdown 값은 보통 정확히 일치하므로 = 사용 (LIKE보다 빠름)
        f_parts.append("s.project = %s")
        f_params.append(filter_project)
    if filter_pack_bc:
        f_parts.append("s.pack_barcode LIKE %s")
        f_params.append(f"%{filter_pack_bc}%")
    if filter_small_no:
        f_parts.append("s.small_pack_no LIKE %s")
        f_params.append(f"%{filter_small_no}%")
    if filter_tray_no:
        f_parts.append("s.tray_no LIKE %s")
        f_params.append(f"%{filter_tray_no}%")
    if filter_pname:
        f_parts.append("s.part_name LIKE %s")
        f_params.append(f"%{filter_pname}%")

    merged_parts = date_parts + f_parts
    merged_params = date_params + f_params
    where_sql = ("WHERE " + " AND ".join(merged_parts)) if merged_parts else ""
    return where_sql, merged_params, f_parts, f_params


def query_all(ci: ConnInfo, from_date: str, to_date: str, ship_to: str, project: str, pack_bc: str,
              small_no: str, tray_no: str, part_name: str, force_fallback: bool = False) -> dict:
    where_sql, params, filter_parts, filter_params = build_where_and_params(
        from_date, to_date, ship_to, project, pack_bc, small_no, tray_no, part_name, force_fallback
    )
    did_fallback = False

    with db_connect(ci) as conn:
        with conn.cursor() as cur:
            cur.execute(f"SELECT COUNT(*) AS c FROM ShipingList s {where_sql}", params)
            total = int((cur.fetchone() or {}).get("c") or 0)

            list_sql = f"""
                SELECT s.*, r.last_return_datetime
                FROM ShipingList s
                LEFT JOIN (
                    SELECT small_pack_no,
                           MAX(COALESCE(return_datetime, return_date)) AS last_return_datetime
                    FROM rmalist
                    GROUP BY small_pack_no
                ) r ON r.small_pack_no = s.small_pack_no
                {where_sql}
                ORDER BY s.ship_datetime DESC, s.id DESC
            """
            cur.execute(list_sql, params)
            rows = list(cur.fetchall() or [])

            sum_sql = f"""
                SELECT TRIM(s.part_name) AS part_name, SUM(s.qty) AS sum_qty
                FROM ShipingList s
                {where_sql}
                GROUP BY TRIM(s.part_name)
                ORDER BY TRIM(s.part_name)
            """
            cur.execute(sum_sql, params)
            sum_rows = list(cur.fetchall() or [])

    # 품번명별 합계 중복 방지: TRIM/공백/개행/유니코드 정규화로 동일 품번명은 합산
    try:
        merged: Dict[str, int] = {}
        for sr in (sum_rows or []):
            nm = normalize_part_name((sr or {}).get("part_name"))
            if not nm:
                continue
            try:
                qv = int((sr or {}).get("sum_qty") or 0)
            except Exception:
                qv = 0
            merged[nm] = merged.get(nm, 0) + qv
        sum_rows = [{"part_name": k, "sum_qty": v} for (k, v) in sorted(merged.items(), key=lambda kv: kv[0])]
    except Exception:
        pass


    for r in rows:
        ret = (r.get("last_return_datetime") or "") or (r.get("return_date") or "") or (r.get("return_datetime") or "")
        if isinstance(ret, str) and ret.startswith("0000-00-00"):
            ret = ""
        r["ret_disp"] = ret

    return dict(rows=rows, total=total, did_fallback=did_fallback, sum_rows=sum_rows)



def query_ship_summary_for_day(ci: ConnInfo, ymd: str) -> dict:
    """오늘 출하 알림용: 날짜(YYYY-MM-DD) 기준 품번명별 합계.
    - 기준: ShipingList.ship_datetime 날짜
    - ship_to/project 필터 없음(전체)
    """
    ymd = (ymd or "").strip()
    if not valid_ymd(ymd):
        return {"ymd": ymd, "sum_rows": [], "row_count": 0, "total_qty": 0}

    dt_from, dt_to = ymd_to_dt_range(ymd, ymd)  # [ymd 00:00, nextday 00:00)
    params = [dt_from, dt_to]
    sql = """
        SELECT TRIM(part_name) AS part_name, SUM(qty) AS sum_qty
        FROM ShipingList
        WHERE ship_datetime >= %s AND ship_datetime < %s
        GROUP BY TRIM(part_name)
        ORDER BY TRIM(part_name)
    """
    with db_connect(ci) as conn:
        with conn.cursor() as cur:
            cur.execute(sql, params)
            rows = list(cur.fetchall() or [])

    merged: Dict[str, int] = {}
    for r in rows:
        pn = normalize_part_name(r.get("part_name"))
        if not pn:
            continue
        try:
            qv = int(r.get("sum_qty") or 0)
        except Exception:
            qv = 0
        merged[pn] = merged.get(pn, 0) + qv

    sum_rows = [{"part_name": k, "sum_qty": v} for (k, v) in sorted(merged.items(), key=lambda kv: kv[0])]
    total_qty = 0
    try:
        total_qty = int(sum(v for v in merged.values()))
    except Exception:
        total_qty = 0

    return {"ymd": ymd, "sum_rows": sum_rows, "row_count": len(sum_rows), "total_qty": total_qty}

def is_valid_ymd(s: str) -> bool:
    s = (s or "").strip()
    if not s:
        return False
    # accept YYYY-MM-DD only
    if not _re.match(r"^\d{4}-\d{2}-\d{2}$", s):
        return False
    try:
        y, mo, d = s.split("-")
        y = int(y); mo = int(mo); d = int(d)
        if y < 1900 or y > 2100:
            return False
        if mo < 1 or mo > 12:
            return False
        if d < 1 or d > 31:
            return False
        return True
    except Exception:
        return False




@dataclass
class SoPojangGroup:
    part_name: str
    cavs: list
    total_qty: int
    rows: list

@dataclass
class SoPojangResult:
    db_ms: int
    groups: list


def query_sopojang_matrix_fast(ci: ConnInfo, from_date: str, to_date: str, ship_to: str, project: str, groups_def: Optional[List[Tuple[str, List[int]]]] = None) -> SoPojangResult:
    """
    Excel의 =SORT(UNIQUE(TOCOL(B2:C200,3,TRUE))) 처럼
    (차수/날짜)에서 빈값은 무시하고, 실수량(수량>0)만 기준으로 매트릭스를 만든다.
    """
    def _parse_cav(v):
        if v is None:
            return None
        # numeric
        if isinstance(v, int):
            return v if 1 <= v <= 8 else None
        if isinstance(v, float):
            try:
                iv = int(v)
                return iv if abs(v - iv) < 1e-9 and 1 <= iv <= 8 else None
            except Exception:
                return None
        s = str(v).strip()
        if not s:
            return None
        m = _re.search(r"(\d+)", s)
        if not m:
            return None
        try:
            iv = int(m.group(1))
        except Exception:
            return None
        return iv if 1 <= iv <= 8 else None

    # 조건: 날짜 + 납품처 + 프로젝트
    params = []
    where = []
    if is_valid_ymd(from_date) and is_valid_ymd(to_date):
        dt_from, dt_to = ymd_to_dt_range(from_date, to_date)
        where.append("ship_datetime >= %s AND ship_datetime < %s")
        params += [dt_from, dt_to]

    ship_to = (ship_to or "").strip()
    project = (project or "").strip()
    if ship_to == "전체":
        ship_to = ""
    if project == "전체":
        project = ""

    if ship_to:
        where.append("ship_to = %s")
        params.append(ship_to)
    if project:
        where.append("project = %s")
        params.append(project)

    where_sql = ("WHERE " + " AND ".join(where)) if where else ""

    sql = f"""
        SELECT
            part_name,
            revision,
            DATE(prod_date) AS prod_date,
            cavity,
            SUM(qty) AS sum_qty
        FROM ShipingList
        {where_sql}
        GROUP BY TRIM(part_name), TRIM(revision), DATE(prod_date), cavity
    """

    t0 = _time.time()
    with db_connect(ci) as conn:
        with conn.cursor() as cur:
            cur.execute(sql, params)
            rows = list(cur.fetchall() or [])
    db_ms = int((_time.time() - t0) * 1000)

    # agg: (pn, rev, date, cav) -> qty
    agg = {}
    cav_seen = {}
    for row in rows:
        pn = normalize_part_name(row.get("part_name") or "")
        pn = MATRIX_PART_ALIAS.get(pn, pn)
        rev = normalize_part_name(row.get("revision") or "")
        d = str(row.get("prod_date") or "").strip()
        cav_i = _parse_cav(row.get("cavity"))
        try:
            q = int(row.get("sum_qty") or 0)
        except Exception:
            q = 0

        # 빈값/무의미(0) 제외 (Excel의 ignore=3 느낌)
        if not pn or not rev or not d:
            continue
        if q <= 0:
            continue
        if cav_i is None:
            continue

        key = (pn, rev, d, cav_i)
        agg[key] = agg.get(key, 0) + q
        cav_seen.setdefault(pn, []).append(cav_i)

    # part totals (qty>0 기준)
    totals = {}
    for (pn, rev, d, cav), q in agg.items():
        totals[pn] = totals.get(pn, 0) + q

    def choose_cavs(pn: str, default_cavs):
        """Cav 헤더는 데이터 기반으로 1~4 또는 5~8 중 하나만 선택."""
        cavs = [c for c in (cav_seen.get(pn, []) or []) if isinstance(c, int) and 1 <= c <= 8]
        # 기본값: 정의된 그룹(1~4 또는 5~8) 유지
        def _default4():
            try:
                mx = max(default_cavs or [])
                mn = min(default_cavs or [])
                if mx <= 4:
                    return [1, 2, 3, 4]
                if mn >= 5:
                    return [5, 6, 7, 8]
            except Exception:
                pass
            return [1, 2, 3, 4]

        if not cavs:
            return _default4()

        c14 = sum(1 for c in cavs if 1 <= c <= 4)
        c58 = sum(1 for c in cavs if 5 <= c <= 8)
        if c14 > c58:
            return [1, 2, 3, 4]
        if c58 > c14:
            return [5, 6, 7, 8]
        # 동률이면 기본 그룹을 우선
        return _default4()

    groups = []
    _defs = groups_def or DEFAULT_MATRIX_GROUPS
    for pn, default_cavs in _defs:
        # pn별 keys (rev/date) - 빈값은 이미 agg에서 제거됨
        keys = [(rev, d) for (pp, rev, d, cav) in agg.keys() if pp == pn]

        cav_cols = choose_cavs(pn, default_cavs)
        key_map = {}
        for (pp, rev, d, cav), q in agg.items():
            if pp != pn:
                continue
            k = (rev, d)
            if k not in key_map:
                key_map[k] = {"cavs": {c: False for c in cav_cols}, "qty": 0}
            if cav in key_map[k]["cavs"]:
                key_map[k]["cavs"][cav] = True
            key_map[k]["qty"] += q

        # qty>0 기준으로 rows 구성 (혹시라도 guard)
        keys_sorted = sorted(
            [k for k, v in key_map.items() if v.get("qty", 0) > 0],
            # ✅ Excel 화면처럼: 차수(REV) -> 날짜(prod_date) 순으로 정렬
            key=lambda x: (
                (0, int(x[0])) if _re.match(r"^\d+$", str(x[0] or "")) else (1, str(x[0] or "")),
                str(x[1] or "")
            )
        )
        rows_out = []
        for rev, d in keys_sorted:
            rec = key_map[(rev, d)]
            rows_out.append({
                "part_name": pn,
                "model": pn.replace("MEM-", ""),
                "rev": rev,
                "date": d,
                "cavs": rec["cavs"],
                "qty": rec["qty"],
            })

        groups.append(SoPojangGroup(part_name=pn, cavs=cav_cols, total_qty=totals.get(pn, 0), rows=rows_out))

    return {"from": from_date, "to": to_date, "db_ms": db_ms, "groups": [asdict(g) for g in groups]}


def query_ship_triples_for_sp_export(ci: ConnInfo, from_date: str, to_date: str, ship_to: str, project: str) -> Dict[str, Any]:
    """ShipingList(출하) 기준으로 (Tool=revision, Cav=cavity, Date=prod_date) 묶음 추출.
    - 대상: X/Y/Z CARRIER만
    - 결과: {mapkey: {(tool,cav,date): ship_qty}}
    """
    params: List[Any] = []
    where: List[str] = []

    if is_valid_ymd(from_date) and is_valid_ymd(to_date):
        dt_from, dt_to = ymd_to_dt_range(from_date, to_date)
        where.append("ship_datetime >= %s AND ship_datetime < %s")
        params += [dt_from, dt_to]

    ship_to = (ship_to or "").strip()
    project = (project or "").strip()
    if ship_to == "전체":
        ship_to = ""
    if project == "전체":
        project = ""
    if ship_to:
        where.append("ship_to = %s")
        params.append(ship_to)
    if project:
        where.append("project = %s")
        params.append(project)

    where_sql = ("WHERE " + " AND ".join(where)) if where else ""

    sql = f"""
        SELECT
            part_name,
            revision,
            DATE(prod_date) AS prod_date,
            cavity,
            SUM(qty) AS sum_qty
        FROM ShipingList
        {where_sql}
        GROUP BY TRIM(part_name), TRIM(revision), DATE(prod_date), cavity
    """

    with db_connect(ci) as conn:
        with conn.cursor() as cur:
            cur.execute(sql, params)
            rows = list(cur.fetchall() or [])

    shipped: Dict[str, Dict[Tuple[str, int, str], int]] = {}
    date_min: Optional[str] = None
    date_max: Optional[str] = None

    for r in rows:
        pn = normalize_part_name(r.get("part_name") or "")
        pn = MATRIX_PART_ALIAS.get(pn, pn)
        mk = ipqc_model_to_mapkey(pn)
        if mk not in {"XCARRIER", "YCARRIER", "ZCARRIER"}:
            continue

        tool = normalize_part_name(r.get("revision") or "")
        d = str(r.get("prod_date") or "").strip()
        cav = _parse_int_cav(r.get("cavity"))
        try:
            qty = int(r.get("sum_qty") or 0)
        except Exception:
            qty = 0

        if not tool or not d or cav is None or qty <= 0:
            continue

        shipped.setdefault(mk, {})[(tool, int(cav), d)] = shipped.setdefault(mk, {}).get((tool, int(cav), d), 0) + qty

        if date_min is None or d < date_min:
            date_min = d
        if date_max is None or d > date_max:
            date_max = d

    return {"shipped": shipped, "date_min": date_min, "date_max": date_max}


def _fetch_ipqc_aoi_part_candidates(ci: ConnInfo, date_min: str, date_max: str) -> Dict[str, List[str]]:
    """ipqc_aoi_header에서 part_name 후보를 뽑아 mapkey별로 분류."""
    out: Dict[str, List[str]] = {"XCARRIER": [], "YCARRIER": [], "ZCARRIER": []}
    if not (date_min and date_max):
        return out
    sql = (
        "SELECT DISTINCT part_name FROM ipqc_aoi_header "
        "WHERE meas_date IS NOT NULL AND DATE(meas_date) >= %s AND DATE(meas_date) <= %s"
    )
    with db_connect(ci) as conn:
        with conn.cursor() as cur:
            cur.execute(sql, (date_min, date_max))
            names = [normalize_part_name(x.get("part_name") or "") for x in (cur.fetchall() or [])]
    seen = set()
    for n in names:
        if not n or n in seen:
            continue
        seen.add(n)
        mk = ipqc_model_to_mapkey(n)
        if mk in out:
            out[mk].append(n)
    return out


def export_sp_data_xlsx(ci: ConnInfo, from_date: str, to_date: str, ship_to: str, project: str, save_path: str) -> Dict[str, Any]:
    """특정치수 Sp_Data 엑셀(xlsx) 생성.
    - ShipingList 출하 기준(현재 From~To/납품처/프로젝트)
    - ipqc AOI DB에서 지정 FAI만 수집
    """
    if Workbook is None:
        raise RuntimeError("openpyxl 모듈이 필요합니다.")

    base = query_ship_triples_for_sp_export(ci, from_date, to_date, ship_to, project)
    shipped: Dict[str, Dict[Tuple[str, int, str], int]] = base.get("shipped") or {}
    date_min = base.get("date_min")
    date_max = base.get("date_max")

    # 출하된 모델이 하나도 없으면 종료
    if not any(shipped.get(k) for (k, _sn, _f) in SP_EXPORT_SHEETS):
        return {"ok": 0, "msg": "해당 조건에서 X/Y/Z CARRIER 출하 데이터가 없습니다."}

    # 1) AOI header에서 part_name 후보
    part_candidates = _fetch_ipqc_aoi_part_candidates(ci, str(date_min or ""), str(date_max or ""))

    # 2) AOI header 매칭: (mapkey, tool, cav, date) -> header_id
    header_best: Dict[Tuple[str, str, int, str], int] = {}
    header_id_to_info: Dict[int, Tuple[str, str, int, str]] = {}

    # 쿼리 범위: date_min~date_max
    date_min_q = str(date_min or "")
    date_max_q = str(date_max or "")
    if not date_min_q or not date_max_q:
        # 방어: ship 결과가 있는데 날짜가 없다면 전체를 쿼리하지 않고 종료
        return {"ok": 0, "msg": "출하(prod_date) 날짜를 확인할 수 없습니다."}

    # part_name IN 후보(전체 union)
    all_parts: List[str] = []
    for mk in ("XCARRIER", "YCARRIER", "ZCARRIER"):
        for nm in (part_candidates.get(mk) or []):
            if nm not in all_parts:
                all_parts.append(nm)

    # 후보가 없으면(낮은 확률) LIKE로 fallback
    where_parts = ""
    part_params: List[Any] = []
    if all_parts:
        placeholders = ",".join(["%s"] * len(all_parts))
        where_parts = f" AND h.part_name IN ({placeholders})"
        part_params = list(all_parts)
    else:
        where_parts = (
            " AND (h.part_name LIKE '%X%CARRIER%' OR h.part_name LIKE '%Y%CARRIER%' OR h.part_name LIKE '%Z%CARRIER%')"
        )

    sql_h = (
        "SELECT h.id, h.part_name, DATE(h.meas_date) AS d, h.tool, h.cavity "
        "FROM ipqc_aoi_header h "
        "WHERE h.meas_date IS NOT NULL AND DATE(h.meas_date) >= %s AND DATE(h.meas_date) <= %s" +
        where_parts
    )

    with db_connect(ci) as conn:
        with conn.cursor() as cur:
            cur.execute(sql_h, [date_min_q, date_max_q] + part_params)
            hdr_rows = list(cur.fetchall() or [])

    shipped_sets: Dict[str, set] = {k: set((shipped.get(k) or {}).keys()) for k in shipped.keys()}

    for r in hdr_rows:
        mk = ipqc_model_to_mapkey(normalize_part_name(r.get("part_name") or ""))
        if mk not in shipped_sets:
            continue
        tool = normalize_part_name(r.get("tool") or "")
        cav = _parse_int_cav(r.get("cavity"))
        d = str(r.get("d") or "").strip()
        if not tool or cav is None or not d:
            continue
        key_trip = (tool, int(cav), d)
        if key_trip not in shipped_sets[mk]:
            continue
        try:
            hid = int(r.get("id") or 0)
        except Exception:
            continue
        kfull = (mk, tool, int(cav), d)
        if header_best.get(kfull, 0) < hid:
            header_best[kfull] = hid

    for kfull, hid in header_best.items():
        header_id_to_info[hid] = kfull

    header_ids = sorted(header_id_to_info.keys())

    # 3) AOI measurements: header_id + fai + spc + row_index + value
    #    - FAI는 (NNN-N) 형태를 기준으로 매칭 (예: 68-1)
    targets = set(SP_EXPORT_ALL_FAIS)

    def _fai_regex_for(db_key: str) -> str:
        s = str(db_key or "").strip()
        m = _re.match(r"^(\d+)\s*[-_]\s*(\d+)$", s)
        if m:
            a = int(m.group(1)); b = int(m.group(2))
            return rf"(^|[^0-9]){a}\s*[-_]\s*{b}($|[^0-9])"
        # fallback: digits-only
        if _re.fullmatch(r"\d+", s):
            return rf"(^|[^0-9]){int(s)}($|[^0-9])"
        # generic escape
        esc = _re.escape(s)
        return rf"(^|\s){esc}($|\s)"

    fai_patterns = [_fai_regex_for(k) for k in SP_EXPORT_ALL_FAIS]

    # (mk, tool, cav, date, fai_key) -> {'spc':..., 'data':[...16...]}
    meas_map: Dict[Tuple[str, str, int, str, str], Dict[str, Any]] = {}

    def _match_fai_key(fai_text: str) -> Optional[str]:
        s = str(fai_text or "")
        # Extract first NNN-N token
        m = _re.search(r"(\d+)\s*[-_]\s*(\d+)", s)
        if m:
            k = f"{int(m.group(1))}-{int(m.group(2))}"
            return k if k in targets else None
        # digits-only token
        m2 = _re.search(r"\b(\d+)\b", s)
        if m2:
            k2 = str(int(m2.group(1)))
            return k2 if k2 in targets else None
        return None

    if header_ids:
        chunk = 400
        with db_connect(ci) as conn:
            with conn.cursor() as cur:
                for i0 in range(0, len(header_ids), chunk):
                    part = header_ids[i0:i0+chunk]
                    ph = ",".join(["%s"] * len(part))
                    where_fai = " OR ".join(["m.fai REGEXP %s"] * len(fai_patterns))
                    sql_m = (
                        f"SELECT m.header_id, m.fai, m.spc, m.row_index, m.value "
                        f"FROM ipqc_aoi_measurements m "
                        f"WHERE m.header_id IN ({ph}) AND (" + where_fai + ") "
                        f"AND m.row_index BETWEEN 1 AND 16"
                    )
                    cur.execute(sql_m, list(part) + fai_patterns)
                    for r in (cur.fetchall() or []):
                        try:
                            hid = int(r.get("header_id") or 0)
                        except Exception:
                            continue
                        info = header_id_to_info.get(hid)
                        if not info:
                            continue
                        mk, tool, cav, d = info
                        fai_key = _match_fai_key(r.get("fai") or "")
                        if fai_key is None:
                            continue
                        idx = _parse_int_cav(r.get("row_index"))
                        if idx is None or not (1 <= idx <= 16):
                            continue
                        key = (mk, tool, int(cav), d, str(fai_key))
                        rec = meas_map.get(key)
                        if rec is None:
                            rec = {"spc": normalize_part_name(r.get("spc") or ""), "data": [None] * 16}
                            meas_map[key] = rec
                        # 첫 spc가 비어있으면 채우기
                        if not rec.get("spc"):
                            rec["spc"] = normalize_part_name(r.get("spc") or "")
                        rec["data"][int(idx) - 1] = _to_number_or_text(r.get("value"))

    # 4) Workbook 작성
    wb = Workbook()
    # remove default sheet
    try:
        if wb.worksheets:
            wb.remove(wb.worksheets[0])
    except Exception:
        pass

    header_font = Font(bold=True) if Font else None
    center = Alignment(horizontal="center", vertical="center") if Alignment else None
    center_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True) if Alignment else None

    def _has_value(v: Any) -> bool:
        if v is None:
            return False
        if isinstance(v, (int, float)):
            return True
        try:
            s = str(v).strip()
        except Exception:
            return False
        return s != ""

    made = 0

    for mk, sheet_name, specs in SP_EXPORT_SHEETS:
        ship_map = shipped.get(mk) or {}
        if not ship_map:
            continue
        # rows: sort like AOI view (Tool -> Date -> Cavity)
        def _tool_key(t: str):
            ts = str(t or "")
            if _re.fullmatch(r"\d+", ts):
                return (0, int(ts))
            return (1, ts)

        triplets = sorted(ship_map.items(), key=lambda kv: (_tool_key(kv[0][0]), str(kv[0][2]), int(kv[0][1])))

        # 1) 먼저 AOI 데이터가 있는 행만 수집하고, Data/편차 컬럼 수를 최대로 맞춘다.
        #    - 편차값: (pair1 - pair2) 를 Data 1..N별로 계산해서 '값'으로 넣는다(엑셀 함수/수식 금지).
        #    - pair 구조는 specs의 순서를 2개씩 묶는다. (예: 68-1/68-2, 70-1/70-2)
        sheet_rows: List[Tuple[List[Any], List[Any], List[Any]]] = []  # (base6, data16, dev16)
        merge_pairs: List[Tuple[int,int]] = []  # sheet_rows index pairs for vertical-merge of dev columns
        max_data = 0

        def _normalize_rec(rec: Optional[Dict[str, Any]]) -> Optional[Tuple[str, List[Any], int]]:
            if not rec:
                return None
            spc = (rec or {}).get("spc") or ""
            data = (rec or {}).get("data") or []
            try:
                data = list(data)
            except Exception:
                data = []
            if len(data) < 16:
                data = data + [None] * (16 - len(data))
            # 데이터가 전부 비어있으면 스킵
            if not any(_has_value(v) for v in data):
                return None
            last = 0
            for ii, v in enumerate(data, start=1):
                if _has_value(v):
                    last = ii
            return (str(spc), data, int(last))

        def _as_num(v: Any) -> Optional[float]:
            if v is None:
                return None
            if isinstance(v, (int, float)):
                return float(v)
            try:
                s = str(v).strip().replace(",", "")
            except Exception:
                return None
            if not s:
                return None
            try:
                return float(s)
            except Exception:
                return None

        for (tool, cav, d), qty in triplets:
            qty_written = False  # Ship_Qty는 (Tool,Cavity,Date) 묶음에서 1회만 표시
            # 각 spec의 rec를 미리 가져와둔다
            rec_by_fai: Dict[str, Optional[Tuple[str, List[Any], int]]] = {}
            for (db_fai, _out_label) in specs:
                rec_by_fai[str(db_fai)] = _normalize_rec(meas_map.get((mk, tool, int(cav), d, str(db_fai))))

            si = 0
            while si < len(specs):
                db1, lbl1 = specs[si]
                db2, lbl2 = (specs[si + 1] if si + 1 < len(specs) else (None, None))

                r1 = rec_by_fai.get(str(db1)) if db1 is not None else None
                r2 = rec_by_fai.get(str(db2)) if db2 is not None else None        # r1 행
                idx1 = None
                idx2 = None
                if r1 is not None:
                    spc1, data1, last1 = r1
                    if last1 > max_data:
                        max_data = last1
                    dev1 = [None] * 16
                    # 짝(r2)이 있으면 편차 계산 (v1 - v2)
                    if r2 is not None:
                        _spc2, data2, _last2 = r2
                        for ii in range(16):
                            v1 = _as_num(data1[ii])
                            v2 = _as_num(data2[ii])
                            if v1 is None or v2 is None:
                                dev1[ii] = None
                            else:
                                dev1[ii] = (v1 - v2)
                    ship_qty = (int(qty) if not qty_written else None)
                    base6 = [tool, int(cav), d, ship_qty, str(lbl1), str(spc1)]
                    qty_written = True
                    sheet_rows.append((base6, data1, dev1))
                    idx1 = len(sheet_rows) - 1

                # r2 행 (편차는 위(r1) 행에 표시하고, r2는 공백)
                if r2 is not None:
                    spc2, data2, last2 = r2
                    if last2 > max_data:
                        max_data = last2
                    ship_qty = (int(qty) if not qty_written else None)
                    base6 = [tool, int(cav), d, ship_qty, str(lbl2), str(spc2)]
                    qty_written = True
                    sheet_rows.append((base6, data2, [None] * 16))
                    idx2 = len(sheet_rows) - 1

                # (디스플레이) 편차값은 r1/r2 두 행에 걸쳐 세로 병합해서 한 번만 보이게
                if idx1 is not None and idx2 is not None:
                    merge_pairs.append((idx1, idx2))

                si += 2
# 2) 시트 생성/헤더 작성 (Data N + 편차값)
        ws = wb.create_sheet(title=sheet_name)
        if max_data < 0:
            max_data = 0
        header_row = (
            ["Tool", "Cavity", "Date", "Ship_Qty", "FAI", "SPC"]
            + [f"Data {i}" for i in range(1, max_data + 1)]
            + ([""] if max_data > 0 else [])
            + [f"Data {i}\n편차값" for i in range(1, max_data + 1)]
        )
        ws.append(header_row)

        # style header
        if header_font or center:
            for c in range(1, len(header_row) + 1):
                cell = ws.cell(row=1, column=c)
                if header_font:
                    cell.font = header_font
                if center_wrap:
                    cell.alignment = center_wrap

        # header row height (wrap)
        try:
            ws.row_dimensions[1].height = 36
        except Exception:
            pass

        # 3) 데이터 행 추가 (필요한 Data/편차 컬럼까지만)
        cur_key = None  # (Tool, Cavity, Date)
        start_row = None
        excel_row = 2  # header is row 1
        for base6, data, dev in sheet_rows:
            key = (base6[0], base6[1], base6[2])
            if cur_key is None:
                cur_key = key
                start_row = excel_row
            elif key != cur_key:
                end_row = excel_row - 1
                if start_row is not None and end_row > start_row:
                    for _col in (1, 2, 3, 4):
                        try:
                            ws.merge_cells(start_row=int(start_row), start_column=int(_col), end_row=int(end_row), end_column=int(_col))
                            if center:
                                ws.cell(row=int(start_row), column=int(_col)).alignment = center
                        except Exception:
                            pass
                cur_key = key
                start_row = excel_row
        
            out = list(base6)
            if max_data > 0:
                out += list(data[:max_data])
                # blank separator column between Data and Deviation
                out += [None]
                out += list(dev[:max_data])
            ws.append(out)
            excel_row += 1
        
        # last group merge
        end_row = excel_row - 1
        if start_row is not None and end_row > start_row:
            for _col in (1, 2, 3, 4):
                try:
                    ws.merge_cells(start_row=int(start_row), start_column=int(_col), end_row=int(end_row), end_column=int(_col))
                    if center:
                        ws.cell(row=int(start_row), column=int(_col)).alignment = center
                except Exception:
                    pass

# freeze
        try:
            ws.freeze_panes = "A2"
        except Exception:
            pass

        # center align all data cells
        if center:
            for r in range(2, ws.max_row + 1):
                for c in range(1, len(header_row) + 1):
                    ws.cell(row=r, column=c).alignment = center

        # Ship_Qty: thousands separator format (ex: 6,480)
        try:
            for r in range(2, ws.max_row + 1):
                ws.cell(row=r, column=4).number_format = "#,##0"
        except Exception:
            pass
        # column widths (simple)
        if get_column_letter:
            try:
                # base columns
                base_w = {
                    1: 8,   # Tool
                    2: 8,   # Cav
                    3: 12,  # Date
                    4: 10,  # Ship_Qty
                    5: 12,  # FAI
                    6: 8,   # SPC
                }
                for c, w in base_w.items():
                    ws.column_dimensions[get_column_letter(c)].width = w

                # Data columns
                for c in range(7, 7 + int(max_data)):
                    ws.column_dimensions[get_column_letter(c)].width = 10

                # Blank separator column + Dev columns (after Data)
                if int(max_data) > 0:
                    blank_col = 7 + int(max_data)
                    ws.column_dimensions[get_column_letter(blank_col)].width = 3

                    dev_start_col = blank_col + 1
                    for c in range(dev_start_col, dev_start_col + int(max_data)):
                        ws.column_dimensions[get_column_letter(c)].width = 14
            except Exception:
                pass

        # vertical-merge dev cells across pair rows (clean display; no blank/skip)
        if max_data > 0 and merge_pairs:
            dev_start_col = 8 + int(max_data)
            for i1, i2 in merge_pairs:
                r1 = 2 + int(i1)
                r2 = 2 + int(i2)
                if r2 <= r1:
                    continue
                for c in range(dev_start_col, dev_start_col + int(max_data)):
                    try:
                        ws.merge_cells(start_row=r1, start_column=c, end_row=r2, end_column=c)
                        if center:
                            ws.cell(row=r1, column=c).alignment = center
                    except Exception:
                        pass

        made += 1

    # 저장
    os.makedirs(os.path.dirname(save_path) or ".", exist_ok=True)
    wb.save(save_path)
    return {"ok": 1, "path": save_path, "sheets": made}


def log_to_console(msg: str):
    """디버그/에러 메시지는 UI가 아니라 콘솔(커널)에만 출력."""
    try:
        print(msg, file=sys.stderr)
    except Exception:
        pass


def detect_ship_date_col(conn) -> str:
    """ShipingList 날짜 컬럼 후보를 스캔해서 실제 존재하는 컬럼을 반환."""
    candidates = [
        "ship_date", "출하일자", "shipping_date", "out_date", "date",
        "pack_date", "포장일자",
    ]
    try:
        with conn.cursor() as cur:
            cur.execute("SHOW COLUMNS FROM ShipingList")
            cols = [str(r.get("Field") or "").strip() for r in (cur.fetchall() or [])]
        cols_l = {c.lower(): c for c in cols if c}
        for cand in candidates:
            if cand.lower() in cols_l:
                return cols_l[cand.lower()]
    except Exception:
        return "ship_date"
    return "ship_date"


def fetch_filter_options(ci: "ConnInfo", from_ymd: str, to_ymd: str) -> dict:
    """
    납품처(ship_to), 프로젝트(project) 드롭다운용 DISTINCT 목록 (날짜범위 기반).
    - From~To 범위 내 데이터에 존재하는 값만 제공.
    - 범위 내 데이터가 없으면(빈 목록) 전체 기준으로 fallback (드롭다운 목록만).
    """
    def _clean(v: Any) -> str:
        try:
            s = "" if v is None else str(v)
        except Exception:
            s = ""
        s = s.replace("\u00A0", " ").replace("\u200b", " ")
        s = s.replace("\r", " ").replace("\n", " ").replace("\t", " ")
        s = _re.sub(r"\s+", " ", s).strip()
        return s

    use_range = is_valid_ymd(from_ymd) and is_valid_ymd(to_ymd)
    if use_range:
        try:
            dt_from, dt_to = ymd_to_dt_range(from_ymd, to_ymd)
        except Exception:
            use_range = False
            dt_from, dt_to = "", ""
    else:
        dt_from, dt_to = "", ""

    ship_list: List[str] = []
    proj_list: List[str] = []

    with db_connect(ci) as conn:
        with conn.cursor() as cur:
            def _q_ship(all_scope: bool) -> List[str]:
                if all_scope:
                    cur.execute("SELECT DISTINCT ship_to FROM ShipingList WHERE ship_to IS NOT NULL AND ship_to <> '' ORDER BY ship_to")
                else:
                    cur.execute(
                        "SELECT DISTINCT ship_to FROM ShipingList "
                        "WHERE ship_to IS NOT NULL AND ship_to <> '' "
                        "AND ship_datetime >= %s AND ship_datetime < %s "
                        "ORDER BY ship_to",
                        (dt_from, dt_to),
                    )
                return [_clean(r.get("ship_to")) for r in (cur.fetchall() or []) if _clean(r.get("ship_to"))]

            def _q_proj(all_scope: bool) -> List[str]:
                if all_scope:
                    cur.execute("SELECT DISTINCT project FROM ShipingList WHERE project IS NOT NULL AND project <> '' ORDER BY project")
                else:
                    cur.execute(
                        "SELECT DISTINCT project FROM ShipingList "
                        "WHERE project IS NOT NULL AND project <> '' "
                        "AND ship_datetime >= %s AND ship_datetime < %s "
                        "ORDER BY project",
                        (dt_from, dt_to),
                    )
                return [_clean(r.get("project")) for r in (cur.fetchall() or []) if _clean(r.get("project"))]

            if use_range:
                ship_list = _q_ship(all_scope=False)
                proj_list = _q_proj(all_scope=False)

                # 범위 내 데이터가 없으면 전체 목록으로 fallback (드롭다운 목록만)
                if not ship_list:
                    ship_list = _q_ship(all_scope=True)
                if not proj_list:
                    proj_list = _q_proj(all_scope=True)
            else:
                ship_list = _q_ship(all_scope=True)
                proj_list = _q_proj(all_scope=True)

    # ensure '전체' exists and is first (as first)
    ship_list = ["전체"] + [s for s in ship_list if s and s != "전체"]
    proj_list = ["전체"] + [p for p in proj_list if p and p != "전체"]
    return {"ship_to": ship_list, "project": proj_list}


class Worker(QtCore.QObject):
    finished = QtCore.pyqtSignal(object)
    failed = QtCore.pyqtSignal(str)

    def __init__(self, fn, *args, **kwargs):
        super().__init__()
        self.fn = fn
        self.args = args
        self.kwargs = kwargs

    @QtCore.pyqtSlot()
    def run(self):
        try:
            t0 = _time.perf_counter()
            res = self.fn(*self.args, **self.kwargs)
            if is_dataclass(res):
                res = asdict(res)
            elif isinstance(res, dict):
                res = dict(res)
            else:
                res = {"res": res}
            res["_elapsed_ms"] = int((_time.perf_counter() - t0) * 1000)
            self.finished.emit(res)
        except Exception:
            self.failed.emit(traceback.format_exc())


class _ThreadKeeperMixin:
    def _init_keeper(self):
        self._active_jobs: List[Tuple[QtCore.QThread, Worker]] = []

    def _start_job(self, thread: QtCore.QThread, worker: Worker):
        self._active_jobs.append((thread, worker))

    def _end_job(self, thread: QtCore.QThread):
        self._active_jobs = [(t, w) for (t, w) in self._active_jobs if t is not thread]



def authenticate_account(ci: ConnInfo, user_id: str, user_pw: str) -> Tuple[Optional[dict], str]:
    """
    dp.account 테이블로 로그인 체크
    - status가 'approved'면 통과 (status가 비어있으면 통과)
    - PW는 현재 DB 값과 입력값을 그대로 비교
    """
    user_id = (user_id or "").strip()
    user_pw = (user_pw or "").strip()
    if not user_id:
        return None, "ID를 입력하세요."
    if not user_pw:
        return None, "PW를 입력하세요."

    try:
        with db_connect(ci) as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT `ID`, `NAME`, `PW`, `role`, `status`, `lv` "
                    "FROM `account` WHERE `ID`=%s LIMIT 1",
                    (user_id,),
                )
                row = cur.fetchone()
                if not row:
                    return None, "해당 ID가 존재하지 않습니다."
                db_pw = str(row.get("PW") or "")
                if db_pw != user_pw:
                    return None, "해당 패스워드가 틀렸습니다."
                st = str(row.get("status") or "").strip().lower()
                if st and st != "approved":
                    return None, f"승인되지 않은 계정(status={row.get('status')})."

                # last_login_at 갱신 (없어도 무시)
                try:
                    cur.execute("UPDATE `account` SET `last_login_at`=NOW() WHERE `ID`=%s", (user_id,))
                except Exception:
                    pass

                return {
                    "id": row.get("ID"),
                    "name": row.get("NAME"),
                    "role": row.get("role"),
                    "status": row.get("status"),
                    "lv": row.get("lv"),
                }, ""
    except Exception as e:
        return None, f"DB 오류: {e}"


class LoginDialog(QtWidgets.QDialog):
    def __init__(self, parent=None, ci: Optional[ConnInfo] = None):
        super().__init__(parent)
        self.ci = ci or conn_info_default()
        self.user: Optional[dict] = None
        self.setWindowTitle("로그인")
        self.setFixedSize(420, 210)

        v = QtWidgets.QVBoxLayout(self)
        form = QtWidgets.QGridLayout()
        v.addLayout(form)

        form.addWidget(QtWidgets.QLabel("ID"), 0, 0)
        self.ed_id = QtWidgets.QLineEdit()
        self.ed_id.setPlaceholderText("아이디")
        form.addWidget(self.ed_id, 0, 1)

        form.addWidget(QtWidgets.QLabel("PW"), 1, 0)
        self.ed_pw = QtWidgets.QLineEdit()
        self.ed_pw.setEchoMode(QtWidgets.QLineEdit.Password)
        self.ed_pw.setPlaceholderText("비밀번호")
        form.addWidget(self.ed_pw, 1, 1)

        self.lbl_err = QtWidgets.QLabel("")
        self.lbl_err.setStyleSheet("color:#ff8080;")
        v.addWidget(self.lbl_err)

        btns = QtWidgets.QHBoxLayout()
        v.addLayout(btns)
        btns.addStretch(1)
        self.btn_login = QtWidgets.QPushButton("로그인")
        self.btn_cancel = QtWidgets.QPushButton("취소")
        btns.addWidget(self.btn_login)
        btns.addWidget(self.btn_cancel)

        self.btn_login.clicked.connect(self._do_login)
        self.btn_cancel.clicked.connect(self.reject)
        self.ed_pw.returnPressed.connect(self._do_login)
        self.ed_id.returnPressed.connect(lambda: self.ed_pw.setFocus())

        self.ed_id.setFocus()

    def _do_login(self):
        uid = self.ed_id.text()
        upw = self.ed_pw.text()
        user, err = authenticate_account(self.ci, uid, upw)
        if not user:
            self.lbl_err.setText(err)
            return
        self.user = user
        self.accept()





class SoPojangDialog(QtWidgets.QDialog, _ThreadKeeperMixin):
    def __init__(self, parent, ci: ConnInfo, from_date: str, to_date: str, ship_to: str, project: str, user: Optional[dict] = None):
        super().__init__(parent)
        self._init_keeper()
        self.setWindowTitle("소포장내역")
        self.resize(1750, 900)
        self.ci = ci
        self.user = user
        self.from_date = from_date
        self.to_date = to_date
        self.ship_to = ship_to
        self.project = project

        layout = QtWidgets.QVBoxLayout(self)
        self.lbl_top = QtWidgets.QLabel("불러오는 중...", self)
        self.lbl_top.setObjectName("muted")
        self.lbl_top.setWordWrap(True)
        layout.addWidget(self.lbl_top)

        # 소포장내역(매트릭스)
        self.scroll = QtWidgets.QScrollArea(self)
        self.scroll.setWidgetResizable(True)
        self.scroll.setFrameShape(QtWidgets.QFrame.NoFrame)
        self.scroll.setHorizontalScrollBarPolicy(QtCore.Qt.ScrollBarAsNeeded)
        self.scroll.setVerticalScrollBarPolicy(QtCore.Qt.ScrollBarAsNeeded)

        # ✅ 5종(IR/X/Y/Z/ZS)을 한 줄(가로)로 쭉 배치 (줄바꿈 없음, 가로 스크롤로 처리)
        self.scroll_content = QtWidgets.QWidget(self.scroll)
        self.row = QtWidgets.QHBoxLayout(self.scroll_content)
        self.row.setContentsMargins(0, 0, 0, 0)
        self.row.setSpacing(12)
        try:
            self.row.setAlignment(QtCore.Qt.AlignTop | QtCore.Qt.AlignLeft)
        except Exception:
            pass

        self.scroll.setWidget(self.scroll_content)

        layout.addWidget(self.scroll, 1)

        self._group_tables: List[QtWidgets.QTableWidget] = []

        btns = QtWidgets.QHBoxLayout()
        btn_close = QtWidgets.QPushButton("닫기")
        btn_close.clicked.connect(self.accept)
        btns.addStretch(1)
        btns.addWidget(btn_close)
        layout.addLayout(btns)

        self._run_async()

    def _clear_groups(self):
        # 기존 그룹 위젯/테이블 제거 (단일 가로행)
        def _clear_layout(lay: QtWidgets.QLayout):
            try:
                while lay.count():
                    item = lay.takeAt(0)
                    w = item.widget()
                    if w is not None:
                        w.setParent(None)
                        w.deleteLater()
            except Exception:
                pass
        _clear_layout(self.row)
        self._group_tables = []


    def _run_async(self):
        th = QtCore.QThread(self)
        worker = Worker(query_sopojang_matrix_fast, self.ci, self.from_date, self.to_date, self.ship_to, self.project, DEFAULT_MATRIX_GROUPS)
        worker.moveToThread(th)
        self._start_job(th, worker)
        worker.finished.connect(lambda res: self._apply(res, th))
        worker.failed.connect(lambda trace: self._err(trace, th))
        th.started.connect(worker.run)
        th.start()


    def _err(self, trace, th):
        _log_error('소포장내역', trace)
        if _is_admin(getattr(self, 'user', None)):
            first = (trace.strip().splitlines() or [''])[0]
            self.lbl_top.setText('오류: ' + first)
        else:
            self.lbl_top.setText('오류가 발생했습니다. 관리자에게 문의하세요.')
        self._end_job(th)
        th.quit()
        th.wait()

    def _apply(self, res: Dict[str, Any], th):
        """소포장내역(매트릭스) UI 적용. (UI traceback 노출 금지 / 콘솔만)"""
        try:
            raw_groups = list(res.get("groups") or [])
            from_d = res.get("from") or self.from_date
            to_d = res.get("to") or self.to_date
            ms = int(res.get("_elapsed_ms") or res.get("db_ms") or 0)

            # ✅ 그룹 순서를 항상 DEFAULT_MATRIX_GROUPS(5종) 순서로 고정
            gmap: Dict[str, Dict[str, Any]] = {}
            for g in raw_groups:
                k0 = normalize_part_name(g.get("part_name") or "")
                k = MATRIX_PART_ALIAS.get(k0, k0)
                if k:
                    gmap[k] = g

            ordered: List[Dict[str, Any]] = []
            for part, cav_def in DEFAULT_MATRIX_GROUPS:
                k0 = normalize_part_name(part)
                k = MATRIX_PART_ALIAS.get(k0, k0)
                g = gmap.get(k)
                if not g:
                    g = {"part_name": part, "cavs": list(cav_def), "rows": [], "total_qty": 0}
                else:
                    g["part_name"] = part  # 표시용 품번명은 5종 기준으로 고정
                    if not g.get("cavs"):
                        g["cavs"] = list(cav_def)
                ordered.append(g)

            # 상단 요약
            parts = []
            for g in ordered:
                pn = str(g.get("part_name") or "")
                total = int(g.get("total_qty") or 0)
                parts.append(f"{pn} : {fmt_ea(total)}")
            self.lbl_top.setText(f"기간: {from_d} ~ {to_d}   (DB {ms} ms)\n" + " | ".join(parts))

            # 그룹 UI 리빌드
            self._clear_groups()

            col_true_fill = QtGui.QColor("#C6EFCE")
            col_true_text = QtGui.QColor("#006100")
            col_false_fill = QtGui.QColor("#FFC7CE")
            col_false_text = QtGui.QColor("#9C0006")

            def _uniq_sorted(rows, cav_header):
                # Excel: =SORT(UNIQUE(TOCOL(B2:C200,3,TRUE)))
                #  - 날짜(prod_date) + 차수(REV)를 flatten한 뒤 UNIQUE+SORT
                #  - 여기서는 날짜는 그대로 나열하고, 차수는 "차수 1,2,3,4 /Cav" 형태로 Cav 포함 표시
                dates = set()
                revs = set()
                rev_to_cavs: Dict[str, set] = {}

                for rr in rows:
                    rv = str(rr.get("rev") or "").strip()
                    dd = str(rr.get("date") or "").strip()
                    if dd:
                        dates.add(dd)
                    if rv:
                        revs.add(rv)
                        cav_map = rr.get("cavs") or {}
                        for c in (cav_header or []):
                            try:
                                ci = int(c)
                            except Exception:
                                continue
                            if bool(cav_map.get(ci, False)):
                                rev_to_cavs.setdefault(rv, set()).add(ci)

                out: List[str] = []
                for d in sorted(dates):
                    out.append(d)
                for rv in sorted(revs):
                    cav_list = sorted(rev_to_cavs.get(rv, set()))
                    cav_str = ",".join(str(x) for x in cav_list) if cav_list else "-"
                    # 고정폭(모노스페이스) + 패딩으로 줄맞춤
                    out.append(f"{rv:<2} {cav_str:<11} /Cav")
                return out

            try:
                self.scroll_content.setUpdatesEnabled(False)
            except Exception:
                pass

            for idx, g in enumerate(ordered):
                pn = str(g.get("part_name") or "")
                total = int(g.get("total_qty") or 0)
                rows = list(g.get("rows") or [])

                # Cav 헤더: 데이터 기반(1~4 / 5~8) 4칸 고정
                cavs = []
                for c in (g.get("cavs") or []):
                    try:
                        cavs.append(int(c))
                    except Exception:
                        continue
                if not cavs:
                    cavs = [1, 2, 3, 4]
                if max(cavs) >= 5:
                    cavs = [5, 6, 7, 8]
                else:
                    cavs = [1, 2, 3, 4]

                gb = QtWidgets.QGroupBox(f"{pn}")
                vb = QtWidgets.QVBoxLayout(gb)
                vb.setContentsMargins(10, 10, 10, 10)
                vb.setSpacing(8)

                # 메인 테이블
                tbl = QtWidgets.QTableWidget(gb)
                tbl.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
                _enable_copyable_table(tbl)
                # focus handled by _enable_copyable_table
                tbl.setAlternatingRowColors(True)
                tbl.setSortingEnabled(False)
                tbl.horizontalHeader().setStretchLastSection(False)
                tbl.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.ResizeToContents)
                hide_row_numbers(tbl)

                headers = ["품번명", "차수", "날짜"] + [f"Cav{c}" for c in cavs] + ["수량"]
                tbl.setColumnCount(len(headers))
                tbl.setHorizontalHeaderLabels(headers)
                try:
                    tbl.horizontalHeader().setDefaultAlignment(QtCore.Qt.AlignCenter)
                except Exception:
                    pass
                tbl.setRowCount(max(len(rows), 1))

                if not rows:
                    muted = QtGui.QBrush(QtGui.QColor("#a0a0a0"))
                    it0 = QtWidgets.QTableWidgetItem(pn)
                    it0.setTextAlignment(QtCore.Qt.AlignCenter)
                    it1 = QtWidgets.QTableWidgetItem("-")
                    it1.setTextAlignment(QtCore.Qt.AlignCenter)
                    it2 = QtWidgets.QTableWidgetItem("-")
                    it2.setTextAlignment(QtCore.Qt.AlignCenter)
                    it0.setForeground(muted); it1.setForeground(muted); it2.setForeground(muted)
                    tbl.setItem(0, 0, it0)
                    tbl.setItem(0, 1, it1)
                    tbl.setItem(0, 2, it2)
                    for i, _c in enumerate(cavs):
                        it = QtWidgets.QTableWidgetItem("-")
                        it.setTextAlignment(QtCore.Qt.AlignCenter)
                        it.setForeground(muted)
                        tbl.setItem(0, 3 + i, it)
                    itq = QtWidgets.QTableWidgetItem(fmt_ea(0))
                    itq.setTextAlignment(QtCore.Qt.AlignCenter)
                    itq.setForeground(muted)
                    tbl.setItem(0, 3 + len(cavs), itq)
                else:
                    for r, row in enumerate(rows):
                        part_name = str(row.get("part_name") or pn or "")
                        it = QtWidgets.QTableWidgetItem(part_name)
                        it.setTextAlignment(QtCore.Qt.AlignCenter)
                        tbl.setItem(r, 0, it)

                        it = QtWidgets.QTableWidgetItem(str(row.get("rev") or ""))
                        it.setTextAlignment(QtCore.Qt.AlignCenter)
                        tbl.setItem(r, 1, it)

                        # ✅ 날짜: 생산일자(prod_date) 기준으로 query에서 내려온 값 사용
                        it = QtWidgets.QTableWidgetItem(str(row.get("date") or ""))
                        it.setTextAlignment(QtCore.Qt.AlignCenter)
                        tbl.setItem(r, 2, it)

                        cav_map = row.get("cavs") or {}
                        for i, cav_no in enumerate(cavs):
                            val = bool(cav_map.get(int(cav_no), False))
                            it = QtWidgets.QTableWidgetItem("TRUE" if val else "FALSE")
                            it.setTextAlignment(QtCore.Qt.AlignCenter)
                            it.setForeground(QtGui.QBrush(col_true_text if val else col_false_text))
                            it.setData(QtCore.Qt.BackgroundRole, col_true_fill if val else col_false_fill)
                            tbl.setItem(r, 3 + i, it)

                        it = QtWidgets.QTableWidgetItem(fmt_ea(row.get("qty") or 0))
                        it.setTextAlignment(QtCore.Qt.AlignCenter)
                        tbl.setItem(r, 3 + len(cavs), it)

                # 표 크기를 내용에 맞춰 고정 (수직 스크롤은 전체 스크롤로)
                tbl.resizeColumnsToContents()
                tbl.resizeRowsToContents()
                header_h = tbl.horizontalHeader().height()
                rows_h = sum(tbl.rowHeight(i) for i in range(tbl.rowCount()))
                tbl_h = header_h + rows_h + (tbl.frameWidth() * 2) + 10
                tbl_w = sum(tbl.columnWidth(i) for i in range(tbl.columnCount()))
                tbl_w += tbl.verticalHeader().width() + (tbl.frameWidth() * 2) + 14

                tbl.setSizePolicy(QtWidgets.QSizePolicy.Fixed, QtWidgets.QSizePolicy.Fixed)
                tbl.setFixedSize(tbl_w, tbl_h)
                tbl.setVerticalScrollBarPolicy(QtCore.Qt.ScrollBarAlwaysOff)

                # 수량 옆 UNIQUE 리스트 (표 높이와 동일하게 맞춰서 "빈공간/떠다님" 방지)
                uniq_items = _uniq_sorted(rows, cavs)

                roww = QtWidgets.QWidget(gb)
                hb = QtWidgets.QHBoxLayout(roww)
                hb.setContentsMargins(0, 0, 0, 0)
                hb.setSpacing(8)
                try:
                    hb.setAlignment(QtCore.Qt.AlignTop | QtCore.Qt.AlignLeft)
                except Exception:
                    pass

                hb.addWidget(tbl, 0, QtCore.Qt.AlignTop)

                side = QtWidgets.QWidget(gb)
                sv = QtWidgets.QVBoxLayout(side)
                sv.setContentsMargins(0, 0, 0, 0)
                sv.setSpacing(6)

                lbl_total = QtWidgets.QLabel(fmt_ea(total), side)
                lbl_total.setAlignment(QtCore.Qt.AlignCenter)
                lbl_total.setObjectName("matrixTotal")
                lbl_total.setFixedHeight(28)
                sv.addWidget(lbl_total)

                lw = QtWidgets.QListWidget(side)
                lw.setObjectName("matrixUnique")
                _enable_copyable_list(lw)
                # focus handled by _enable_copyable_list
                try:
                    lw.setFont(monospace_font())
                except Exception:
                    pass
                lw.addItems(uniq_items)
                lw.setVerticalScrollBarPolicy(QtCore.Qt.ScrollBarAsNeeded)
                sv.addWidget(lw, 1)

                side_w = 160
                side.setFixedWidth(side_w)
                side.setFixedHeight(tbl_h)
                # 리스트 높이도 고정해서 sizeHint 폭주 방지
                lw.setFixedHeight(max(60, tbl_h - lbl_total.height() - 8))

                hb.addWidget(side, 0, QtCore.Qt.AlignTop)
                roww.setFixedHeight(tbl_h)

                vb.addWidget(roww)

                # 블록 고정폭(=표+우측리스트)로 3+2 배치가 흐트러지지 않게
                gb_w = tbl_w + side_w + 8 + 26
                gb.setSizePolicy(QtWidgets.QSizePolicy.Fixed, QtWidgets.QSizePolicy.Fixed)
                gb.setFixedWidth(gb_w)
                gb.setFixedHeight(roww.height() + 42)  # title/margins 여유

                self._group_tables.append(tbl)

                # ✅ 5종을 한 줄(가로)로 배치
                self.row.addWidget(gb, 0, QtCore.Qt.AlignTop)

            # 오른쪽 여백은 stretch로 처리 (블록은 좌측 정렬)
            self.row.addStretch(1)

        except Exception:
            self._err(traceback.format_exc(), th)
            return
        finally:
            try:
                self.scroll_content.setUpdatesEnabled(True)
            except Exception:
                pass
            self._end_job(th)
            th.quit()
            th.wait()



class MainWindow(QtWidgets.QMainWindow, _ThreadKeeperMixin):
    def __init__(self, user: Optional[dict] = None):
        super().__init__()
        self._threads = []  # QThread keep-alive
        self._init_keeper()
        self.setWindowTitle("출하내역")
        self.resize(1700, 920)

        self.last_result: Optional[dict] = None
        self.ci = conn_info_default()

        self.user = user
        self._apply_user_title()


        central = QtWidgets.QWidget(self)
        self.setCentralWidget(central)
        root = QtWidgets.QHBoxLayout(central)

        left = QtWidgets.QVBoxLayout()
        root.addLayout(left, 1)
        right = QtWidgets.QVBoxLayout()
        root.addLayout(right, 0)

        g_f = QtWidgets.QGroupBox("필터")
        left.addWidget(g_f)
        fl = QtWidgets.QGridLayout(g_f)

        self.dt_from = QtWidgets.QDateEdit(self)
        self.dt_to = QtWidgets.QDateEdit(self)
        self.dt_from.setCalendarPopup(True)
        self.dt_to.setCalendarPopup(True)
        self.dt_from.setDisplayFormat("yyyy-MM-dd")
        self.dt_to.setDisplayFormat("yyyy-MM-dd")
        today = QtCore.QDate.currentDate()
        self.dt_from.setDate(today)
        self.dt_to.setDate(today)

        # 날짜 변경 시 드롭다운(납품처/프로젝트) 목록을 날짜범위 기반으로 갱신 (디바운스)
        self._dropdown_reload_timer = QtCore.QTimer(self)
        self._dropdown_reload_timer.setSingleShot(True)
        self._dropdown_reload_timer.setInterval(250)
        self.dt_from.dateChanged.connect(lambda _=None: self._dropdown_reload_timer.start())
        self.dt_to.dateChanged.connect(lambda _=None: self._dropdown_reload_timer.start())
        self._dropdown_reload_timer.timeout.connect(self._load_dropdowns_async)

        self.cb_shipto = QtWidgets.QComboBox()
        self.cb_shipto.setEditable(False)
        self.cb_shipto.setMinimumWidth(180)
        self.cb_shipto.clear()
        self.cb_shipto.setCurrentIndex(-1)
        self.ed_packbc = QtWidgets.QLineEdit()
        self.ed_small = QtWidgets.QLineEdit()
        self.ed_tray = QtWidgets.QLineEdit()
        self.cb_project = QtWidgets.QComboBox()
        self.cb_project.setEditable(False)
        self.cb_project.setMinimumWidth(180)
        self.cb_project.clear()
        self.cb_project.setCurrentIndex(-1)
        # 시작 시에는 전체 목록(범위제한 없음) 기반으로 선택할 수 있게 해둠
        self.cb_shipto.addItem("전체")
        self.cb_project.addItem("전체")

        self.cb_shipto.setEnabled(True)
        self.cb_project.setEnabled(True)
        self.ed_pname = QtWidgets.QLineEdit()

        fl.addWidget(QtWidgets.QLabel("From"), 0, 0); fl.addWidget(self.dt_from, 0, 1)
        fl.addWidget(QtWidgets.QLabel("To"),   0, 2); fl.addWidget(self.dt_to,   0, 3)
        fl.addWidget(QtWidgets.QLabel("납품처"), 0, 4); fl.addWidget(self.cb_shipto, 0, 5)
        fl.addWidget(QtWidgets.QLabel("포장바코드"), 1, 0); fl.addWidget(self.ed_packbc, 1, 1)
        fl.addWidget(QtWidgets.QLabel("소포장NO"), 1, 2); fl.addWidget(self.ed_small, 1, 3)
        fl.addWidget(QtWidgets.QLabel("Tray NO"), 1, 4); fl.addWidget(self.ed_tray, 1, 5)
        fl.addWidget(QtWidgets.QLabel("품번명"), 2, 0); fl.addWidget(self.ed_pname, 2, 1, 1, 3)
        fl.addWidget(QtWidgets.QLabel("프로젝트"), 2, 4); fl.addWidget(self.cb_project, 2, 5)

        btn_wrap = QtWidgets.QHBoxLayout()
        self.btn_search = QtWidgets.QPushButton("조회")
        self.btn_search.clicked.connect(self.on_search)
        self.btn_sopojang = QtWidgets.QPushButton("소포장내역")
        self.btn_sopojang.clicked.connect(self.on_open_sopojang)
        self.btn_sp_export = QtWidgets.QPushButton("특정치수 엑셀출력")
        self.btn_sp_export.clicked.connect(self.on_export_sp_data)
        btn_wrap.addWidget(self.btn_search)
        btn_wrap.addWidget(self.btn_sopojang)
        btn_wrap.addWidget(self.btn_sp_export)
        btn_wrap.addStretch(1)
        fl.addLayout(btn_wrap, 3, 0, 1, 2)

        self.lbl_status = QtWidgets.QLabel("")
        self.lbl_status.setObjectName("muted")
        fl.addWidget(self.lbl_status, 3, 2, 1, 4)

        self.table = QtWidgets.QTableWidget(self)
        self.table.setColumnCount(len(COLUMNS))
        self.table.setHorizontalHeaderLabels([t for _, t in COLUMNS])
        self.table.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
        self.table.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
        self.table.setAlternatingRowColors(True)
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.Interactive)
        try:
            self.table.horizontalHeader().setDefaultAlignment(QtCore.Qt.AlignCenter)
        except Exception:
            pass
        hide_row_numbers(self.table)
        left.addWidget(self.table, 1)

        g_sum = QtWidgets.QGroupBox("모델별 출고수량 합계")
        right.addWidget(g_sum, 1)
        sl = QtWidgets.QVBoxLayout(g_sum)

        self.sum_table = QtWidgets.QTableWidget(self)
        self.sum_table.setColumnCount(2)
        self.sum_table.setHorizontalHeaderLabels(["모델", "합계"])
        self.sum_table.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
        self.sum_table.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
        self.sum_table.setAlternatingRowColors(True)
        self.sum_table.horizontalHeader().setStretchLastSection(True)
        self.sum_table.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.Interactive)
        try:
            self.sum_table.horizontalHeader().setDefaultAlignment(QtCore.Qt.AlignCenter)
        except Exception:
            pass
        hide_row_numbers(self.sum_table)
        sl.addWidget(self.sum_table, 1)

        self.status = self.statusBar()
        self._test_connection()
        self._load_dropdowns_initial_sync()
        # 앱 시작 시 '전체 목록'을 바로 표시 (로딩 화면 없이 백그라운드 조회)
        try:
            QtCore.QTimer.singleShot(0, self.on_search)
        except Exception:
            pass
        # 출하 알림(토스트): 하루 1회(오늘 출하가 있으면) - 옵션에서 오늘 출하 상세 보기
        self._toast = None
        self._shiptoast_last_summary = None
        try:
            # window icon + tray (optional)
            self._init_window_icon()
            self._init_tray()
            self._shiptoast_settings = QtCore.QSettings('dp', 'ShipingListViewer')
        except Exception:
            self._shiptoast_settings = None
        self._shiptoast_timer = QtCore.QTimer(self)
        self._shiptoast_timer.setInterval(10 * 60 * 1000)  # 10 min
        self._shiptoast_timer.timeout.connect(self._check_daily_ship_toast)
        self._shiptoast_timer.start()
        try:
            QtCore.QTimer.singleShot(2500, self._check_daily_ship_toast)
        except Exception:
            pass


    def _test_connection(self):
        try:
            with db_connect(self.ci) as conn:
                with conn.cursor() as cur:
                    cur.execute("SELECT 1 AS ok")
                    _ = cur.fetchone()
            self.status.showMessage(f"DB OK  |  {self.ci.host}:{self.ci.port}/{self.ci.db} (user={self.ci.user})", 12000)
        except Exception as e:
            self.status.showMessage("DB 연결 실패: " + str(e), 15000)

    def _apply_user_title(self):
        if getattr(self, "user", None) and isinstance(self.user, dict) and self.user.get("id"):
            nm = self.user.get("name") or ""
            uid = self.user.get("id") or ""
            role = self.user.get("role") or ""
            tail = f"{nm} ({uid})" if nm else f"{uid}"
            if role:
                tail += f" / {role}"
            self.setWindowTitle(f"출하내역 - {tail}")
        else:
            self.setWindowTitle("출하내역")


    def _range_from_ymd(self) -> str:
        try:
            return self.dt_from.date().toString("yyyy-MM-dd")
        except Exception:
            return ""

    def _range_to_ymd(self) -> str:
        try:
            return self.dt_to.date().toString("yyyy-MM-dd")
        except Exception:
            return ""

    def _get_required_ship_project(self):
        """납품처/프로젝트 현재 선택값을 (ship_to, project)로 반환.
        - 비어있으면 '전체'처럼 필터 미적용(=빈 문자열)로 처리.
        """
        ship_to = (self.cb_shipto.currentText() if hasattr(self, "cb_shipto") else "").strip()
        project = (self.cb_project.currentText() if hasattr(self, "cb_project") else "").strip()

        if ship_to == "":
            ship_to = ""
        if project == "":
            project = ""

        # '전체'는 필터 미적용
        if ship_to == "전체":
            ship_to = ""
        if project == "전체":
            project = ""
        return ship_to, project

    def _set_status(self, text: str):
        try:
            self.lbl_status.setText(text)
        except Exception:
            pass

    def _apply_dropdown_items(self, shiptos, projects, keep_shipto=None, keep_project=None):
        # keep current selection if possible
        cur_ship = (keep_shipto if keep_shipto is not None else (self.cb_shipto.currentText() if hasattr(self, "cb_shipto") else "")).strip()
        cur_proj = (keep_project if keep_project is not None else (self.cb_project.currentText() if hasattr(self, "cb_project") else "")).strip()

        def _build(items):
            # addItems에 문자열이 들어가면 글자 단위로 쪼개져 들어가므로 방어
            if items is None:
                items = []
            elif isinstance(items, str):
                items = [items]
            elif isinstance(items, dict):
                items = list(items)
            cleaned = []
            seen = set()
            for x in items:
                s = str(x).strip()
                if not s:
                    continue
                if s in seen:
                    continue
                seen.add(s)
                cleaned.append(s)

            # ensure '전체' exists and is first
            if "전체" in seen:
                cleaned = [x for x in cleaned if x != "전체"]
            cleaned.insert(0, "전체")
            return cleaned

        ship_list = _build(shiptos)
        proj_list = _build(projects)

        def _apply(cb: QtWidgets.QComboBox, items, keep_value: str):
            cb.blockSignals(True)
            cb.clear()
            cb.addItems(items)
            cb.setEnabled(True)

            # 기본 선택 규칙:
            # 1) 이전 선택(keep_value)이 새 목록에 있으면 유지
            # 2) 없으면 '전체'가 아닌 첫 항목을 자동 선택
            # 3) 그마저도 없으면 첫 항목(보통 '전체')
            idx = 0
            if keep_value and keep_value in items:
                idx = items.index(keep_value)
            else:
                for i, v in enumerate(items):
                    if str(v).strip() and str(v).strip() != "전체":
                        idx = i
                        break
            cb.setCurrentIndex(idx)
            cb.blockSignals(False)

        _apply(self.cb_shipto, ship_list, cur_ship)
        _apply(self.cb_project, proj_list, cur_proj)

    def _load_dropdowns_initial_sync(self):
        # On first open: From~To 범위 기준으로 shipto/project 로드 (없으면 전체로 fallback)
        try:
            from_ymd = self._range_from_ymd()
            to_ymd = self._range_to_ymd()
        except Exception:
            from_ymd, to_ymd = "", ""
        try:
            opts = fetch_filter_options(self.ci, from_ymd, to_ymd)
            shiptos = opts.get("ship_to", []) if isinstance(opts, dict) else []
            projects = opts.get("project", []) if isinstance(opts, dict) else []
            # 기본은 '전체' 유지
            self._apply_dropdown_items(shiptos, projects, keep_shipto="전체", keep_project="전체")
            self._set_status("납품처/프로젝트 목록 로드됨")
        except Exception as e:
            log_to_console("납품처/프로젝트 목록 로드 실패: " + str(e))
            self._set_status("납품처/프로젝트 목록 로드 실패 (콘솔 확인)")

    def _drop_thread(self, th):
        try:
            self._threads = [(t, w) for (t, w) in self._threads if t is not th]
        except Exception:
            pass


    def _load_dropdowns_async(self):
        # After From/To change: refresh shipto/project list by date range
        try:
            from_ymd = self._range_from_ymd()
            to_ymd = self._range_to_ymd()
        except Exception:
            from_ymd, to_ymd = "", ""

        self._set_status("납품처/프로젝트 목록 갱신 중...")

        th = QtCore.QThread()
        worker = Worker(fetch_filter_options, self.ci, from_ymd, to_ymd)

        def _ok(res, th=th):
            try:
                shiptos, projects = [], []
                if isinstance(res, dict):
                    shiptos = res.get("ship_to", []) or []
                    projects = res.get("project", []) or []
                elif isinstance(res, (list, tuple)) and len(res) == 2:
                    shiptos, projects = res

                # 선택 범위에 데이터가 전혀 없으면, 초기(전체) 목록으로 fallback
                if (not shiptos) and (not projects):
                    try:
                        if is_valid_ymd(self._range_from_ymd()) and is_valid_ymd(self._range_to_ymd()):
                            opts_all = fetch_filter_options(self.ci, "", "")
                            if isinstance(opts_all, dict):
                                shiptos = opts_all.get("ship_to", []) or []
                                projects = opts_all.get("project", []) or []
                            elif isinstance(opts_all, (list, tuple)) and len(opts_all) == 2:
                                shiptos, projects = opts_all
                            self._set_status("선택 범위에 데이터 없음 - 목록 전체 표시")
                        else:
                            self._set_status("납품처/프로젝트 목록 갱신됨")
                    except Exception:
                        self._set_status("납품처/프로젝트 목록 갱신됨")
                else:
                    self._set_status("납품처/프로젝트 목록 갱신됨")

                self._apply_dropdown_items(shiptos, projects)
            except Exception as e:
                log_to_console("납품처/프로젝트 목록 적용 실패: " + str(e))
                self._set_status("납품처/프로젝트 목록 적용 실패 (콘솔 확인)")
            finally:
                try:
                    th.quit()
                except Exception:
                    pass

        def _fail(err, th=th):
            log_to_console("납품처/프로젝트 목록 갱신 실패: " + str(err))
            self._set_status("납품처/프로젝트 목록 갱신 실패 (콘솔 확인)")
            try:
                th.quit()
            except Exception:
                pass

        worker.moveToThread(th)
        th.started.connect(worker.run)
        worker.finished.connect(_ok)
        worker.failed.connect(_fail)
        worker.finished.connect(worker.deleteLater)
        worker.failed.connect(worker.deleteLater)
        th.finished.connect(th.deleteLater)
        self._threads.append((th, worker))
        th.start()

    def on_search(self):
        from_d = self.dt_from.date().toString("yyyy-MM-dd")
        to_d = self.dt_to.date().toString("yyyy-MM-dd")
        ship_to, project = self._get_required_ship_project()
        if ship_to is None:
            return
        pack_bc = self.ed_packbc.text().strip()
        small_no = self.ed_small.text().strip()
        tray_no = self.ed_tray.text().strip()
        pname = self.ed_pname.text().strip()

        self.lbl_status.setText("조회 중...")
        self.btn_search.setEnabled(False)
        self.btn_sopojang.setEnabled(False)
        try:
            self.btn_sp_export.setEnabled(False)
        except Exception:
            pass

        th = QtCore.QThread(self)
        worker = Worker(query_all, self.ci, from_d, to_d, ship_to, project, pack_bc, small_no, tray_no, pname)
        worker.moveToThread(th)
        self._start_job(th, worker)

        worker.finished.connect(lambda res: self._apply(res, th))
        worker.failed.connect(lambda trace: self._err(trace, th))
        th.started.connect(worker.run)
        th.start()

    def on_export_sp_data(self):
        """특정치수 Sp_Data_YYYYMMDD.xlsx 생성 (AOI 방식)."""
        from_d = self.dt_from.date().toString("yyyy-MM-dd")
        to_d = self.dt_to.date().toString("yyyy-MM-dd")
        ship_to, project = self._get_required_ship_project()
        if ship_to is None:
            return

        ymd = ""
        try:
            if valid_ymd(to_d):
                ymd = to_d.replace("-", "")
        except Exception:
            ymd = ""
        if not ymd:
            ymd = _today_yyyymmdd()

        save_path = os.path.join(os.getcwd(), f"Sp_Data_{ymd}.xlsx")

        # UI disable
        self.lbl_status.setText("Sp_Data 엑셀 생성 중...")
        self.btn_search.setEnabled(False)
        self.btn_sopojang.setEnabled(False)
        try:
            self.btn_sp_export.setEnabled(False)
        except Exception:
            pass

        th = QtCore.QThread(self)
        worker = Worker(export_sp_data_xlsx, self.ci, from_d, to_d, ship_to, project, save_path)
        worker.moveToThread(th)
        self._start_job(th, worker)

        worker.finished.connect(lambda res: self._apply_sp_export(res, th, save_path))
        worker.failed.connect(lambda trace: self._err_sp_export(trace, th))
        th.started.connect(worker.run)
        th.start()

    def _apply_sp_export(self, res, th, save_path: str):
        ok = int(res.get("ok") or 0)
        if ok:
            sheets = int(res.get("sheets") or 0)
            self.lbl_status.setText(f"Sp_Data 엑셀 완료: {sheets} sheet | {os.path.basename(save_path)}")
            self.status.showMessage(f"Sp_Data 저장 완료: {save_path}", 15000)
        else:
            msg = str(res.get("msg") or "Sp_Data 생성 실패")
            self.lbl_status.setText(msg)
            self.status.showMessage(msg, 12000)

        self.btn_search.setEnabled(True)
        self.btn_sopojang.setEnabled(True)
        try:
            self.btn_sp_export.setEnabled(True)
        except Exception:
            pass
        self._end_job(th)
        th.quit(); th.wait()

    def _err_sp_export(self, trace, th):
        _log_error("Sp_Data", trace)
        if _is_admin(getattr(self, 'user', None)):
            first = (trace.strip().splitlines() or [''])[0]
            self.lbl_status.setText('오류: ' + first)
            self.status.showMessage('오류: ' + first, 15000)
        else:
            self.lbl_status.setText('오류가 발생했습니다. 관리자에게 문의하세요.')
            self.status.showMessage('오류가 발생했습니다.', 8000)

        self.btn_search.setEnabled(True)
        self.btn_sopojang.setEnabled(True)
        try:
            self.btn_sp_export.setEnabled(True)
        except Exception:
            pass
        self._end_job(th)
        th.quit(); th.wait()

    def _err(self, trace, th):
        _log_error('조회', trace)
        if _is_admin(getattr(self, 'user', None)):
            first = (trace.strip().splitlines() or [''])[0]
            self.lbl_status.setText('오류: ' + first)
            self.status.showMessage('오류: ' + first, 15000)
        else:
            self.lbl_status.setText('오류가 발생했습니다. 관리자에게 문의하세요.')
            self.status.showMessage('오류가 발생했습니다.', 8000)
        self.btn_search.setEnabled(True)
        self.btn_sopojang.setEnabled(True)
        try:
            self.btn_sp_export.setEnabled(True)
        except Exception:
            pass
        self._end_job(th)
        th.quit()
        th.wait()

    def _apply(self, res, th):
        rows = res.get("rows") or []
        total = int(res.get("total") or 0)
        did_fallback = bool(res.get("did_fallback"))
        ms = int(res.get("_elapsed_ms") or 0)

        t_render0 = _time.perf_counter()

        self.table.setUpdatesEnabled(False)
        self.table.setRowCount(len(rows))
        for r, row in enumerate(rows):
            for c, (k, _) in enumerate(COLUMNS):
                v = row.get(k, "")
                item = QtWidgets.QTableWidgetItem("" if v is None else str(v))
                item.setTextAlignment(QtCore.Qt.AlignCenter)
                if k in POPUP_COLS:
                    item.setForeground(QtGui.QBrush(QtGui.QColor("#9bdcff")))
                self.table.setItem(r, c, item)
        self.table.setUpdatesEnabled(True)

        sum_rows = res.get("sum_rows") or []
        self.sum_table.setUpdatesEnabled(False)
        self.sum_table.setRowCount(len(sum_rows))
        for r, srow in enumerate(sum_rows):
            itn = QtWidgets.QTableWidgetItem(str(srow.get("part_name") or ""))
            itn.setTextAlignment(QtCore.Qt.AlignCenter)
            self.sum_table.setItem(r, 0, itn)
            itq = QtWidgets.QTableWidgetItem(fmt_ea(srow.get("sum_qty") or 0))
            itq.setTextAlignment(QtCore.Qt.AlignCenter)
            self.sum_table.setItem(r, 1, itq)
        self.sum_table.setUpdatesEnabled(True)

        render_ms = int((_time.perf_counter() - t_render0) * 1000)

        msg = f" 상태 : {total}건  |  DB {ms} ms  |  UI {render_ms} ms"
        self.lbl_status.setText(msg)

        self.btn_search.setEnabled(True)
        self.btn_sopojang.setEnabled(True)
        try:
            self.btn_sp_export.setEnabled(True)
        except Exception:
            pass

        self._end_job(th)
        th.quit()
        th.wait()

    def on_open_sopojang(self):
        from_d = self.dt_from.date().toString("yyyy-MM-dd")
        to_d = self.dt_to.date().toString("yyyy-MM-dd")
        ship_to, project = self._get_required_ship_project()
        if ship_to is None:
            return
        dlg = SoPojangDialog(self, self.ci, from_d, to_d, ship_to, project)
        dlg.exec_()



    

    # ===== System tray / window icon =====
    def _init_window_icon(self):
        try:
            p = _viewer_app_icon_path()
            if p:
                self.setWindowIcon(QtGui.QIcon(p))
        except Exception:
            pass

    def _init_tray(self):
        self._tray = None
        self._tray_quit = False
        try:
            if not QtWidgets.QSystemTrayIcon.isSystemTrayAvailable():
                return
            icon = self.windowIcon()
            if icon is None or icon.isNull():
                p = _viewer_app_icon_path()
                if p:
                    icon = QtGui.QIcon(p)
            self._tray = QtWidgets.QSystemTrayIcon(icon, self)
            self._tray.setToolTip("출하내역")

            menu = QtWidgets.QMenu(self)

            # (timesheet tray menu compatible)
            act_show = QtWidgets.QAction("열기", self)
            act_show.triggered.connect(self.show_from_tray)
            menu.addAction(act_show)

            act_hide = QtWidgets.QAction("숨기기", self)
            act_hide.triggered.connect(self.hide_to_tray)
            menu.addAction(act_hide)

            act_test = QtWidgets.QAction("테스트 알림", self)
            act_test.triggered.connect(self._debug_test_toast)
            menu.addAction(act_test)

            menu.addSeparator()

            act_exit = QtWidgets.QAction("종료", self)
            act_exit.triggered.connect(self._quit_from_tray)
            menu.addAction(act_exit)

            self._tray.setContextMenu(menu)
            self._tray.activated.connect(self._on_tray_activated)
            self._tray.show()
        except Exception:
            # Do not crash app; tray is optional
            self._tray = None

    def _on_tray_activated(self, reason):
        try:
            if reason in (QtWidgets.QSystemTrayIcon.Trigger, QtWidgets.QSystemTrayIcon.DoubleClick):
                if self.isVisible():
                    self.hide()
                else:
                    self._bring_to_front()
        except Exception:
            pass

    def _quit_from_tray(self):
        try:
            self._tray_quit = True
            if getattr(self, "_tray", None):
                try:
                    self._tray.hide()
                except Exception:
                    pass
            QtWidgets.QApplication.quit()
        except Exception:
            QtWidgets.QApplication.quit()

    def _open_today_shipments(self):
        try:
            today = QtCore.QDate.currentDate()
            try:
                self.dt_from.setDate(today)
                self.dt_to.setDate(today)
            except Exception:
                pass
            self._bring_to_front()
            try:
                self.on_search()
            except Exception:
                pass
        except Exception:
            pass

    def closeEvent(self, event):
        # close button -> tray minimize (do NOT quit) unless user chose Quit
        try:
            if getattr(self, "_tray_quit", False):
                event.accept()
                return
            if getattr(self, "_tray", None) and QtWidgets.QSystemTrayIcon.isSystemTrayAvailable():
                event.ignore()
                self.hide()
                try:
                    self._tray.showMessage("출하내역", "트레이로 이동했습니다.", QtWidgets.QSystemTrayIcon.Information, 2000)
                except Exception:
                    pass
                return
        except Exception:
            pass
        event.accept()

# ===== Daily shipment toast =====
    def _bring_to_front(self):
        try:
            self.show()
            self.raise_()
            self.activateWindow()
        except Exception:
            pass

    def hide_to_tray(self):
        try:
            self.hide()
        except Exception:
            pass

    def show_from_tray(self):
        try:
            self._bring_to_front()
        except Exception:
            pass

    def _debug_test_toast(self):
        """트레이 메뉴에서 토스트가 정상 동작하는지 빠르게 확인."""
        try:
            ymd = _dt.date.today().strftime("%Y-%m-%d")
            msg = f"(TEST) {ymd}\n출하등록이 되었습니다."

            # 가능하면 오늘 출하 요약을 같이 넣음(없으면 기본 메시지)
            try:
                res = query_ship_summary_for_day(self.ci, ymd)
                rows = list(res.get("sum_rows") or [])
                total_qty = int(res.get("total_qty") or 0)
                if rows and total_qty > 0:
                    try:
                        top = sorted(rows, key=lambda r: int(r.get("sum_qty") or 0), reverse=True)[:6]
                    except Exception:
                        top = rows[:6]
                    parts = []
                    for r in top:
                        pn = str(r.get("part_name") or "")
                        if not pn:
                            continue
                        try:
                            q = int(r.get("sum_qty") or 0)
                        except Exception:
                            q = 0
                        parts.append(f"{pn} {q:,}EA")
                    line = " | ".join(parts)
                    if len(line) > 180:
                        line = line[:180] + "…"
                    msg = f"(TEST) 출하등록이 되었습니다.\n{line}"
            except Exception:
                pass

            self._show_ship_toast("테스트 알림", msg, payload=ymd)
        except Exception:
            print(traceback.format_exc())

    def _check_daily_ship_toast(self):
        """하루 1회: 오늘 출하가 있으면 '출하등록이 되었습니다' 토스트."""
        try:
            ymd = _dt.date.today().strftime("%Y-%m-%d")
            if self._shiptoast_settings is not None:
                try:
                    last = str(self._shiptoast_settings.value("shiptoast_last_ymd", "") or "")
                except Exception:
                    last = ""
                if last == ymd:
                    return

            # 백그라운드 DB 조회(빠른 summary)
            th = QtCore.QThread(self)
            worker = Worker(query_ship_summary_for_day, self.ci, ymd)
            worker.moveToThread(th)

            worker.finished.connect(lambda res: self._on_shiptoast_summary(res, th))
            worker.failed.connect(lambda tb: self._on_shiptoast_failed(tb, th))
            th.started.connect(worker.run)

            th.start()
            self._start_job(th, worker)
            self._threads.append(th)
        except Exception:
            # UI에는 절대 노출하지 않음
            print(traceback.format_exc())

    def _on_shiptoast_failed(self, tb: str, th: QtCore.QThread):
        try:
            print(tb)
        except Exception:
            pass
        try:
            self._end_job(th)
            th.quit()
            th.wait()
        except Exception:
            pass

    def _on_shiptoast_summary(self, res: dict, th: QtCore.QThread):
        try:
            self._end_job(th)
            th.quit()
            th.wait()
        except Exception:
            pass

        try:
            ymd = str(res.get("ymd") or "")
            rows = list(res.get("sum_rows") or [])
            total_qty = int(res.get("total_qty") or 0)
        except Exception:
            ymd = _dt.date.today().strftime("%Y-%m-%d")
            rows = []
            total_qty = 0

        if not rows or total_qty <= 0:
            return

        # store as last summary
        self._shiptoast_last_summary = {"ymd": ymd, "sum_rows": rows, "total_qty": total_qty}

        # mark shown today
        if self._shiptoast_settings is not None:
            try:
                self._shiptoast_settings.setValue("shiptoast_last_ymd", ymd)
            except Exception:
                pass

        # message (short, but informative)
        # show major parts first (by qty)
        try:
            top = sorted(rows, key=lambda r: int(r.get("sum_qty") or 0), reverse=True)[:6]
        except Exception:
            top = rows[:6]
        parts = []
        for r in top:
            pn = str(r.get("part_name") or "")
            try:
                q = int(r.get("sum_qty") or 0)
            except Exception:
                q = 0
            if not pn:
                continue
            parts.append(f"{pn} {q:,}EA")
        line = " | ".join(parts)
        if len(line) > 180:
            line = line[:180] + "…"
        msg = f"출하등록이 되었습니다.\n{line}"

        self._show_ship_toast("출하 등록", msg, payload=ymd)

    def _show_ship_toast(self, title: str, message: str, payload: str):
        try:
            if getattr(self, "_toast", None) is not None:
                try:
                    self._toast.close()
                except Exception:
                    pass
                self._toast = None

            self._toast = ToastNotification(title, message, payload=payload)
            try:
                self._toast.openRequested.connect(self._on_shiptoast_open)
                self._toast.activateRequested.connect(self._bring_to_front)
            except Exception:
                pass
            self._toast.show_top_right(margin=18)
        except Exception:
            print(traceback.format_exc())

    def _on_shiptoast_open(self, payload: str):
        """토스트 옵션: 오늘 출하 보기"""
        try:
            ymd = (payload or "").strip() or _dt.date.today().strftime("%Y-%m-%d")
            rows = []
            if isinstance(self._shiptoast_last_summary, dict) and self._shiptoast_last_summary.get("ymd") == ymd:
                rows = list(self._shiptoast_last_summary.get("sum_rows") or [])
            dlg = TodayShipDialog(ymd, rows, parent=self)
            try:
                dlg.btn_open_matrix.clicked.connect(lambda: self._open_matrix_for_day(ymd))
            except Exception:
                pass
            dlg.exec_()
        except Exception:
            print(traceback.format_exc())

    def _open_matrix_for_day(self, ymd: str):
        try:
            dlg = SoPojangDialog(self, self.ci, ymd, ymd, "전체", "전체", self.user)
            dlg.exec_()
        except Exception:
            print(traceback.format_exc())

def main():
    app = QtWidgets.QApplication([])
    try:
        app.setQuitOnLastWindowClosed(False)
    except Exception:
        pass
    app.setStyleSheet(APP_DARK_QSS)

    ci = conn_info_default()
    dlg = LoginDialog(None, ci)
    if dlg.exec_() != QtWidgets.QDialog.Accepted or not dlg.user:
        return

    w = MainWindow(dlg.user)
    w.show()
    app.exec_()


if __name__ == "__main__":
    main()