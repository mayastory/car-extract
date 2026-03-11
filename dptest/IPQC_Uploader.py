# -*- coding: utf-8 -*-
"""
IPQC Uploader (XML-FAST, BUNDLE: OMM+AOI+CMM in ONE workbook open)
=================================================================

왜 v1/v2가 느렸냐?
- "AOI 따로 열고, OMM 따로 열고, CMM 따로 열고" (같은 파일을 2~3번 열면 당연히 느림)
- openpyxl(느림) 사용하면 더 느림
- DB도 "1개 통합 테이블 + type 컬럼"이면 조회/삭제/인덱스가 비효율적일 수 있음

이 버전(v3)은:
- OMM 폴더의 "*.xlsx(파일명에 omm 포함)"만 스캔해도,
  파일 1개를 **딱 1번만** 열어서 OMM+AOI+ (있으면) CMM Raw data 까지 같이 업로드한다.
- xlsx를 zip+XML iterparse로 읽어서 openpyxl 없이 빠르게 처리한다.
- OQC 업로더처럼 state(json)로 "이미 업로드한 파일은 재스캔해도 스킵"한다(파일이 바뀐 경우만 재업로드).

시트 규칙
- AOI: 1CAV~4CAV 시트만 읽음
- OMM: "OMM FAI Raw Data"(또는 모델별 변형) 시트만 읽음 (JMP Assist 규칙)
- CMM: "CMM Raw data" (또는 공백/대소문자 변형) 시트만 읽음

엑셀 레이아웃(샘플 기준)
- 1CAV~4CAV:
  - 공통: point_no = B, spc_code = C
  - OMM: USL=L, LSL=M, 측정값=AK..AZ (16칸)
  - AOI: USL=BG, LSL=BH, 측정값=BB..BD (3칸)
- CMM Raw data:
  - point_code = B, feature/desc = C
  - 각 cavity 블럭의 CMM 값(3칸)을 읽음
    (기본: 1CAV F/G/H, 2CAV Y/Z/AA, 3CAV AR/AS/AT, 4CAV BK/BL/BM)

DB 테이블(기본 기대값)
- ipqc_omm_header / ipqc_omm_measurements / ipqc_omm_result
- ipqc_aoi_header / ipqc_aoi_measurements / ipqc_aoi_result
- ipqc_cmm_header / ipqc_cmm_measurements / (선택) ipqc_cmm_result

※ 테이블명이 다르면 아래 TABLES 매핑만 바꿔주면 됨.

Python 3.8+
"""

import os
import re
import time
import random
import json
import datetime
import zipfile
import hashlib
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from xml.etree.ElementTree import iterparse

import pymysql

# --------------------------- DB retry helpers ---------------------------
RETRYABLE_DB_ERR = (pymysql.err.InterfaceError, pymysql.err.OperationalError)

def safe_rollback(conn):
    try:
        if conn is not None and getattr(conn, "open", False):
            conn.rollback()
    except Exception:
        pass

def safe_close(obj):
    try:
        if obj is not None:
            obj.close()
    except Exception:
        pass

def reconnect_db(conn):
    """Close old connection (if any) and open a fresh one."""
    safe_close(conn)
    new_conn = connect_db()
    new_cur = new_conn.cursor()
    set_session_tuning(new_cur)
    ensure_keymap_tables(new_cur)
    return new_conn, new_cur

def ping_db(cur):
    """Ping before large writes; raises if connection is unusable."""
    conn = getattr(cur, "connection", None)
    if conn is None:
        return
    # PyMySQL supports ping(reconnect=True)
    conn.ping(reconnect=True)


# ------------------------- CONFIG (EDIT) -------------------------
DB_HOST = "dpams.ddns.net"
DB_PORT = 3306
DB_USER = "maya"
DB_PASS = "##Gmlakd2323"
DB_NAME = "dp"

# Use JMP assist base candidates exactly
OS_ENV = "WIN"  # WIN / OSX

# Backfill start (YYYY, M)
START_YEAR, START_MONTH = 2025, 1

# Run mode
RUN_MODE = "BACKFILL"   # BACKFILL / HOURLY / TEST
LOOP_INTERVAL_SEC = 120   # 0=one shot, >0=watch loop (sec). ex) 120=2min

# TEST mode knobs (used when RUN_MODE="TEST")
TEST_MONTHS = [(2025, 10)]          # [(YYYY, M), ...]
TEST_MODELS = []                    # ["Memphis IR BASE"] or [] for all
TEST_TOOLS  = []                    # ["A"] or [] for all
MAX_FILES_PER_DIR = 0               # 0=unlimited, else cap per directory

# Store full path for uniqueness + debugging
STORE_SOURCE_PATH = True

# BUNDLE ingest switches
INGEST_OMM = True
INGEST_AOI = True
INGEST_CMM = True

# OMM은 'OMM FAI Raw Data' 시트 기준(JMP Assist 규칙)
USE_OMM_RAWDATA = True

# Scan only these directory types (추천: ["OMM"]만)
SCAN_TYPES = ["OMM"]  # ["OMM"] 권장

# ----------------------------------------------------------------


# Excel layout (1CAV~4CAV)
START_ROW = 11

OMM_DATA_COLS = ["AK","AL","AM","AN","AO","AP","AQ","AR","AS","AT","AU","AV","AW","AX","AY","AZ"]  # 16
AOI_DATA_COLS = ["AK","AL","AM","AN","AO","AP","AQ","AR","AS","AT","AU","AV","AW","AX","AY","AZ"]  # 16
# OMM
OMM_USL_COL = "L"
OMM_LSL_COL = "M"

# AOI
AOI_USL_COL = "BG"
AOI_LSL_COL = "BH"

# Always read these from cav sheets
WANTED_COLS = set(["B","C", OMM_USL_COL, OMM_LSL_COL, AOI_USL_COL, AOI_LSL_COL] + OMM_DATA_COLS + AOI_DATA_COLS)

# CMM Raw data: cavity block start columns (샘플 기준)
CMM_BLOCK_START_COL = {1: "E", 2: "X", 3: "AQ", 4: "BJ"}  # each has 3 cols (Data1..3)

# Retry / batching
CHUNK = 1500
MAX_RETRIES = 6
RETRY_BASE_SLEEP = 0.5


# --------------------- JMP-assist maps ---------------------
MODEL_MAP = {
    "Memphis IR BASE": ("IRBase", "1. IR Base"),
    "Memphis X Carrier": ("X_Carrier", "3. X_Carrier"),
    "Memphis Y Carrier": ("Y_Carrier", "4. Y_Carrier"),
    "Memphis Z Carrier": ("Z_Carrier", "2. Z_Carrier"),
    "Memphis Z Stopper": ("Z_Stopper", "5. Z_Stopper"),
}
TOOL_MAP = {
    "Memphis IR BASE": ["A","B","C","D","E","F","G","H","J","K","L","M","N","P","Q"],
    "Memphis X Carrier": ["A","B","C","D","E","F","G","H","J","K","L","M","N","P","Q"],
    "Memphis Y Carrier": ["A","B","C","D","E","F","G","H","J","K","L","M","N","P","Q"],
    "Memphis Z Carrier": ["A","B","C","D","E","F","G","H","J","K","L","M","N","P","Q"],
    "Memphis Z Stopper": ["A","B","C","D","E","F","G","H","J","K","L","M","N","P","Q"],
}

# ------------------------- table mapping -------------------------
TABLES = {
    "OMM": {"header": "ipqc_omm_header", "meas": "ipqc_omm_measurements", "res": "ipqc_omm_result"},
    "AOI": {"header": "ipqc_aoi_header", "meas": "ipqc_aoi_measurements", "res": "ipqc_aoi_result"},
    "CMM": {"header": "ipqc_cmm_header", "meas": "ipqc_cmm_measurements", "res": "ipqc_cmm_result"},
}


# ------------------------- keymap (pivot key cache) -------------------------
# OMM/CMM export 성능을 위해, 업로드 시점에 keymap을 같이 누적한다.
# - OMM key: (part_name, fai=point_no, spc=spc_code)
# - CMM key: (part_name, point_no)
# AOI는 Data1~DataN 고정이라 keymap 불필요.
KEYMAP_OMM_TABLE = "ipqc_omm_keymap"
KEYMAP_CMM_TABLE = "ipqc_cmm_keymap"
ENABLE_KEYMAP_OMM = True
ENABLE_KEYMAP_CMM = True

def ensure_keymap_tables(cur):
    """Create keymap tables if they don't exist (safe no-op if already exists)."""
    # OMM (may already exist)
    if ENABLE_KEYMAP_OMM:
        cur.execute(f"""
        CREATE TABLE IF NOT EXISTS {KEYMAP_OMM_TABLE} (
          part_name VARCHAR(64) NOT NULL,
          fai VARCHAR(64) NOT NULL,
          spc VARCHAR(64) NOT NULL,
          col_order INT NOT NULL DEFAULT 0,
          PRIMARY KEY (part_name, fai, spc),
          INDEX idx_part_order (part_name, col_order)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
    # CMM
    if ENABLE_KEYMAP_CMM:
        cur.execute(f"""
        CREATE TABLE IF NOT EXISTS {KEYMAP_CMM_TABLE} (
          part_name VARCHAR(64) NOT NULL,
          point_no VARCHAR(64) NOT NULL,
          col_order INT NOT NULL DEFAULT 0,
          PRIMARY KEY (part_name, point_no),
          INDEX idx_part_order (part_name, col_order)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)

def insert_keymap_omm(cur, part_name: str, keys: List[Tuple[str, str]]):
    """keys: list[(fai, spc)]"""
    if not (ENABLE_KEYMAP_OMM and keys):
        return
    # Try with col_order column; fallback if schema is older
    sql1 = f"INSERT IGNORE INTO {KEYMAP_OMM_TABLE} (part_name, fai, spc, col_order) VALUES (%s,%s,%s,0)"
    sql2 = f"INSERT IGNORE INTO {KEYMAP_OMM_TABLE} (part_name, fai, spc) VALUES (%s,%s,%s)"
    rows = [(part_name, fai, (spc or "")) for (fai, spc) in keys if fai]
    if not rows:
        return
    ping_db(cur)
    try:
        for i in range(0, len(rows), CHUNK):
            cur.executemany(sql1, rows[i:i+CHUNK])
    except Exception:
        # fallback for schema without col_order
        for i in range(0, len(rows), CHUNK):
            cur.executemany(sql2, rows[i:i+CHUNK])

def insert_keymap_cmm(cur, part_name: str, keys: List[str]):
    """keys: list[point_no]"""
    if not (ENABLE_KEYMAP_CMM and keys):
        return
    sql1 = f"INSERT IGNORE INTO {KEYMAP_CMM_TABLE} (part_name, point_no, col_order) VALUES (%s,%s,0)"
    sql2 = f"INSERT IGNORE INTO {KEYMAP_CMM_TABLE} (part_name, point_no) VALUES (%s,%s)"
    rows = [(part_name, pn) for pn in keys if pn]
    if not rows:
        return
    ping_db(cur)
    try:
        for i in range(0, len(rows), CHUNK):
            cur.executemany(sql1, rows[i:i+CHUNK])
    except Exception:
        for i in range(0, len(rows), CHUNK):
            cur.executemany(sql2, rows[i:i+CHUNK])


# ------------------------- state (OQC-style) -------------------------
_THIS_DIR = Path(__file__).resolve().parent
STATE_FILE = _THIS_DIR / "ipqc_uploader_state.json"
LOG_FILE = _THIS_DIR / "ipqc_uploader.log"


# ------------------------- logging -------------------------
LOG_TAG = "IPQC"
def log(msg: str):
    ts = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    line = f"[{ts}]	[{LOG_TAG}]	{msg}"
    print(line, flush=True)
    try:
        with open(LOG_FILE, "a", encoding="utf-8") as f:
            f.write(line + "\n")
    except Exception:
        pass


# ------------------------- state helpers -------------------------
def load_state(path: Path) -> Dict[str, str]:
    if path.exists():
        try:
            return json.loads(path.read_text(encoding="utf-8"))
        except Exception as e:
            try:
                bak = path.with_suffix(path.suffix + ".bad")
                bak.write_text(path.read_text(encoding="utf-8", errors="ignore"), encoding="utf-8")
            except Exception:
                pass
            log(f"[STATE] invalid json -> reset ({e})")
            return {}
    return {}

def save_state(path: Path, state: Dict[str, str]) -> None:
    try:
        tmp = path.with_suffix(path.suffix + ".tmp")
        tmp.write_text(json.dumps(state, ensure_ascii=False, indent=2), encoding="utf-8")
        tmp.replace(path)
    except Exception:
        pass

def file_signature(path_str: str) -> str:
    st = os.stat(path_str)
    return f"{int(st.st_size)}:{int(st.st_mtime)}"

def build_state_key(model_disp: str, tool: str, rel_path: str) -> str:
    rp = str(rel_path).replace("\\", "/")
    return f"BUNDLE::{model_disp}::{tool}::{rp}"
_LEGACY_MODEL_PREFIX_RE = re.compile(r"^\s*\d+\.\s*")

def build_legacy_state_key(rel_path: str) -> str:
    """Backward compatible state key used by older IPQC uploader versions.
    Old keys omitted the leading model index like '5. Z_Stopper' -> 'Z_Stopper'.
    """
    rp = str(rel_path).replace("\\", "/").lstrip("/")
    if not rp:
        return rp
    parts = rp.split("/", 1)
    first = _LEGACY_MODEL_PREFIX_RE.sub("", parts[0])
    return first if len(parts) == 1 else (first + "/" + parts[1])



# ------------------------- misc helpers -------------------------
def is_excel_file(name: str) -> bool:
    lname = str(name).lower()
    return lname.endswith('.xlsx') or lname.endswith('.xlsm')

def should_skip_file(name: str, *, bundle: str = "OMM+AOI+CMM") -> bool:
    """
    IPQC 파일 스캔용 필터 (SQL 입력 금지 규칙)

    ✅ 허용(whitelist) 기본:
      - OMM+AOI 번들 파일: 파일명에 'OMM+AOI' 포함
      - CMM 파일: 파일명에 'CMM' 포함  (CMM이 OMM+AOI 파일 안에 같이 들어있는 경우도 있으므로 OMM+AOI는 항상 허용)

    ❌ 금지(blacklist):
      - 임시/락(~, ~$) 및 @로 시작하는 파일
      - 확장자: xlsx/xlsm 외 전부
      - 템플릿/예제/샷/테스트 등:
          * '@@@@' 포함 (예제 날짜)
          * (NOTE) '_1SHOT'은 실제 데이터 파일일 수 있어, suffix가 '...OMM+AOI_1SHOT.xlsx/.xlsm' 인 경우만 허용
          * 'TEST' 포함 (대소문자 무관)
          * '_@ Tool_' / ' @ Tool' 등 템플릿 표기
          * '예제', '샘플', 'sample', 'template'
    """
    if not name:
        return True

    lname = str(name).lower()

    # temp/lock + hidden templates
    if lname.startswith("~$") or lname.startswith("~") or lname.startswith("@"):
        return True

    if not is_excel_file(name):
        return True

    # --- hard blacklist ---
    if "@@@@" in lname:
        return True
    if "test" in lname:
        return True
    if re.search(r"(^|[_\s-])\d+shot($|[_\s.-])", lname):
        # allow only the real data file suffix: ...OMM+AOI_1SHOT.xlsx/.xlsm
        if not (lname.endswith("omm+aoi_1shot.xlsx") or lname.endswith("omm+aoi_1shot.xlsm")):
            return True
    if re.search(r"[_\s]@\s*tool", lname, re.IGNORECASE):
        return True
    if ("예제" in name) or ("샘플" in name) or ("sample" in lname) or ("template" in lname):
        return True
    # --- whitelist: STRICT naming (파일명 끝이 정확히 OMM+AOI.xlsx / OMM+AOI_1SHOT.xlsx 이어야만 허용) ---
    allow = False

    # 사람들이 파일 제목을 제각각 붙이므로, '...OMM+AOI - 일부치수.xlsx' 같은 건 전부 금지.
    # ✅ 허용:
    #   - ...OMM+AOI.xlsx / ...OMM+AOI.xlsm
    #   - ...OMM+AOI_1SHOT.xlsx / ...OMM+AOI_1SHOT.xlsm
    if (lname.endswith("omm+aoi.xlsx") or lname.endswith("omm+aoi.xlsm") or
        lname.endswith("omm+aoi_1shot.xlsx") or lname.endswith("omm+aoi_1shot.xlsm")):
        allow = True
    # If user configured bundle to exclude certain types, honor it.
    b = (bundle or "").upper()
    if "CMM" not in b and "cmm" in lname:
        allow = False
    if ("OMM+AOI" not in b) and ("omm+aoi" in lname) and ("cmm" not in lname):
        allow = False

    if not allow:
        return True

    return False



def parse_meas_date_from_filename(filename: str) -> Optional[str]:
    """
    파일명에서 측정일 추출.

    기존(v4): 첫 번째 6자리 숫자를 잡아서 817118 같은 부품번호를 날짜로 오인할 수 있었음.
    개선(v5):
      - YYYYMMDD(8자리) / YYMMDD(6자리) 후보를 전부 찾고
      - 월/일 범위가 유효한 것만 통과
      - 파일명 "끝쪽"에 있는(가장 뒤 match) 날짜를 선택
      - 템플릿(@@@@)은 None
    """
    if not filename:
        return None
    if "@@@@" in filename:
        return None

    candidates: List[Tuple[int, str]] = []

    # 8-digit first (YYYYMMDD)
    for m in re.finditer(r"(?<!\d)(\d{8})(?!\d)", filename):
        s = m.group(1)
        yyyy = int(s[0:4])
        mm = int(s[4:6])
        dd = int(s[6:8])
        if 2000 <= yyyy <= 2035 and 1 <= mm <= 12 and 1 <= dd <= 31:
            candidates.append((m.start(1), f"{yyyy:04d}-{mm:02d}-{dd:02d}"))

    # 6-digit (YYMMDD)
    for m in re.finditer(r"(?<!\d)(\d{6})(?!\d)", filename):
        s = m.group(1)
        yy = int(s[0:2])
        mm = int(s[2:4])
        dd = int(s[4:6])
        # YY: 20~35 정도만 허용 (IPQC 파일명 규칙 상)
        if 20 <= yy <= 35 and 1 <= mm <= 12 and 1 <= dd <= 31:
            yyyy = 2000 + yy
            candidates.append((m.start(1), f"{yyyy:04d}-{mm:02d}-{dd:02d}"))

    if not candidates:
        return None

    # choose the last occurrence in filename
    candidates.sort(key=lambda x: x[0])
    return candidates[-1][1]
def parse_cavity_int(sheet_name: str) -> Optional[int]:
    s = str(sheet_name).strip().upper().replace(" ", "")
    m = re.match(r"^(\d+)CAV$", s)
    if not m:
        return None
    try:
        return int(m.group(1))
    except Exception:
        return None

def safe_float(x: Optional[str]) -> Optional[float]:
    if x is None:
        return None
    s = str(x).strip()
    if s == "":
        return None
    try:
        return float(s)
    except Exception:
        return None

def month_label(m: int) -> str:
    return f"{m}월"


# --------------------- base candidates (JMP order) ---------------------
def get_base_src_candidates(os_choice: str) -> List[str]:
    mapped_win = r"X:\3.측정실\Memphis(25)\2.0"
    mapped_osx = r"A:\3.측정실\Memphis(25)\2.0"
    unc_candidates = [
        r"\\192.168.1.136\품질\3.측정실\Memphis(25)\2.0",
        r"\\192.168.1.136\3.측정실\Memphis(25)\2.0",
        r"\\192.168.1.136\품질\Memphis(25)\2.0",
    ]
    return ([mapped_win] if os_choice == "WIN" else [mapped_osx]) + unc_candidates


# --------------------- months range ---------------------
def iter_months_backfill() -> List[Tuple[int,int,str]]:
    now = datetime.datetime.now()
    ey, em = now.year, now.month
    y, m = START_YEAR, START_MONTH
    out = []
    while (y < ey) or (y == ey and m <= em):
        out.append((y, m, month_label(m)))
        m += 1
        if m == 13:
            y += 1
            m = 1
    return out
def iter_months_hourly() -> List[Tuple[int,int,str]]:
    now = datetime.datetime.now()
    cur = (now.year, now.month)
    py, pm = now.year, now.month - 1
    if pm == 0:
        py -= 1
        pm = 12
    return [(py, pm, month_label(pm)), (cur[0], cur[1], month_label(cur[1]))]

def iter_months_test() -> List[Tuple[int,int,str]]:
    return [(y,m,month_label(m)) for (y,m) in TEST_MONTHS]


# --------------------- XLSX XML stream parsing ---------------------
_NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_NS_R = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
def _q(ns, tag): return f"{{{ns}}}{tag}"

def split_cell_ref(ref: str) -> Tuple[str, int]:
    m = re.match(r"^([A-Z]+)(\d+)$", ref.upper())
    if not m:
        return "", -1
    return m.group(1), int(m.group(2))

def read_shared_strings(zf: zipfile.ZipFile) -> List[str]:
    path = "xl/sharedStrings.xml"
    if path not in zf.namelist():
        return []
    out = []
    with zf.open(path) as f:
        context = iterparse(f, events=("end",))
        for _, el in context:
            if el.tag == _q(_NS_MAIN, "si"):
                texts = []
                for t in el.iter(_q(_NS_MAIN, "t")):
                    if t.text:
                        texts.append(t.text)
                out.append("".join(texts))
                el.clear()
    return out

def sheet_rid_map(zf: zipfile.ZipFile) -> Dict[str, str]:
    wb_path = "xl/workbook.xml"
    if wb_path not in zf.namelist():
        return {}
    name_to_rid = {}
    with zf.open(wb_path) as f:
        context = iterparse(f, events=("end",))
        for _, el in context:
            if el.tag == _q(_NS_MAIN, "sheet"):
                name = el.attrib.get("name")
                rid = el.attrib.get(_q(_NS_R, "id"))
                if name and rid:
                    name_to_rid[name] = rid
                el.clear()
    return name_to_rid

def rid_to_target_map(zf: zipfile.ZipFile) -> Dict[str, str]:
    rel_path = "xl/_rels/workbook.xml.rels"
    if rel_path not in zf.namelist():
        return {}
    rid_to_target = {}
    with zf.open(rel_path) as f:
        context = iterparse(f, events=("end",))
        for _, el in context:
            if el.tag.endswith("Relationship"):
                rid = el.attrib.get("Id")
                tgt = el.attrib.get("Target")
                if rid and tgt:
                    if not tgt.startswith("/"):
                        rid_to_target[rid] = "xl/" + tgt.lstrip("./")
                el.clear()
    return rid_to_target

def list_sheets(zf: zipfile.ZipFile) -> Dict[str, str]:
    """Return {sheet_name: sheet_xml_path}"""
    name_to_rid = sheet_rid_map(zf)
    rid_to_tgt = rid_to_target_map(zf)
    out = {}
    for name, rid in name_to_rid.items():
        xml_path = rid_to_tgt.get(rid)
        if xml_path and xml_path in zf.namelist():
            out[str(name)] = xml_path
    return out

def list_cav_sheets(all_sheets: Dict[str, str]) -> List[Tuple[str, str]]:
    out = []
    for name, xml_path in all_sheets.items():
        if re.match(r"^\d+\s*CAV$", str(name).strip(), re.I):
            out.append((name, xml_path))
    return out

def normalize_sheet_name(name: str) -> str:
    return re.sub(r"\s+", "", str(name).strip().lower())

def find_cmm_sheet(all_sheets: Dict[str, str]) -> Optional[Tuple[str, str]]:
    # allow: "CMM Raw data", "CMM RawData", etc
    for name, xml_path in all_sheets.items():
        nn = normalize_sheet_name(name)
        if nn in ("cmmrawdata", "cmmrawdata(1)", "cmmraw_data"):
            return (name, xml_path)
        if "cmm" in nn and "raw" in nn and "data" in nn:
            return (name, xml_path)

    return None


# --------------------------- OMM Raw Data sheet parser (JMP Assist-compatible) ---------------------------
def _parse_omm_point_and_spc(raw_b: str) -> Tuple[str, str]:
    """
    OMM:
      - DB에는 엑셀의 FAI 라벨(B열)을 **있는 그대로** 저장한다.
      - 공백/슬래시/문자 하나라도 정리/변환하면 매핑 순서/라벨 매칭이 깨진다.
      - 따라서 라벨 파싱/분리/접두어 보정/대소문자 변경을 절대 하지 않는다.
      - OMM에서 spc는 사용하지 않으므로 항상 빈값('')로 저장한다.
    returns: (fai_label_raw, '')
    """
    s = (str(raw_b) if raw_b is not None else "")
    if s == "":
        return ("", "")
    return (s, "")
def parse_omm_rawdata_openpyxl(filepath: str, model_disp: str) -> Tuple[bool, Dict[int, List[Tuple[str, str, List[Optional[float]], Optional[float], Optional[float]]]], str]:
    """
    Returns (ok, rows_by_cav, reason)
    rows_by_cav[cav] = list[(point_no, spc_code, vals[3], usl, lsl)]
    - OMM Raw Data layout follows JMP Assist.py make_jmp_data()
    """
    try:
        from openpyxl import load_workbook
        import math
    except Exception as e:
        return (False, {}, f"openpyxl not available: {e}")

    try:
        wb = load_workbook(filepath, data_only=True, read_only=True)
    except Exception as e:
        return (False, {}, f"load_workbook failed: {e}")

    sheetnames = list(wb.sheetnames or [])
    low = [s.lower() for s in sheetnames]

    # pick sheet name per model (JMP Assist rules)
    target = None
    if model_disp == "Memphis Y Carrier":
        for i, ln in enumerate(low):
            if "omm fai data" in ln:
                target = sheetnames[i]; break
    elif model_disp == "Memphis Z Stopper":
        for key in ("omm fai raw data", "fai raw data"):
            for i, ln in enumerate(low):
                if key in ln:
                    target = sheetnames[i]; break
            if target: break
    elif model_disp == "Memphis Z Carrier":
        for i, ln in enumerate(low):
            if "omm fai raw data" in ln:
                target = sheetnames[i]; break
    else:
        for i, ln in enumerate(low):
            if "omm fai raw data" in ln:
                target = sheetnames[i]; break

    if not target:
        return (False, {}, "OMM raw data sheet not found")

    ws = wb[target]

    # default fixed cav columns (F~Q): 4 cavities x 3 cols
    cav_titles = ["1CAV", "2CAV", "3CAV", "4CAV"]
    cav_cols_fixed = {
        "1CAV": [6, 7, 8],
        "2CAV": [9, 10, 11],
        "3CAV": [12, 13, 14],
        "4CAV": [15, 16, 17],
    }

    def _coerce_num(x):
        if x is None:
            return None
        if isinstance(x, (int, float)):
            try:
                if isinstance(x, float) and math.isnan(x):
                    return None
            except Exception:
                pass
            return float(x)
        s = str(x).strip()
        if s == "":
            return None
        try:
            return float(s)
        except Exception:
            return None

    # Find data start row: first row whose FAI cell starts with "FAI"
    # (Z Stopper uses col 3 in JMP Assist, others use col 2)
    fai_col = 3 if model_disp == "Memphis Z Stopper" else 2
    design_col, tol_up_col, tol_dn_col = 3, 4, 5

    max_row = ws.max_row or 0
    data_start_row = None
    for r in range(1, max_row + 1):
        v = ws.cell(row=r, column=fai_col).value
        if str(v or "").strip().upper().startswith("FAI"):
            data_start_row = r
            break
    if not data_start_row:
        return (False, {}, "No FAI start row found")

    # For Y Carrier, try dynamic cavity columns like JMP Assist
    cav_order = list(cav_titles)
    cav_cols = cav_cols_fixed

    if model_disp == "Memphis Y Carrier":
        max_col = ws.max_column or 0
        cav_map = {}
        cav_order = []
        data_start_row2 = None
        for row in range(1, 20):
            for col in range(1, max_col + 1):
                val = ws.cell(row=row, column=col).value
                if val and str(val).strip().upper().endswith("CAV"):
                    data_row = row + 1
                    labels = [ws.cell(row=data_row, column=col + i).value for i in range(3)]
                    try:
                        lab_ok = (labels == [1, 2, 3]) or all(str(x).isdigit() for x in labels)
                    except Exception:
                        lab_ok = False
                    if lab_ok:
                        cav_map[str(val).strip()] = [col, col + 1, col + 2]
                        cav_order.append(str(val).strip())
                        if data_start_row2 is None or data_row > data_start_row2:
                            data_start_row2 = data_row
        if cav_map and data_start_row2:
            cav_cols = cav_map
            # FAI list for Y Carrier is in col2 starting at data_start_row2+1
            fai_col = 2
            data_start_row = data_start_row2 + 1

    # rows_by_cav
    rows_by_cav: Dict[int, List[Tuple[str, str, List[Optional[float]], Optional[float], Optional[float]]]] = {1: [], 2: [], 3: [], 4: []}

    for r in range(data_start_row, max_row + 1):
        raw_fai = ws.cell(row=r, column=fai_col).value
        fai_raw = "" if raw_fai is None else str(raw_fai)
        fai_chk = fai_raw.strip()
        if fai_chk == "" or fai_chk.lower() in ("left", "right"):
            break
        if "RE" in fai_chk:
            continue

        point_no, spc_code = _parse_omm_point_and_spc(fai_raw)

        design = _coerce_num(ws.cell(row=r, column=design_col).value)
        tol_up = _coerce_num(ws.cell(row=r, column=tol_up_col).value)
        tol_dn = _coerce_num(ws.cell(row=r, column=tol_dn_col).value)
        usl = (design + tol_up) if (design is not None and tol_up is not None) else None
        lsl = (design + tol_dn) if (design is not None and tol_dn is not None) else None

        # Determine which cavities exist in this sheet
        for cav_label in cav_order:
            cav_norm = str(cav_label).strip().upper().replace(" ", "")
            if cav_norm.startswith("1"):
                cav_i = 1
            elif cav_norm.startswith("2"):
                cav_i = 2
            elif cav_norm.startswith("3"):
                cav_i = 3
            elif cav_norm.startswith("4"):
                cav_i = 4
            else:
                continue

            cols = cav_cols.get(cav_label) if isinstance(cav_cols, dict) else None
            if cols is None and isinstance(cav_cols, dict):
                cols = cav_cols.get(cav_norm)  # try normalized key
            if not cols:
                # fallback to fixed
                cols = cav_cols_fixed.get(f"{cav_i}CAV")
            if not cols:
                continue

            vals = [_coerce_num(ws.cell(row=r, column=cols[i]).value) for i in range(3)]
            # Only keep if at least one measurement exists
            if any(v is not None for v in vals):
                rows_by_cav[cav_i].append((point_no, spc_code, vals, usl, lsl))

    # prune empty cavities
    rows_by_cav = {c: rows for c, rows in rows_by_cav.items() if rows}
    if not rows_by_cav:
        return (False, {}, "No OMM rows parsed from raw data sheet")
    return (True, rows_by_cav, "OK")




def parse_omm_rawdata_xml_stream(zf: zipfile.ZipFile, all_sheets: Dict[str, str], shared: List[str], model_disp: str) -> Tuple[bool, Dict[int, List[Tuple[str, str, List[Optional[float]], Optional[float], Optional[float]]]], str]:
    """
    Fast XML-stream parser for 'OMM FAI Raw Data' sheet (no openpyxl).

    Returns (ok, rows_by_cav, reason)
    rows_by_cav[cav] = list[(point_no, spc_code, vals[3], usl, lsl)]
    - Layout follows JMP Assist.py make_jmp_data()
    """
    # pick sheet name per model (same rules as parse_omm_rawdata_openpyxl)
    target_xml = None
    for name, xml_path in (all_sheets or {}).items():
        ln = str(name).lower()
        if model_disp == "Memphis Y Carrier":
            if "omm fai data" in ln:
                target_xml = xml_path
                break
        elif model_disp == "Memphis Z Stopper":
            if ("omm fai raw data" in ln) or ("fai raw data" in ln):
                target_xml = xml_path
                break
        elif model_disp == "Memphis Z Carrier":
            if "omm fai raw data" in ln:
                target_xml = xml_path
                break
        else:
            if "omm fai raw data" in ln:
                target_xml = xml_path
                break

    if not target_xml:
        return (False, {}, "OMM raw data sheet not found")

    # default fixed cav columns (F~Q): 4 cavities x 3 cols
    cav_titles = ["1CAV", "2CAV", "3CAV", "4CAV"]
    cav_cols_fixed = {
        "1CAV": ["F", "G", "H"],
        "2CAV": ["I", "J", "K"],
        "3CAV": ["L", "M", "N"],
        "4CAV": ["O", "P", "Q"],
    }

    # Find data start row: first row whose FAI cell starts with "FAI"
    # (Z Stopper uses col C in JMP Assist, others use col B)
    fai_col = "C" if model_disp == "Memphis Z Stopper" else "B"
    design_col, tol_up_col, tol_dn_col = "C", "D", "E"

    cav_order = list(cav_titles)
    cav_cols: Dict[str, List[str]] = dict(cav_cols_fixed)

    # For Y Carrier, try dynamic cavity columns like JMP Assist (header scan only)
    data_start_row: Optional[int] = None
    if model_disp == "Memphis Y Carrier":
        try:
            header = scan_sheet_header_rows(zf, target_xml, shared, max_row=25)
        except Exception:
            header = {}

        cav_map: Dict[str, List[str]] = {}
        cav_order2: List[str] = []
        data_start_row2: Optional[int] = None

        for row in range(1, 20):
            cells = header.get(row) or {}
            for col, raw in cells.items():
                if raw is None:
                    continue
                sval = str(raw).strip()
                if not sval:
                    continue
                if sval.upper().endswith("CAV"):
                    data_row = row + 1
                    row_cells = header.get(data_row) or {}
                    c1, c2, c3 = col, _next_col(col), _next_col(_next_col(col))
                    labels = [row_cells.get(c1), row_cells.get(c2), row_cells.get(c3)]
                    try:
                        lab_ok = (labels == [1, 2, 3]) or (labels == ["1", "2", "3"]) or all(str(x).strip().isdigit() for x in labels)
                    except Exception:
                        lab_ok = False
                    if lab_ok:
                        cav_map[sval] = [c1, c2, c3]
                        cav_order2.append(sval)
                        if data_start_row2 is None or data_row > data_start_row2:
                            data_start_row2 = data_row

        if cav_map and data_start_row2:
            cav_cols = cav_map
            cav_order = cav_order2
            # FAI list for Y Carrier is in col B starting at (label row + 1)
            fai_col = "B"
            data_start_row = data_start_row2 + 1

    # If not Y Carrier override, search for the first "FAI*" row (streaming, early-stop)
    if data_start_row is None:
        with zf.open(target_xml) as f:
            context = iterparse(f, events=("end",))
            for _, el in context:
                if el.tag != _q(_NS_MAIN, "row"):
                    continue
                r_attr = el.attrib.get("r")
                try:
                    row_no = int(r_attr) if r_attr else -1
                except Exception:
                    row_no = -1
                if row_no == -1:
                    el.clear()
                    continue

                for c in el.findall(_q(_NS_MAIN, "c")):
                    ref = c.attrib.get("r")
                    if not ref:
                        continue
                    col, rr = split_cell_ref(ref)
                    if rr != row_no or col != fai_col:
                        continue
                    val = get_cell_value(c, shared)
                    if val is not None and str(val).strip().upper().startswith("FAI"):
                        data_start_row = row_no
                    break

                el.clear()
                if data_start_row is not None:
                    break

        if not data_start_row:
            return (False, {}, "No FAI start row found")

    # wanted columns (parse only what we need)
    wanted = set([fai_col, design_col, tol_up_col, tol_dn_col])
    for cav_label in cav_order:
        cols = cav_cols.get(cav_label)
        if cols:
            wanted.update(cols)
    # keep fixed fallback columns available (cheap)
    for cols in cav_cols_fixed.values():
        wanted.update(cols)

    rows_by_cav: Dict[int, List[Tuple[str, str, List[Optional[float]], Optional[float], Optional[float]]]] = {1: [], 2: [], 3: [], 4: []}

    with zf.open(target_xml) as f:
        context = iterparse(f, events=("end",))
        for _, el in context:
            if el.tag != _q(_NS_MAIN, "row"):
                continue
            r_attr = el.attrib.get("r")
            try:
                row_no = int(r_attr) if r_attr else -1
            except Exception:
                row_no = -1

            if row_no == -1 or row_no < data_start_row:
                el.clear()
                continue

            cell_vals: Dict[str, str] = {}
            for c in el.findall(_q(_NS_MAIN, "c")):
                ref = c.attrib.get("r")
                if not ref:
                    continue
                col, rr = split_cell_ref(ref)
                if rr != row_no:
                    continue
                if col not in wanted:
                    continue
                v = get_cell_value(c, shared)
                if v is not None:
                    cell_vals[col] = v

            raw_fai = cell_vals.get(fai_col)
            fai_raw = "" if raw_fai is None else str(raw_fai)
            fai_chk = fai_raw.strip()
            if fai_chk == "" or fai_chk.lower() in ("left", "right"):
                el.clear()
                break
            if "RE" in fai_chk:
                el.clear()
                continue

            point_no, spc_code = _parse_omm_point_and_spc(fai_raw)

            design = safe_float(cell_vals.get(design_col))
            tol_up = safe_float(cell_vals.get(tol_up_col))
            tol_dn = safe_float(cell_vals.get(tol_dn_col))
            usl = (design + tol_up) if (design is not None and tol_up is not None) else None
            lsl = (design + tol_dn) if (design is not None and tol_dn is not None) else None

            for cav_label in cav_order:
                cav_norm = str(cav_label).strip().upper().replace(" ", "")
                if cav_norm.startswith("1"):
                    cav_i = 1
                elif cav_norm.startswith("2"):
                    cav_i = 2
                elif cav_norm.startswith("3"):
                    cav_i = 3
                elif cav_norm.startswith("4"):
                    cav_i = 4
                else:
                    continue

                cols = cav_cols.get(cav_label) or cav_cols.get(cav_norm)
                if not cols:
                    cols = cav_cols_fixed.get(f"{cav_i}CAV")
                if not cols:
                    continue

                vals = [safe_float(cell_vals.get(cols[i])) for i in range(3)]
                if any(v is not None for v in vals):
                    rows_by_cav[cav_i].append((point_no, spc_code, vals, usl, lsl))

            el.clear()

    # prune empty cavities
    rows_by_cav = {c: rows for c, rows in rows_by_cav.items() if rows}
    if not rows_by_cav:
        return (False, {}, "No OMM rows parsed from raw data sheet")
    return (True, rows_by_cav, "OK")


def _excel_col_to_num(col: str) -> int:
    col = (col or "").strip().upper()
    n = 0
    for ch in col:
        if "A" <= ch <= "Z":
            n = n * 26 + (ord(ch) - 64)
    return n

def scan_sheet_header_rows(zf: zipfile.ZipFile, sheet_xml_path: str, shared: List[str],
                          max_row: int = 25) -> Dict[int, Dict[str, str]]:
    """
    Fast header scan (XML iterparse):
      - reads only rows <= max_row
      - returns {row_no: {col_letter: text}}
    Used for detecting CMM layout variations without openpyxl.
    """
    row_map: Dict[int, Dict[str, str]] = {}
    with zf.open(sheet_xml_path) as f:
        context = iterparse(f, events=("end",))
        for _, el in context:
            if el.tag == _q(_NS_MAIN, "row"):
                r_attr = el.attrib.get("r")
                try:
                    row_no = int(r_attr) if r_attr else -1
                except Exception:
                    row_no = -1

                if row_no > max_row and row_no != -1:
                    el.clear()
                    break

                cell_vals = {}
                for c in el.findall(_q(_NS_MAIN, "c")):
                    ref = c.attrib.get("r")
                    if not ref:
                        continue
                    col, _r = split_cell_ref(ref)
                    if not col:
                        continue
                    v = get_cell_value(c, shared)
                    if v is None:
                        continue
                    cell_vals[col] = str(v)

                if cell_vals:
                    row_map[row_no] = cell_vals

                el.clear()
    return row_map

def detect_cmm_layout(zf: zipfile.ZipFile, sheet_xml_path: str, shared: List[str],
                      max_scan_row: int = 25) -> Tuple[Dict[int, str], bool, str, Dict[int, str]]:
    """
    Detect CMM block start columns in 'CMM Raw data' with 2-step strategy:
      1) If header contains literal 'CMM' cells: use those columns as block starts (ordered left->right).
      2) Else: find numeric header row that has consecutive '1','2','3' triples; use '1' columns as starts.
         Optionally match starts to cavity labels like '1Cavity' / misspelling '1Cavtiy'.
    Returns:
      (cav_start_cols, ok, reason, cav_label_cols)
        - cav_start_cols: {cavity:int -> start_col:str}
        - ok: True if layout detected, False otherwise
        - cav_label_cols: {cavity:int -> label_col:str} (best-effort)
    """
    header = scan_sheet_header_rows(zf, sheet_xml_path, shared, max_row=max_scan_row)

    # cavity label hints (best-effort)
    cav_labels: Dict[int, str] = {}
    for _rn, cells in header.items():
        for col, raw in cells.items():
            s = str(raw).strip().lower().replace(" ", "")
            m = re.match(r"^([1-9])cav", s)
            if not m and "cav" in s:
                # allow: 1cavtiy / 1cavity / 1cav...
                m = re.match(r"^([1-9]).*cav", s)
            if m:
                cav = int(m.group(1))
                if 1 <= cav <= 9 and cav not in cav_labels:
                    cav_labels[cav] = col

    # (1) literal 'CMM' header cells
    cmm_cols = []
    for _rn, cells in header.items():
        for col, raw in cells.items():
            if str(raw).strip().lower() == "cmm":
                cmm_cols.append(col)
    cmm_cols = sorted(set(cmm_cols), key=_excel_col_to_num)
    if cmm_cols:
        # map in order (max 4 cavities in our flows)
        cav_start = {i + 1: c for i, c in enumerate(cmm_cols[:4])}
        return cav_start, True, "CMM header columns", cav_labels

    # (2) numeric 1/2/3 row scan
    best_row = None
    best_starts: List[str] = []
    for rn, cells in header.items():
        starts = []
        for col, raw in cells.items():
            if str(raw).strip() == "1":
                c2 = _next_col(col)
                c3 = _next_col(c2)
                if str(cells.get(c2, "")).strip() == "2" and str(cells.get(c3, "")).strip() == "3":
                    starts.append(col)
        # prefer rows with more triples; tie-breaker: smaller row number
        if len(starts) > len(best_starts) or (len(starts) == len(best_starts) and starts and best_row and rn < best_row):
            best_row = rn
            best_starts = starts

    best_starts = sorted(set(best_starts), key=_excel_col_to_num)
    if best_starts:
        # if cavity labels exist, map by nearest start to the right of label col
        cav_start: Dict[int, str] = {}
        if cav_labels:
            used = set()
            for cav in sorted(cav_labels.keys()):
                lbl_col = cav_labels[cav]
                lbl_idx = _excel_col_to_num(lbl_col)
                # candidates not used
                cands = [(abs(_excel_col_to_num(sc) - lbl_idx), _excel_col_to_num(sc), sc) for sc in best_starts if sc not in used]
                if not cands:
                    continue
                # prefer start col >= label col (to the right); else closest
                right = [(d, idx, sc) for (d, idx, sc) in cands if idx >= lbl_idx]
                pick = min(right, default=min(cands))
                _, _, sc = pick
                cav_start[cav] = sc
                used.add(sc)
            # If some starts remain (labels missing), append sequentially after max cav
            if len(cav_start) < min(4, len(best_starts)):
                next_cav = 1
                while next_cav in cav_start:
                    next_cav += 1
                for sc in best_starts:
                    if sc in used:
                        continue
                    if next_cav > 4:
                        break
                    cav_start[next_cav] = sc
                    used.add(sc)
                    next_cav += 1
        else:
            cav_start = {i + 1: sc for i, sc in enumerate(best_starts[:4])}

        if cav_start:
            return cav_start, True, f"1/2/3 numeric header row (row {best_row})", cav_labels

    return {}, False, "CMM layout not detected (no 'CMM' header, no 1/2/3 row)", cav_labels


def get_cell_value(cell_el, shared: List[str]) -> Optional[str]:
    t = cell_el.attrib.get("t")
    v_el = cell_el.find(_q(_NS_MAIN, "v"))
    if t == "s":
        if v_el is None or v_el.text is None:
            return None
        try:
            idx = int(v_el.text)
            return shared[idx] if 0 <= idx < len(shared) else None
        except Exception:
            return None
    if t == "inlineStr":
        is_el = cell_el.find(_q(_NS_MAIN, "is"))
        if is_el is None:
            return None
        texts = []
        for t_el in is_el.iter(_q(_NS_MAIN, "t")):
            if t_el.text:
                texts.append(t_el.text)
        return "".join(texts) if texts else None
    if v_el is None or v_el.text is None:
        return None
    return v_el.text

def parse_sheet_rows(zf: zipfile.ZipFile, sheet_xml_path: str, shared: List[str],
                     start_row: int, wanted_cols: set) -> Dict[int, Dict[str, str]]:
    """Return {row_no: {col_letter: raw_text}} for only wanted cols."""
    row_map: Dict[int, Dict[str, str]] = {}
    with zf.open(sheet_xml_path) as f:
        context = iterparse(f, events=("end",))
        for _, el in context:
            if el.tag == _q(_NS_MAIN, "row"):
                r_attr = el.attrib.get("r")
                try:
                    row_no = int(r_attr) if r_attr else -1
                except Exception:
                    row_no = -1

                if row_no >= start_row:
                    cell_vals = {}
                    for c in el.findall(_q(_NS_MAIN, "c")):
                        ref = c.attrib.get("r")
                        if not ref:
                            continue
                        col, rr = split_cell_ref(ref)
                        if rr != row_no:
                            continue
                        if col not in wanted_cols:
                            continue
                        val = get_cell_value(c, shared)
                        if val is not None:
                            cell_vals[col] = val
                    if cell_vals:
                        row_map[row_no] = cell_vals
                el.clear()
    return row_map


# --------------------------- DB -----------------------------------
def connect_db():
    return pymysql.connect(
        host=DB_HOST, port=DB_PORT, user=DB_USER, password=DB_PASS, database=DB_NAME,
        charset="utf8mb4", autocommit=False, cursorclass=pymysql.cursors.DictCursor,
        connect_timeout=10, read_timeout=120, write_timeout=120
    )

def set_session_tuning(cur):
    # keep minimal
    try:
        cur.execute("SET SESSION innodb_lock_wait_timeout=10")
    except Exception:
        pass

def make_source_key(tool: str, cavity: int, source_file: str) -> str:
    """
    DB header unique key = (tool, cavity, source_key)
    - source_key는 '파일명 기반'으로 고정(파일이 수정되어도 같은 row를 UPDATE하기 위함)
    - 파일명 자체에 날짜가 들어가므로(샘플 기준) 충돌 위험 낮음
    """
    raw = f"{tool}|{cavity}|{source_file}".encode("utf-8", "ignore")
    return hashlib.sha1(raw).hexdigest()


def upsert_header(cur, table: str, part_name: str, meas_date: Optional[str],
                  tool: str, cavity: int, source_file: str, source_path: Optional[str],
                  source_key: str, file_mtime: int, file_size: int) -> int:
    # Expect UNIQUE key on (tool,cavity,source_key)
    if STORE_SOURCE_PATH:
        cur.execute(
            f"""
            INSERT INTO {table}
              (part_name, meas_date, tool, cavity, source_file, source_path, source_key, file_mtime, file_size)
            VALUES
              (%s,%s,%s,%s,%s,%s,%s,%s,%s)
            ON DUPLICATE KEY UPDATE
              part_name=VALUES(part_name),
              meas_date=VALUES(meas_date),
              tool=VALUES(tool),
              cavity=VALUES(cavity),
              source_file=VALUES(source_file),
              source_path=VALUES(source_path),
              source_key=VALUES(source_key),
              file_mtime=VALUES(file_mtime),
              file_size=VALUES(file_size),
              id=LAST_INSERT_ID(id)
            """,
            (part_name, meas_date, tool, cavity, source_file, source_path, source_key, file_mtime, file_size),
        )
    else:
        cur.execute(
            f"""
            INSERT INTO {table}
              (part_name, meas_date, tool, cavity, source_file, source_key, file_mtime, file_size)
            VALUES
              (%s,%s,%s,%s,%s,%s,%s,%s)
            ON DUPLICATE KEY UPDATE
              part_name=VALUES(part_name),
              meas_date=VALUES(meas_date),
              tool=VALUES(tool),
              cavity=VALUES(cavity),
              source_file=VALUES(source_file),
              source_key=VALUES(source_key),
              file_mtime=VALUES(file_mtime),
              file_size=VALUES(file_size),
              id=LAST_INSERT_ID(id)
            """,
            (part_name, meas_date, tool, cavity, source_file, source_key, file_mtime, file_size),
        )
    cur.execute("SELECT LAST_INSERT_ID() AS id")
    return int(cur.fetchone()["id"])

def clear_existing(cur, meas_table: str, res_table: Optional[str], header_id: int):
    cur.execute(f"DELETE FROM {meas_table} WHERE header_id=%s", (header_id,))
    if res_table:
        cur.execute(f"DELETE FROM {res_table} WHERE header_id=%s", (header_id,))

def bulk_insert_measurements(cur, typ: str, table: str, rows: List[Tuple]):
    if not rows:
        return

    if typ in ("OMM", "AOI"):
        # dp.sql: (header_id, fai, spc, row_index, value)
        sql = f"""
          INSERT INTO {table} (header_id, fai, spc, row_index, value)
          VALUES (%s,%s,%s,%s,%s)
          ON DUPLICATE KEY UPDATE
            value=VALUES(value),
            spc=VALUES(spc)
        """
    elif typ == "CMM":
        # dp.sql: (header_id, point_no, row_index, value)
        sql = f"""
          INSERT INTO {table} (header_id, point_no, row_index, value)
          VALUES (%s,%s,%s,%s)
          ON DUPLICATE KEY UPDATE
            value=VALUES(value)
        """
    else:
        raise ValueError(f"Unknown typ={typ}")

    # OMM 라벨 포맷 변경(FAI 라벨 원문 저장) 마이그레이션 대비:
    # - header_id 단위로 기존 데이터를 먼저 지우고, 현재 파일 내용으로 다시 채운다.
    # - (기존에 'FAI 1' + spc='A' 처럼 잘려 저장된 행이 남아서 컬럼이 분리되는 것을 방지)
    if typ == "OMM":
        header_ids = sorted({r[0] for r in rows if r and r[0]})
        if header_ids:
            ping_db(cur)
            for j in range(0, len(header_ids), CHUNK):
                chunk = header_ids[j:j+CHUNK]
                ph = ",".join(["%s"] * len(chunk))
                cur.execute(f"DELETE FROM {table} WHERE header_id IN ({ph})", tuple(chunk))

    # keep connection alive for big batches
    ping_db(cur)
    for i in range(0, len(rows), CHUNK):
        cur.executemany(sql, rows[i:i+CHUNK])

def bulk_insert_results(cur, typ: str, table: str, rows: List[Tuple]):
    if not rows:
        return

    if typ in ("OMM", "AOI"):
        # dp.sql: (header_id, fai, max_val, min_val, mean_val, usl, lsl, result_ok)
        sql = f"""
          INSERT INTO {table} (header_id, fai, max_val, min_val, mean_val, usl, lsl, result_ok)
          VALUES (%s,%s,%s,%s,%s,%s,%s,%s)
          ON DUPLICATE KEY UPDATE
            max_val=VALUES(max_val),
            min_val=VALUES(min_val),
            mean_val=VALUES(mean_val),
            usl=VALUES(usl),
            lsl=VALUES(lsl),
            result_ok=VALUES(result_ok)
        """
    elif typ == "CMM":
        # dp.sql: (header_id, point_no, feature_kind, feature_name, max_val, min_val, mean_val, usl, lsl, result_ok)
        sql = f"""
          INSERT INTO {table} (header_id, point_no, feature_kind, feature_name, max_val, min_val, mean_val, usl, lsl, result_ok)
          VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
          ON DUPLICATE KEY UPDATE
            feature_kind=VALUES(feature_kind),
            feature_name=VALUES(feature_name),
            max_val=VALUES(max_val),
            min_val=VALUES(min_val),
            mean_val=VALUES(mean_val),
            usl=VALUES(usl),
            lsl=VALUES(lsl),
            result_ok=VALUES(result_ok)
        """
    else:
        raise ValueError(f"Unknown typ={typ}")

    # OMM 라벨 포맷 변경(FAI 라벨 원문 저장) 마이그레이션 대비:
    # - header_id 단위로 기존 결과 데이터를 먼저 지우고, 현재 파일 내용으로 다시 채운다.
    if typ == "OMM":
        header_ids = sorted({r[0] for r in rows if r and r[0]})
        if header_ids:
            ping_db(cur)
            for j in range(0, len(header_ids), CHUNK):
                chunk = header_ids[j:j+CHUNK]
                ph = ",".join(["%s"] * len(chunk))
                cur.execute(f"DELETE FROM {table} WHERE header_id IN ({ph})", tuple(chunk))

    # keep connection alive for big batches
    ping_db(cur)
    for i in range(0, len(rows), CHUNK):
        cur.executemany(sql, rows[i:i+CHUNK])

def compute_stats(vals: List[Optional[float]]) -> Tuple[Optional[float], Optional[float], Optional[float]]:
    xs = [v for v in vals if v is not None]
    if not xs:
        return (None, None, None)
    mx = max(xs)
    mn = min(xs)
    mean = sum(xs) / float(len(xs))
    return (mx, mn, mean)

def compute_ok(vals: List[Optional[float]], usl: Optional[float], lsl: Optional[float]) -> int:
    xs = [v for v in vals if v is not None]
    if not xs:
        return 1
    for v in xs:
        if usl is not None and v > usl:
            return 0
        if lsl is not None and v < lsl:
            return 0
    return 1


# --------------------------- scanning paths ---------------------------
def build_ipqc_dirs(base_root: str, model_folder: str, tool: str, ipqc_type: str,
                    y: int, m: int, mlabel: str) -> List[str]:
    # per v2
    tool_dir = f"{tool} Tool"
    if ipqc_type == "OMM":
        return [
            os.path.join(base_root, model_folder, tool_dir, "IPQC", "OMM", str(y), mlabel),
            os.path.join(base_root, model_folder, tool_dir, "IPQC", "OMM", mlabel),
        ]
    if ipqc_type == "CMM":
        return [
            os.path.join(base_root, model_folder, tool_dir, "IPQC", "OMM", str(y), mlabel),  # JMP: try OMM first
            os.path.join(base_root, model_folder, tool_dir, "IPQC", "CMM", str(y), mlabel),
            os.path.join(base_root, model_folder, tool_dir, "IPQC", "CMM", mlabel),
        ]
    if ipqc_type == "AOI":
        return [
            os.path.join(base_root, model_folder, tool_dir, "IPQC", "AOI", "성적서", str(y), mlabel),
            os.path.join(base_root, model_folder, tool_dir, "IPQC", "AOI", "성적서", mlabel),
        ]
    return []

@dataclass
class FileItem:
    model_disp: str
    model_folder: str
    tool: str
    abs_path: str
    rel_path: str  # stable key across base roots (best effort)

def _bundle_file_priority(filename: str) -> int:
    """Lower is better. Prefer normal OMM+AOI over OMM+AOI_1SHOT."""
    lname = str(filename or "").lower()
    if lname.endswith("omm+aoi.xlsx") or lname.endswith("omm+aoi.xlsm"):
        return 0
    if lname.endswith("omm+aoi_1shot.xlsx") or lname.endswith("omm+aoi_1shot.xlsm"):
        return 1
    return 99

def _dedupe_items_by_meas_date(items: List[FileItem]) -> List[FileItem]:
    """If multiple files have the same (model, tool, meas_date), keep only one.
    Rule:
      - prefer ...OMM+AOI.xlsx over ...OMM+AOI_1SHOT.xlsx
      - if same priority, keep newer mtime (fallback: larger size)
    """
    best: Dict[Tuple[str, str, str], FileItem] = {}
    keep_no_date: List[FileItem] = []

    for it in items:
        fname = os.path.basename(it.abs_path)
        d = parse_meas_date_from_filename(fname)
        if not d:
            keep_no_date.append(it)
            continue

        key = (it.model_disp, it.tool, d)
        prev = best.get(key)
        if prev is None:
            best[key] = it
            continue

        p_new = _bundle_file_priority(fname)
        p_old = _bundle_file_priority(os.path.basename(prev.abs_path))

        if p_new < p_old:
            best[key] = it
            continue
        if p_new > p_old:
            continue

        try:
            st_new = os.stat(it.abs_path)
            st_old = os.stat(prev.abs_path)
            if int(st_new.st_mtime) > int(st_old.st_mtime):
                best[key] = it
            elif int(st_new.st_mtime) == int(st_old.st_mtime) and int(st_new.st_size) > int(st_old.st_size):
                best[key] = it
        except Exception:
            pass

    merged = list(best.values()) + keep_no_date
    try:
        merged.sort(key=lambda x: (
            x.model_disp,
            x.tool,
            parse_meas_date_from_filename(os.path.basename(x.abs_path)) or "9999-99-99",
            os.path.basename(x.abs_path).lower(),
        ))
    except Exception:
        pass
    return merged

def collect_files_for_month(months: List[Tuple[int,int,str]]) -> List[FileItem]:
    base_candidates = get_base_src_candidates(OS_ENV)
    out: List[FileItem] = []
    seen = set()

    for (y, m, mlabel) in months:
        for model_disp, (_, model_folder) in MODEL_MAP.items():
            if TEST_MODELS and model_disp not in TEST_MODELS:
                continue
            tools = TOOL_MAP.get(model_disp, [])
            for tool in tools:
                if TEST_TOOLS and tool not in TEST_TOOLS:
                    continue
                for base_root in base_candidates:
                    if not os.path.isdir(base_root):
                        continue
                    for ipqc_type in SCAN_TYPES:
                        dirs = build_ipqc_dirs(base_root, model_folder, tool, ipqc_type, y, m, mlabel)
                        for d in dirs:
                            if not os.path.isdir(d):
                                continue
                            try:
                                names = os.listdir(d)
                            except Exception:
                                continue
                            if MAX_FILES_PER_DIR and MAX_FILES_PER_DIR > 0:
                                names = names[:MAX_FILES_PER_DIR]

                            for name in names:
                                if should_skip_file(name):
                                    continue
                                abs_path = os.path.join(d, name)
                                if not os.path.isfile(abs_path):
                                    continue

                                # stable-ish rel_path: strip base_root prefix
                                rel_path = abs_path
                                br = os.path.normpath(base_root)
                                ap = os.path.normpath(abs_path)
                                if ap.lower().startswith(br.lower()):
                                    rel_path = ap[len(br):].lstrip("\\/")

                                k = (model_disp, tool, rel_path)
                                if k in seen:
                                    continue
                                seen.add(k)
                                out.append(FileItem(model_disp, model_folder, tool, abs_path, rel_path))
    return _dedupe_items_by_meas_date(out)


# --------------------------- core ingest ---------------------------
def ingest_bundle_one_file(cur, item: FileItem, state: Dict[str, str]) -> bool:
    """
    Returns True if uploaded (or updated), False if skipped.
    """
    abs_path = item.abs_path
    fname = os.path.basename(abs_path)
    st = os.stat(abs_path)
    sig = f"{int(st.st_size)}:{int(st.st_mtime)}"
    skey = build_state_key(item.model_disp, item.tool, item.rel_path)
    legacy_key = build_legacy_state_key(item.rel_path)

    if state.get(skey) == sig or state.get(legacy_key) == sig:
        return False

    meas_date = parse_meas_date_from_filename(fname)
    if not meas_date:
        log(f"[SKIP] invalid/placeholder date in filename: {fname}")
        state[skey] = sig  # avoid endless retry; file rename will be treated as new
        state[legacy_key] = sig
        return False
    source_file = fname
    source_path = abs_path if STORE_SOURCE_PATH else None
    part_name = item.model_disp
    tool = item.tool
    omm_keyset = set()  # (fai, spc)
    cmm_keyset = set()  # point_no
    file_mtime = int(st.st_mtime)
    file_size = int(st.st_size)

    t0 = time.time()

    with zipfile.ZipFile(abs_path, "r") as zf:
        shared = read_shared_strings(zf)
        all_sheets = list_sheets(zf)

        cav_sheets = list_cav_sheets(all_sheets)
        if not cav_sheets:
            if INGEST_AOI:
                log(f"[WARN] no CAV sheets (AOI skip): {fname}")
            # OMM Raw Data 모드면 OMM은 계속 진행
            if not (INGEST_OMM and USE_OMM_RAWDATA):
                log(f"[SKIP] no CAV sheets: {fname}")
                state[skey] = sig  # treat as done to avoid endless retry
                state[legacy_key] = sig
                return False

                # --- parse CMM Raw data once (optional)
        cmm_rows_by_cav: Dict[int, List[Tuple[str, Optional[str], List[Optional[float]]]]] = {}
        cmm_parse_fail_by_cav: Dict[int, str] = {}
        cmm_sheet = find_cmm_sheet(all_sheets) if INGEST_CMM else None
        if cmm_sheet:
            cmm_name, cmm_xml = cmm_sheet

            cav_start_cols, ok, reason, cav_label_cols = detect_cmm_layout(zf, cmm_xml, shared, max_scan_row=25)

            if ok and cav_start_cols:
                # wanted cols: A,B,C + detected cavity CMM cols (3 each)
                wanted = set(["A","B","C"])
                for cav, start_col in cav_start_cols.items():
                    cols = [start_col, _next_col(start_col), _next_col(_next_col(start_col))]
                    wanted.update(cols)

                cmm_map = parse_sheet_rows(zf, cmm_xml, shared, start_row=3, wanted_cols=wanted)
                for row_no, cells in cmm_map.items():
                    point = cells.get("B")
                    if not point:
                        continue
                    p = str(point).strip()
                    if not p:
                        continue
                    cmm_keyset.add(p)
                    feat = cells.get("C")
                    for cav, start_col in cav_start_cols.items():
                        cols = [start_col, _next_col(start_col), _next_col(_next_col(start_col))]
                        vals = [safe_float(cells.get(c)) for c in cols]
                        if any(v is not None for v in vals):
                            cmm_rows_by_cav.setdefault(cav, []).append((p, feat, vals))
            else:
                # Layout unknown: DO NOT guess columns (it can corrupt DB).
                # Instead insert NULL placeholders to keep shape + log warning.
                warn = f"{reason or 'CMM layout not detected'}"
                print(f"[WARN] CMM layout detection failed: {source_file} / sheet='{cmm_name}' -> placeholders(NULL) will be inserted. ({warn})")

                # Determine cavities to keep (use detected cavity labels if any; else default 1..4)
                if cav_label_cols:
                    cavs = sorted([c for c in cav_label_cols.keys() if 1 <= c <= 4])
                    if not cavs:
                        cavs = [1,2,3,4]
                else:
                    cavs = [1,2,3,4]

                # Read only point_code(B) and feature(C) to build placeholder rows
                base_wanted = set(["B","C"])
                base_map = parse_sheet_rows(zf, cmm_xml, shared, start_row=3, wanted_cols=base_wanted)

                seen = set()
                points: List[Tuple[str, Optional[str]]] = []
                for _rn, cells in base_map.items():
                    point = cells.get("B")
                    if not point:
                        continue
                    p = str(point).strip()
                    if not p or p in seen:
                        continue
                    seen.add(p)
                    feat = cells.get("C")
                    points.append((p, feat))
                    cmm_keyset.add(p)

                for cav in cavs:
                    cmm_parse_fail_by_cav[cav] = warn
                    for p, feat in points:
                        cmm_rows_by_cav.setdefault(cav, []).append((p, feat, [None, None, None]))
# --- for each cavity sheet: parse OMM/AOI
        # We'll accumulate per (type,cavity) measurement rows & result rows.
        meas_rows: Dict[Tuple[str,int], List[Tuple]] = {}
        res_rows: Dict[Tuple[str,int], List[Tuple]] = {}
        header_ids: Dict[Tuple[str,int], int] = {}

        # --- OMM Raw Data (JMP Assist 기준) ---
        omm_rows_by_cav: Dict[int, List[Tuple[str, str, List[Optional[float]], Optional[float], Optional[float]]]] = {}
        if INGEST_OMM and USE_OMM_RAWDATA:
            ok_omm, omm_rows_by_cav, reason_omm = parse_omm_rawdata_xml_stream(zf, all_sheets, shared, item.model_disp)
            if not ok_omm:
                log(f"[WARN] OMM rawdata parse skipped: {fname} / {reason_omm}")
                omm_rows_by_cav = {}
            else:
                # OMM은 cavity별 header_id가 필요하므로 여기서 먼저 upsert+clear
                for cav, _rows in omm_rows_by_cav.items():
                    sk = make_source_key(tool, cav, source_file)
                    hid = upsert_header(cur, TABLES["OMM"]["header"], part_name, meas_date, tool, cav,
                                        source_file, source_path, sk, file_mtime, file_size)
                    header_ids[("OMM", cav)] = hid
                    clear_existing(cur, TABLES["OMM"]["meas"], TABLES["OMM"].get("res"), hid)


        for sheet_name, sheet_xml in cav_sheets:
            cav = parse_cavity_int(sheet_name)
            if not cav:
                continue
            row_map = parse_sheet_rows(zf, sheet_xml, shared, start_row=START_ROW, wanted_cols=WANTED_COLS)

            # Upsert headers (per cavity per type)
            if INGEST_OMM and (not USE_OMM_RAWDATA):
                sk = make_source_key(tool, cav, source_file)
                hid = upsert_header(cur, TABLES["OMM"]["header"], part_name, meas_date, tool, cav,
                                    source_file, source_path, sk, file_mtime, file_size)
                header_ids[("OMM", cav)] = hid
                clear_existing(cur, TABLES["OMM"]["meas"], TABLES["OMM"].get("res"), hid)
            if INGEST_AOI:
                sk = make_source_key(tool, cav, source_file)
                hid = upsert_header(cur, TABLES["AOI"]["header"], part_name, meas_date, tool, cav,
                                    source_file, source_path, sk, file_mtime, file_size)
                header_ids[("AOI", cav)] = hid
                clear_existing(cur, TABLES["AOI"]["meas"], TABLES["AOI"].get("res"), hid)

            for _, cells in row_map.items():
                point_no = cells.get("B")
                if not point_no:
                    continue
                point_no = str(point_no).strip()
                spc_code = cells.get("C")
                spc_code = str(spc_code).strip() if spc_code is not None else None
                if (INGEST_OMM and (not USE_OMM_RAWDATA)) and spc_code:
                    omm_keyset.add((point_no, spc_code))

                if INGEST_OMM and (not USE_OMM_RAWDATA):
                    usl = safe_float(cells.get(OMM_USL_COL))
                    lsl = safe_float(cells.get(OMM_LSL_COL))
                    vals = [safe_float(cells.get(c)) for c in OMM_DATA_COLS]
                    hid = header_ids[("OMM", cav)]
                    # measurements
                    for idx, v in enumerate(vals, start=1):
                        if v is None:
                            continue
                        meas_rows.setdefault(("OMM", cav), []).append((hid, point_no, spc_code, idx, v))
                    # results
                    mx, mn, mean = compute_stats(vals)
                    ok = compute_ok(vals, usl, lsl)
                    res_rows.setdefault(("OMM", cav), []).append((hid, point_no, mx, mn, mean, usl, lsl, ok))

                if INGEST_AOI:
                    usl = safe_float(cells.get(AOI_USL_COL))
                    lsl = safe_float(cells.get(AOI_LSL_COL))
                    vals = [safe_float(cells.get(c)) for c in AOI_DATA_COLS]
                    hid = header_ids[("AOI", cav)]
                    for idx, v in enumerate(vals, start=1):
                        if v is None:
                            continue
                        meas_rows.setdefault(("AOI", cav), []).append((hid, point_no, spc_code, idx, v))
                    mx, mn, mean = compute_stats(vals)
                    ok = compute_ok(vals, usl, lsl)
                    res_rows.setdefault(("AOI", cav), []).append((hid, point_no, mx, mn, mean, usl, lsl, ok))


        # --- OMM Raw Data rows -> DB rows (measurements/results) ---
        if INGEST_OMM and USE_OMM_RAWDATA and omm_rows_by_cav:
            for cav, rows in omm_rows_by_cav.items():
                hid = header_ids.get(("OMM", cav))
                if not hid:
                    sk = make_source_key(tool, cav, source_file)
                    hid = upsert_header(cur, TABLES["OMM"]["header"], part_name, meas_date, tool, cav,
                                        source_file, source_path, sk, file_mtime, file_size)
                    header_ids[("OMM", cav)] = hid
                    clear_existing(cur, TABLES["OMM"]["meas"], TABLES["OMM"].get("res"), hid)
                for point_no, spc_code, vals, usl, lsl in rows:
                    spc = (spc_code or "")
                    if point_no:
                        omm_keyset.add((point_no, spc))
                    for idx, v in enumerate(vals, start=1):
                        if v is None:
                            continue
                        meas_rows.setdefault(("OMM", cav), []).append((hid, point_no, spc, idx, v))
                    mx, mn, mean = compute_stats(vals)
                    ok = compute_ok(vals, usl, lsl)
                    res_rows.setdefault(("OMM", cav), []).append((hid, point_no, mx, mn, mean, usl, lsl, ok))

                # --- CMM per cavity insert
        if INGEST_CMM and cmm_rows_by_cav:
            for cav, rows in cmm_rows_by_cav.items():
                sk = make_source_key(tool, cav, source_file)
                hid = upsert_header(cur, TABLES["CMM"]["header"], part_name, meas_date, tool, cav,
                                    source_file, source_path, sk, file_mtime, file_size)
                header_ids[("CMM", cav)] = hid
                clear_existing(cur, TABLES["CMM"]["meas"], TABLES["CMM"].get("res"), hid)

                is_parse_fail = False
                fail_reason = None
                try:
                    # cmm_parse_fail_by_cav exists only in newer code path
                    is_parse_fail = cav in cmm_parse_fail_by_cav
                    fail_reason = cmm_parse_fail_by_cav.get(cav)
                except Exception:
                    is_parse_fail = False
                    fail_reason = None

                cmm_meas = []
                cmm_res = []
                for point_code, feat, vals in rows:
                    if is_parse_fail:
                        # keep shape: always write 3 rows with NULL values (no wrong numbers)
                        for idx in (1, 2, 3):
                            cmm_meas.append((hid, point_code, idx, None))
                        # mark parse failure in result table (if exists)
                        fk = "PARSE_FAIL"
                        fn = feat if feat is not None else (fail_reason or "CMM layout undetected")
                        cmm_res.append((hid, point_code, fk, fn, None, None, None, None, None, 0))
                    else:
                        for idx, v in enumerate(vals, start=1):
                            if v is None:
                                continue
                            cmm_meas.append((hid, point_code, idx, v))
                        mx, mn, mean = compute_stats(vals)
                        # CMM은 limit이 없으니 result_ok=1로 둠
                        cmm_res.append((hid, point_code, None, feat, mx, mn, mean, None, None, 1))

                # Measurements insert: allow NULL in parse-fail mode. If DB rejects NULL, degrade gracefully.
                try:
                    bulk_insert_measurements(cur, "CMM", TABLES["CMM"]["meas"], cmm_meas)
                except Exception as e:
                    if is_parse_fail:
                        print(f"[WARN] CMM placeholder(NULL) insert failed for {source_file} cav={cav}: {e}")
                    else:
                        raise

                if TABLES["CMM"].get("res"):
                    try:
                        bulk_insert_results(cur, "CMM", TABLES["CMM"]["res"], cmm_res)
                    except Exception:
                        # result 테이블이 없을 수 있음(선택)
                        pass
# --- keymap upsert (pivot key cache)
        if INGEST_OMM and ENABLE_KEYMAP_OMM and omm_keyset:
            insert_keymap_omm(cur, part_name, sorted(list(omm_keyset)))
        if INGEST_CMM and ENABLE_KEYMAP_CMM and cmm_keyset:
            insert_keymap_cmm(cur, part_name, sorted(list(cmm_keyset)))

        # --- bulk insert OMM/AOI
        if INGEST_OMM:
            for (typ, cav), rows in meas_rows.items():
                if typ != "OMM":
                    continue
                bulk_insert_measurements(cur, "OMM", TABLES["OMM"]["meas"], rows)
            for (typ, cav), rows in res_rows.items():
                if typ != "OMM":
                    continue
                bulk_insert_results(cur, "OMM", TABLES["OMM"]["res"], rows)

        if INGEST_AOI:
            for (typ, cav), rows in meas_rows.items():
                if typ != "AOI":
                    continue
                bulk_insert_measurements(cur, "AOI", TABLES["AOI"]["meas"], rows)
            for (typ, cav), rows in res_rows.items():
                if typ != "AOI":
                    continue
                bulk_insert_results(cur, "AOI", TABLES["AOI"]["res"], rows)

    elapsed = time.time() - t0
    state[skey] = sig
    state[legacy_key] = sig
    log(f"[OK] {fname} / {part_name} / tool={tool} / {meas_date or '-'} / {elapsed:.2f}s")
    return True


def _next_col(col: str) -> str:
    """Excel column increment: A->B, Z->AA"""
    col = col.strip().upper()
    n = 0
    for ch in col:
        n = n * 26 + (ord(ch) - 64)
    n += 1
    # back to letters
    out = ""
    while n > 0:
        n, rem = divmod(n - 1, 26)
        out = chr(65 + rem) + out
    return out


# --------------------------- retry wrapper ---------------------------
def with_retries(fn, *args, **kwargs):
    for attempt in range(MAX_RETRIES):
        try:
            return fn(*args, **kwargs)
        except Exception as e:
            if attempt == MAX_RETRIES - 1:
                raise
            sleep = RETRY_BASE_SLEEP * (2 ** attempt) + random.random() * 0.2
            log(f"[RETRY] {e} (sleep {sleep:.2f}s)")
            time.sleep(sleep)


# --------------------------- main loop ---------------------------
def choose_months() -> List[Tuple[int,int,str]]:
    if RUN_MODE == "TEST":
        return iter_months_test()
    if RUN_MODE == "HOURLY":
        return iter_months_hourly()
    return iter_months_backfill()

def run_once(state: Dict[str, str]) -> None:
    months = choose_months()
    items = collect_files_for_month(months)
    log(f"[MODE] {RUN_MODE} months={len(months)} files={len(items)} scan_types={SCAN_TYPES} bundle=OMM+AOI+CMM")

    if not items:
        return

    conn = connect_db()
    cur = conn.cursor()
    set_session_tuning(cur)
    ensure_keymap_tables(cur)

    ok = 0
    skipped = 0

    try:
        for item in items:
            fname = os.path.basename(item.abs_path)
            try:
                changed = ingest_bundle_one_file(cur, item, state)
                if changed:
                    conn.commit()
                    save_state(STATE_FILE, state)
                    ok += 1
                else:
                    skipped += 1

            except RETRYABLE_DB_ERR as e:
                # connection likely dropped mid-write; rollback may also fail
                safe_rollback(conn)
                log(f"[DB-RETRY] {fname} : {e} -> reconnect & retry once")

                try:
                    conn, cur = reconnect_db(conn)
                except Exception as e2:
                    log(f"[DB-RETRY-FAIL] reconnect failed: {e2}")
                    continue

                try:
                    changed = ingest_bundle_one_file(cur, item, state)
                    if changed:
                        conn.commit()
                        save_state(STATE_FILE, state)
                        ok += 1
                    else:
                        skipped += 1
                except Exception as e3:
                    safe_rollback(conn)
                    log(f"[ERR] {fname} (after retry) : {e3}")
                    continue

            except Exception as e:
                safe_rollback(conn)
                log(f"[ERR] {fname} : {e}")
                continue

        log(f"[DONE] ok={ok} skipped={skipped} total={len(items)}")
    finally:
        safe_close(cur)
        safe_close(conn)
def main():
    # startup banner (for unified XAMPP-like console)
    log("=== START ===")
    log(f"DB={DB_HOST}:{DB_PORT}/{DB_NAME} user={DB_USER}")
    log(f"RUN_MODE={RUN_MODE} interval={LOOP_INTERVAL_SEC}s scan_types={SCAN_TYPES} bundle=OMM+AOI+CMM")
    log(f"STATE_FILE={STATE_FILE.resolve()}")
    log(f"LOG_FILE={LOG_FILE.resolve()}")
    state = load_state(STATE_FILE)
    while True:
        run_once(state)
        if LOOP_INTERVAL_SEC and LOOP_INTERVAL_SEC > 0:
            time.sleep(LOOP_INTERVAL_SEC)
        else:
            break

if __name__ == "__main__":
    main()