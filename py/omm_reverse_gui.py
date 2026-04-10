# -*- coding: utf-8 -*-
import csv
import datetime
import re
import traceback
from collections import OrderedDict, defaultdict

import tkinter as tk
from tkinter import ttk, filedialog, messagebox

import pymysql

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    OPENPYXL_OK = True
except Exception:
    OPENPYXL_OK = False

APP_TITLE = "OMM 역추출기 (Rough GUI)"

# DB 설정은 코드에만 둠
DB_HOSTS = ["dpams.ddns.net", "211.212.182.110"]
DB_PORT = 3306
DB_USER = "maya"
DB_PASS = "##Gmlakd2323"
DB_NAME = "dp"

MODEL_OPTIONS = [
    "ALL",
    "Memphis IR BASE",
    "Memphis X Carrier",
    "Memphis Y Carrier",
    "Memphis Z Carrier",
    "Memphis Z Stopper",
]

BASE_COLUMNS = ["fai", "usl", "lsl"]

IR_BASE_FAI_ORDER_TEXT = 'FAI 1 / SPC A\nFAI 3-1 / SPC AL\nFAI 3-2 / SPC AL\nFAI 4 / SPC C\nFAI 5\nFAI 6-1 / SPC AM\nFAI 6-2 / SPC AM\nFAI 17-1 / SPC U \nFAI 17-2/ SPC U \nFAI 17-3 /SPC U \nFAI 17 (틱契똑)\nFAI 18 / SPC AB\nFAI 19 / SPC AC\nFAI 20-1 SPC AD\nFAI 20-2 SPC AD\nFAI 20-3 SPC AD\nFAI 22-1\nFAI 22-2\nFAI 22-3\nFAI 22-4\nFAI 22-5\nFAI 150_\nFAI149-1\nFAI149-2\nFAI 151-1(S8)\nFAI 151-1(S9)\nFAI 151\nFAI 23-1 \nFAI 23-2 \nFAI 25-1\nFAI 25-2\nFAI 25\nFAI 26\nFAI 27.1 / SPC Z\nFAI 27.2 / SPC Z\nFAI 28-1\nFAI 28-2\nFAI 29 / SPC AF\nFAI 30 / SPC T\nFAI 31 / SPC AE\nFAI 32-1 /SPC W (DC)\nFAI 32-2 /SPC W (DC)\nFAI 33-1\nFAI 33-2\nFAI 35-1 / SPC Y\nFAI 35-2 / SPC Y\nFAI 37-1(U) / SPC X\nFAI 37-2(M) / SPC X\nFAI 37-3(D) / SPC X\nFAI 88-1(U) / SPC AP\nFAI 88-2(D) / SPC AP\nFAI 39-1\nFAI 39-2\nFAI 40-1(L) / SPC AK\nFAI 40-2(M) / SPC AK\nFAI 40-3(R) / SPC AK\nFAI 41-1 / SPC AH\nFAI 41-2 / SPC AH\nFAI 41-3 / SPC AH\nFAI 41-4 / SPC AH\nFAI 43\nFAI 46-1\nFAI 46-2\nFAI 82-1 / SPC V\nFAI 82-2 / SPC V\nFAI 82-3 / SPC V\nFAI 82-4 / SPC V\nFAI 82-5 / SPC V\nFAI 82-6 / SPC V\nFAI 38-1/ SPC AG\nFAI 38-2/ SPC AG\nFAI 89-1\nFAI 89-2\ne1\ne2\ne3\ne4\ne5\ne6\ne7\ne8\ne9\ne10\ne11\ne12\ne13\ne14\ne15\nFAI94-1\nFAI94-2\nFAI94-3\nFAI94-4\nFAI94-5\nFAI94-6\nFAI94-7\nFAI94-8\nFAI94-9\nFAI94-10\nFAI94-11\nFAI94-12\nFAI94-13\nFAI94-14\nFAI94-15\nFAI94-16\nFAI94-17\nFAI94-18\nFAI94-19\nFAI94-20\nFAI94-21\nFAI153\nFAI 154\nFAI 155\nFAI 97 / SPC AV\nFAI 97-1 / SPC AV\nFAI 97-2 / SPC AV\nFAI 97-3 / SPC AV\nFAI 102\nFAI 103\nFAI 107-1\nFAI 107-2\nFAI 107-3\nFAI 111\nFAI 140/SPC BH\nFAI 141/SCP BI\nG1X\nG1Y\nG1X1\nG1Y1\nG1X1 Position\nG2X\nG2Y\nG2X1\nG2Y1\nG2X1 Position\nG3X1\nG3Y1\nG3X1\nG3Y1\nG3X1 Position.\nG4XX\nG4YY\nG4XX\nG4YY\nG4X1 Position\nG5XX\nG5YY\nG5X1\nG5Y1\nG5X1 Position\nG6XX\nG6YY\nG6X1\nG6Y1\nG6X1 Position\nG7X1\nG7Y1\nG7X1\nG7Y1\nG7X1 Position\nG8XX\nG8YY\nG8XX1\nG8Y1\nG8X1 Position\nG9XX\nG9YY\nG9X1\nG9Y1\nG9X1 Position\nFAI143\nG10XX\nG10YY\nG10X1\nG10Y1\nG10X1 Position\nG11XX\nG11YY\nG11X1\nG11Y11\nG11X1 Position\nG12X\nG12Y\nG12X1\nG12Y1\nG12X1 Position\nG13XX\nG13YY\nG13XX1\nG13YY1\nG13X1 Position\nG14XX\nG14YY\nG14X1\nG14Y1\nG14X1 Position\nG15X\nG15Y\nG15X1\nG15Y1\nG15X1 Position\nG16X\nG16Y\nG16X1\nG16Y1\nG16X1 Position\nG17X\nG17Y\nG17X1\nG17Y1\nG17X1 Position\nG18X\nG18Y\nG18X1\nG18Y1\nG18X1 Position\nFAI 144\nFAI42-1/SPC AN(DC)\nFAI42-2/SPC AN(DC)\nFAI42-3/SPC AN(DC)\nFAI42-4/SPC AN(DC)\nFAI45-1/SPC AJ(DC)\nFAI45-2/SPC AJ(DC)\nFAI45-3/SPC AJ(DC)\nFAI45-4/SPC AJ(DC)\nFAI152-1 (S6)\nFAI152-1 (S7)\nFAI152\nFAI 157 /SPC CA\nFAI 158 /SPC CB\nFAI98\nFAI99\nFAI 100_\nFAI 9/SPC F\nFAI 9.1\nFAI 9.2\nFAI 9.3\nFAI 9.4\nFAI 9.5\nFAI 9.6\nFAI 9.7\nFAI 10/SPC G\nFAI 10.01\nFAI 10.02\nFAI 10.03\nFAI 10.04\nFAI 10.05\nFAI 10.06\nFAI 10.07\nFAI 10.08\nFAI 10.09\nFAI 10.10_\nFAI 10.11\nFAI 10.12\nFAI 10.13\nFAI 71-1/SPC AS\nFAI 71-2/SPC AS\nFAI 72-1/SPC AT\nFAI 72-2/SPC AT\nFAI 73-1/SPC AU\nFAI 73-2/SPC AU\nFAI 74-1\nFAI 74-2\nFAI 75-1\nFAI 75-2\nFAI 76\nFAI 77\nFAI 78-1\nFAI 78-2\nFAI 126/SPC BA （BOTTOM）\nFAI 32-1/SPC W\nFAI 32-2/SPC W\nFAI 42-1/SPC AN （BOTTOM）\nFAI 42-2/SPC AN （BOTTOM）\nFAI 42-3/SPC AN （BOTTOM）\nFAI 45-1/SPC AJ \nFAI 45-2/SPC AJ \nFAI 45-3/SPC AJ \nFAI 127-1/SPC BB \nFAI 127-2/SPC BB \nFAI 128-1/SPC BC \nFAI 128-2/SPC BC \nFAI 129-1/SPC BD \nFAI 129-2/SPC BD \nFAI 130-1/SPC BE\nFAI 130-2/SPC BE\nFAI 133-F1\nFAI 133-F2\nFAI 133-F3\nFAI 133-F4\nF1튤뀌\nF2튤뀌\nF3튤뀌\nF4튤뀌\nFAI 133\nFAI 135-G1\nFAI 135-G2\nFAI 135-G3\nFAI 135 離댕튤뀌\nFAI 135\nFAI 136-1\nFAI 136-2\nFAI 136-I3\nFAI 137-H1\nFAI 137-H2\nFAI 137-H3\nFAI 137-H4\nFAI 137-H5\nFAI 137\nFAI 138-J1\nFAI 138-J2\nFAI 138-J3\nFAI 139-L6\nFAI 139-L5\nFAI 139-L4\nFAI 139-L3\nFAI 139-L2\nFAI 139-L1\nFAI 139 離댕튤뀌\nFAI 139\nFAI 148-P1\nFAI 148-P2\nFAI 148\nFAI 13\nFAI  14-1\n60-X\nFAI 60_\nFAI  61-1\nFAI  61-2\nFAI  61-3\n70-Y\nFAI  70_\nFAI  87-1\nFAI  87-2\nFAI  87-3\nFAI  109-1\nFAI  109-2\nFAI  124\nFAI  145 居스\n145 居스-P1\n145 居스-P2\n145 居스-P3\n145 居스-P4\n145 居스-P5\n145 居스-P6\n145 居스-P7\n145 居스-P8Z\nFAI  145 똥綾\n145 똥綾-P9\n145 똥綾-P10_\n145 똥綾-P11\n145 똥綾-P12\n145 똥綾-P13\n145 똥綾-P14\n145 똥綾-P15\n145 똥綾-P16\n145 똥綾-P17\n145 똥綾-P18\n145 똥綾-P19\n145 똥綾-P20_\n145 똥綾-P21\n145 똥綾-P22\n145 똥綾-P23\n145 똥綾-P24\nFAI  145\n145-P1\n145-P2\n145-P3\n145-P4\n145-P5\n145-P6\n145-P7\n145-P8\n145-P9\n145-P10_\n145-P11\n145-P12\n145-P13\n145-P14\n145-P15\n145-P16\n145-P17\n145-P18\n145-P19\n145-P20_\n145-P21\n145-P22\n145-P23\n145-P24\nFAI 14-2\nFAI 101-4\nFAI 131/SPC BF\n132-X\n132-Y\nFAI 132/SPC BG\nFAI 65-1\nFAI 65-2\nFAI 65-3\n66-X\nFAI 66\nFAI 69-1\nFAI 69-2\nFAI 69-3\n86-Y\nFAI 86\nFAI 110-1\nFAI 110-2\nFAI 125\nFAI146居스\n146居스-P1\n146居스-P2\n146居스-P3\n146居스-P4\n146居스-P5\n146居스-P6\n146居스-P7\n146居스-P8\nFAI 146똥綾\n146똥綾-P9\n146똥綾-P10_\n146똥綾-P11\n146똥綾-P12\n146똥綾-P13\n146똥綾-P14\n146똥綾-P15\n146똥綾-P16\n146똥綾-P17\n146똥綾-P18\n146똥綾-P19\n146똥綾-P20_\n146똥綾-P21\nFAI 146\n146-P1\n146-P2\n146-P3\n146-P4\n146-P5\n146-P6\n146-P7\n146-P8\n146-P9\n146-P10_\n146-P11\n146-P12\n146-P13\n146-P14\n146-P15\n146-P16\n146-P17\n146-P18\n146-P19\n146-P20_\n146-P21\n147-P1\n147-P2\n147-P3\n147-P4\n147-P5\n147-P6\n147-P7\n147-P8\nFAI 147\n14-2(DC)\n132-X(DC)\n132-Y(DC)\nFAI 132(DC)\n86-Y(DC)\nFAI 86(DC)\n2점각도\n3점각도\nFAI 101-1\nFAI 101-2\nFAI 101-3\n63-Y\nFAI 63\nFAI 64-1\nFAI 64-2\nFAI 64-3\n67-X\nFAI 67\nFAI 68-1\nFAI 68-2\nFAI 68-3\nFAI 108-1\nFAI 108-2\nFAI105_1\nFAI105_2\nFAI105_3\nFAI106_1\nFAI106_2\nFAI106_3'


def chunked(items, size):
    i = 0
    while i < len(items):
        yield items[i:i + size]
        i += size


def canonical_fai_label(text):
    s = str(text or "")
    s = s.replace("\u3000", " ").strip()
    s = re.sub(r"\s+", " ", s)
    s = re.sub(r"\s*/\s*", "/", s)
    s = re.sub(r"\s*\.\s*", ".", s)
    s = re.sub(r"\s*\(\s*", "(", s)
    s = re.sub(r"\s*\)\s*", ")", s)
    return s.upper()


def build_fai_order_map(multiline_text):
    order_map = OrderedDict()
    for raw in (multiline_text or "").splitlines():
        line = str(raw or "").strip()
        if not line:
            continue
        key = canonical_fai_label(line)
        if key and key not in order_map:
            order_map[key] = len(order_map)
    return order_map


FAI_ORDER_MAP_BY_MODEL = {
    "Memphis IR BASE": build_fai_order_map(IR_BASE_FAI_ORDER_TEXT),
}


class OMMReverseService(object):
    def __init__(self, hosts, port, user, password, database):
        self.hosts = [str(x).strip() for x in hosts if str(x).strip()]
        self.port = int(port)
        self.user = user
        self.password = password
        self.database = database

    def connect(self):
        last_exc = None
        for host in self.hosts:
            try:
                return pymysql.connect(
                    host=host,
                    port=self.port,
                    user=self.user,
                    password=self.password,
                    database=self.database,
                    charset="utf8mb4",
                    autocommit=True,
                    cursorclass=pymysql.cursors.DictCursor,
                    connect_timeout=5,
                    read_timeout=60,
                    write_timeout=60,
                )
            except Exception as exc:
                last_exc = exc
        raise RuntimeError("DB 연결 실패: %s" % last_exc)

    def fetch_headers(self, start_date, end_date, model, selected_tcs):
        tool_set = []
        cav_set = []
        for tc in selected_tcs:
            if "#" not in tc:
                continue
            tool, cav = tc.split("#", 1)
            tool = tool.strip().upper()
            cav = cav.strip()
            if not tool or not cav.isdigit():
                continue
            if tool not in tool_set:
                tool_set.append(tool)
            cav_i = int(cav)
            if cav_i not in cav_set:
                cav_set.append(cav_i)

        if not tool_set or not cav_set:
            return []

        sql = []
        params = []
        sql.append("SELECT id, meas_date, part_name, tool, cavity")
        sql.append("FROM ipqc_omm_header")
        sql.append("WHERE meas_date BETWEEN %s AND %s")
        params.append(start_date)
        params.append(end_date)

        if model != "ALL":
            sql.append("AND part_name = %s")
            params.append(model)

        sql.append("AND tool IN (%s)" % ",".join(["%s"] * len(tool_set)))
        params.extend(tool_set)
        sql.append("AND cavity IN (%s)" % ",".join(["%s"] * len(cav_set)))
        params.extend(cav_set)
        sql.append("ORDER BY meas_date, tool, cavity, id")

        out = []
        conn = self.connect()
        try:
            cur = conn.cursor()
            try:
                cur.execute("\n".join(sql), params)
                for row in cur.fetchall():
                    out.append({
                        "header_id": int(row.get("id") or 0),
                        "meas_date": str(row.get("meas_date") or ""),
                        "part_name": str(row.get("part_name") or ""),
                        "tool": str(row.get("tool") or "").upper(),
                        "cavity": int(row.get("cavity") or 0),
                    })
            finally:
                cur.close()
        finally:
            conn.close()
        return out

    def fetch_wide_rows(self, headers, selected_tcs, model_name):
        if not headers:
            empty_meta = {
                "date_order": [],
                "preview_columns": list(BASE_COLUMNS),
                "xlsx_header_row1": list(BASE_COLUMNS),
                "xlsx_header_row2": list(BASE_COLUMNS),
                "merge_ranges": [],
                "date_tc_to_header": OrderedDict(),
            }
            return list(BASE_COLUMNS), [], empty_meta

        date_order = []
        for h in headers:
            d = str(h.get("meas_date") or "")
            if d and d not in date_order:
                date_order.append(d)
        date_order.sort()

        date_tc_to_header = OrderedDict()
        for meas_date in date_order:
            date_tc_to_header[meas_date] = OrderedDict((tc, None) for tc in selected_tcs)

        # 같은 날짜 + Tool#Cavity 조합이 여러 건이면 가장 마지막 header(id 큰 것)를 사용
        for h in headers:
            meas_date = str(h.get("meas_date") or "")
            tc = "%s#%s" % (h.get("tool"), h.get("cavity"))
            if meas_date in date_tc_to_header and tc in date_tc_to_header[meas_date]:
                date_tc_to_header[meas_date][tc] = h

        chosen_headers = []
        seen_hids = set()
        for meas_date in date_order:
            tc_map = date_tc_to_header.get(meas_date) or {}
            for tc in selected_tcs:
                h = tc_map.get(tc)
                if h is None:
                    continue
                hid = int(h.get("header_id") or 0)
                if hid and hid not in seen_hids:
                    chosen_headers.append(h)
                    seen_hids.add(hid)

        if not chosen_headers:
            empty_meta = {
                "date_order": date_order,
                "preview_columns": list(BASE_COLUMNS),
                "xlsx_header_row1": list(BASE_COLUMNS),
                "xlsx_header_row2": list(BASE_COLUMNS),
                "merge_ranges": [],
                "date_tc_to_header": date_tc_to_header,
            }
            return list(BASE_COLUMNS), [], empty_meta

        header_ids = [int(h["header_id"]) for h in chosen_headers]
        result_by_key = {}
        meas_by_key = defaultdict(dict)
        fai_seen_order = OrderedDict()

        conn = self.connect()
        try:
            cur = conn.cursor()
            try:
                for chunk in chunked(header_ids, 500):
                    placeholders = ",".join(["%s"] * len(chunk))

                    cur.execute(
                        "SELECT header_id, fai, usl, lsl FROM ipqc_omm_result WHERE header_id IN (%s) ORDER BY header_id, fai" % placeholders,
                        chunk,
                    )
                    for row in cur.fetchall():
                        hid = int(row.get("header_id") or 0)
                        fai = str(row.get("fai") or "")
                        if fai and fai not in fai_seen_order:
                            fai_seen_order[fai] = len(fai_seen_order)
                        result_by_key[(hid, fai)] = {
                            "usl": row.get("usl"),
                            "lsl": row.get("lsl"),
                        }

                    cur.execute(
                        "SELECT header_id, fai, row_index, value FROM ipqc_omm_measurements WHERE header_id IN (%s) ORDER BY header_id, fai, row_index" % placeholders,
                        chunk,
                    )
                    for row in cur.fetchall():
                        hid = int(row.get("header_id") or 0)
                        fai = str(row.get("fai") or "")
                        idx = int(row.get("row_index") or 0)
                        val = row.get("value")
                        if fai and fai not in fai_seen_order:
                            fai_seen_order[fai] = len(fai_seen_order)
                        meas_by_key[(hid, fai)][idx] = val
                        if (hid, fai) not in result_by_key:
                            result_by_key[(hid, fai)] = {"usl": None, "lsl": None}
            finally:
                cur.close()
        finally:
            conn.close()

        template_order_map = FAI_ORDER_MAP_BY_MODEL.get(model_name, {})
        ordered_fais = list(fai_seen_order.keys())
        ordered_fais.sort(key=lambda fai: (
            0 if canonical_fai_label(fai) in template_order_map else 1,
            template_order_map.get(canonical_fai_label(fai), 10**9),
            fai_seen_order.get(fai, 10**9),
        ))

        preview_columns = list(BASE_COLUMNS)
        xlsx_header_row1 = ["", "", ""]
        xlsx_header_row2 = list(BASE_COLUMNS)
        merge_ranges = []
        excel_col_idx = 4

        for meas_date in date_order:
            block_cols = []
            for tc in selected_tcs:
                for idx in (1, 2, 3):
                    col_name = "%s|%s_%s" % (meas_date, tc, idx)
                    preview_columns.append(col_name)
                    xlsx_header_row1.append(meas_date)
                    xlsx_header_row2.append("%s_%s" % (tc, idx))
                    block_cols.append(col_name)
            if block_cols:
                start_col = excel_col_idx
                end_col = excel_col_idx + len(block_cols) - 1
                merge_ranges.append((1, start_col, 1, end_col, meas_date))
                excel_col_idx = end_col + 1

        rows = []
        for fai in ordered_fais:
            row = {"fai": fai, "usl": None, "lsl": None}
            for meas_date in date_order:
                tc_map = date_tc_to_header.get(meas_date) or {}
                for tc in selected_tcs:
                    chosen = tc_map.get(tc)
                    key_base = "%s|%s" % (meas_date, tc)
                    if chosen is None:
                        row[key_base + "_1"] = None
                        row[key_base + "_2"] = None
                        row[key_base + "_3"] = None
                        continue

                    hid = int(chosen.get("header_id") or 0)
                    result_info = result_by_key.get((hid, fai), {})
                    if row.get("usl") is None and result_info.get("usl") is not None:
                        row["usl"] = result_info.get("usl")
                    if row.get("lsl") is None and result_info.get("lsl") is not None:
                        row["lsl"] = result_info.get("lsl")

                    idx_vals = meas_by_key.get((hid, fai), {})
                    row[key_base + "_1"] = idx_vals.get(1)
                    row[key_base + "_2"] = idx_vals.get(2)
                    row[key_base + "_3"] = idx_vals.get(3)
            rows.append(row)

        export_meta = {
            "date_order": date_order,
            "preview_columns": preview_columns,
            "xlsx_header_row1": xlsx_header_row1,
            "xlsx_header_row2": xlsx_header_row2,
            "merge_ranges": merge_ranges,
            "date_tc_to_header": date_tc_to_header,
        }
        return preview_columns, rows, export_meta


class App(tk.Tk):
    def __init__(self):
        tk.Tk.__init__(self)
        self.title(APP_TITLE)
        self.geometry("1600x860")
        self.minsize(1200, 720)

        self.columns = []
        self.rows = []
        self.export_meta = {
            "date_order": [],
            "preview_columns": [],
            "xlsx_header_row1": [],
            "xlsx_header_row2": [],
            "merge_ranges": [],
        }

        today = datetime.date.today().strftime("%Y-%m-%d")
        self.start_date_var = tk.StringVar(value=today)
        self.end_date_var = tk.StringVar(value=today)
        self.model_var = tk.StringVar(value="ALL")
        self.tools_var = tk.StringVar(value="A")
        self.cavities_var = tk.StringVar(value="1-4")
        self.status_var = tk.StringVar(value="대기 중")

        self._build_ui()
        self._fill_defaults()

    def _build_ui(self):
        self.columnconfigure(0, weight=1)
        self.rowconfigure(1, weight=1)

        filter_frame = ttk.LabelFrame(self, text="OMM 조회 조건")
        filter_frame.grid(row=0, column=0, sticky="ew", padx=8, pady=(8, 4))
        for c in range(12):
            filter_frame.columnconfigure(c, weight=1)

        add_labeled_entry(filter_frame, "시작일", self.start_date_var, 0, 0, 12)
        add_labeled_entry(filter_frame, "종료일", self.end_date_var, 0, 2, 12)

        ttk.Label(filter_frame, text="모델").grid(row=0, column=4, sticky="w", padx=(8, 4), pady=6)
        model_cb = ttk.Combobox(filter_frame, textvariable=self.model_var, values=MODEL_OPTIONS, state="readonly", width=24)
        model_cb.grid(row=0, column=5, sticky="ew", padx=(0, 8), pady=6)

        add_labeled_entry(filter_frame, "툴", self.tools_var, 0, 8, 18)
        add_labeled_entry(filter_frame, "캐비티", self.cavities_var, 0, 10, 12)

        tc_frame = ttk.LabelFrame(self, text="Tool#Cavity 설정")
        tc_frame.grid(row=1, column=0, sticky="nsew", padx=8, pady=4)
        tc_frame.columnconfigure(0, weight=0)
        tc_frame.columnconfigure(1, weight=1)
        tc_frame.rowconfigure(1, weight=1)

        helper_frame = ttk.Frame(tc_frame)
        helper_frame.grid(row=0, column=0, columnspan=2, sticky="ew", padx=6, pady=(6, 4))
        helper_frame.columnconfigure(1, weight=1)

        ttk.Label(helper_frame, text="FAI명은 세로 유지 / 날짜별로 Tool#Cavity 3칸씩 오른쪽 누적").grid(row=0, column=0, sticky="w")
        ttk.Button(helper_frame, text="툴/캐비티로 목록 생성", command=self.generate_tc).grid(row=0, column=1, sticky="e", padx=4)
        ttk.Button(helper_frame, text="조회", command=self.run_query).grid(row=0, column=2, sticky="e", padx=4)
        ttk.Button(helper_frame, text="CSV 저장", command=self.export_csv).grid(row=0, column=3, sticky="e", padx=4)
        ttk.Button(helper_frame, text="엑셀 저장", command=self.export_xlsx).grid(row=0, column=4, sticky="e", padx=4)

        left_box = ttk.Frame(tc_frame)
        left_box.grid(row=1, column=0, sticky="nsw", padx=(6, 3), pady=(0, 6))
        ttk.Label(left_box, text="Tool#Cavity 목록\n(쉼표 또는 줄바꿈 구분)").pack(anchor="w")
        self.tc_text = tk.Text(left_box, width=30, height=10, wrap="word")
        self.tc_text.pack(fill="y", expand=False, pady=(4, 0))

        preview_frame = ttk.Frame(tc_frame)
        preview_frame.grid(row=1, column=1, sticky="nsew", padx=(3, 6), pady=(0, 6))
        preview_frame.columnconfigure(0, weight=1)
        preview_frame.rowconfigure(0, weight=1)

        self.tree = ttk.Treeview(preview_frame, show="headings")
        self.tree.grid(row=0, column=0, sticky="nsew")
        ysb = ttk.Scrollbar(preview_frame, orient="vertical", command=self.tree.yview)
        xsb = ttk.Scrollbar(preview_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=ysb.set, xscrollcommand=xsb.set)
        ysb.grid(row=0, column=1, sticky="ns")
        xsb.grid(row=1, column=0, sticky="ew")

        status_bar = ttk.Frame(self)
        status_bar.grid(row=2, column=0, sticky="ew", padx=8, pady=(0, 8))
        status_bar.columnconfigure(0, weight=1)
        ttk.Label(status_bar, textvariable=self.status_var).grid(row=0, column=0, sticky="w")

    def _fill_defaults(self):
        self.tc_text.delete("1.0", "end")
        self.tc_text.insert("1.0", "A#1, A#2, A#3, A#4")

    def make_service(self):
        return OMMReverseService(DB_HOSTS, DB_PORT, DB_USER, DB_PASS, DB_NAME)

    def generate_tc(self):
        try:
            tools = parse_tools(self.tools_var.get())
            cavities = parse_cavities(self.cavities_var.get())
            tc_list = []
            for tool in tools:
                for cav in cavities:
                    tc_list.append("%s#%s" % (tool, cav))
            self.tc_text.delete("1.0", "end")
            self.tc_text.insert("1.0", ", ".join(tc_list))
            self.status_var.set("Tool#Cavity %s개 생성" % len(tc_list))
        except Exception as exc:
            messagebox.showerror(APP_TITLE, str(exc))

    def run_query(self):
        try:
            start_date = normalize_date(self.start_date_var.get().strip())
            end_raw = self.end_date_var.get().strip()
            if not end_raw:
                end_raw = self.start_date_var.get().strip()
            end_date = normalize_date(end_raw)
            selected_tcs = parse_tc_text(self.tc_text.get("1.0", "end"))
            if not selected_tcs:
                raise ValueError("Tool#Cavity 목록이 비어 있음")

            self.status_var.set("조회 중...")
            self.update_idletasks()

            svc = self.make_service()
            model_name = self.model_var.get().strip() or "ALL"
            headers = svc.fetch_headers(start_date, end_date, model_name, selected_tcs)
            columns, rows, export_meta = svc.fetch_wide_rows(headers, selected_tcs, model_name)

            self.columns = columns
            self.rows = rows
            self.export_meta = export_meta
            self.refresh_tree(columns, rows)

            used_date_tc_count = 0
            for meas_date in export_meta.get("date_order", []):
                tc_map = export_meta.get("date_tc_to_header", {}).get(meas_date, {})
                for tc in selected_tcs:
                    if tc_map.get(tc) is not None:
                        used_date_tc_count += 1

            self.status_var.set(
                "조회 완료: header %s건 / 날짜 %s개 / 사용 날짜×Tool#Cavity %s개 / FAI 행 %s건" % (
                    len(headers),
                    len(export_meta.get("date_order", [])),
                    used_date_tc_count,
                    len(rows),
                )
            )
        except Exception as exc:
            self.status_var.set("오류")
            messagebox.showerror(APP_TITLE, "조회 실패\n\n%s\n\n%s" % (exc, traceback.format_exc()))

    def refresh_tree(self, columns, rows):
        self.tree.delete(*self.tree.get_children())
        self.tree["columns"] = columns

        for col in columns:
            width = 110
            if col == "fai":
                width = 240
            elif col in ("usl", "lsl"):
                width = 90
            elif "|" in col:
                width = 96
            self.tree.heading(col, text=col)
            self.tree.column(col, width=width, anchor="center")

        preview_limit = 500
        for row in rows[:preview_limit]:
            values = [format_cell(row.get(col), "") for col in columns]
            self.tree.insert("", "end", values=values)

        if len(rows) > preview_limit:
            self.status_var.set("미리보기는 %s행까지만 표시됨 / 전체 %s행" % (preview_limit, len(rows)))

    def export_csv(self):
        if not self.columns or not self.rows:
            messagebox.showwarning(APP_TITLE, "먼저 조회부터 해줘")
            return

        path = filedialog.asksaveasfilename(
            title="CSV 저장",
            defaultextension=".csv",
            filetypes=[("CSV", "*.csv")],
            initialfile="omm_reverse_export.csv",
        )
        if not path:
            return

        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            row1 = self.export_meta.get("xlsx_header_row1") or self.columns
            row2 = self.export_meta.get("xlsx_header_row2") or self.columns
            writer.writerow(row1)
            writer.writerow(row2)
            for row in self.rows:
                writer.writerow([format_cell(row.get(col), "") for col in self.columns])

        self.status_var.set("CSV 저장 완료: %s" % path)
        messagebox.showinfo(APP_TITLE, "CSV 저장 완료\n%s" % path)

    def export_xlsx(self):
        if not self.columns or not self.rows:
            messagebox.showwarning(APP_TITLE, "먼저 조회부터 해줘")
            return
        if not OPENPYXL_OK:
            messagebox.showerror(APP_TITLE, "openpyxl이 없어서 xlsx 저장 불가")
            return

        path = filedialog.asksaveasfilename(
            title="엑셀 저장",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile="omm_reverse_export.xlsx",
        )
        if not path:
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "OMM"

        row1 = self.export_meta.get("xlsx_header_row1") or self.columns
        row2 = self.export_meta.get("xlsx_header_row2") or self.columns
        ws.append(row1)
        ws.append(row2)
        for row in self.rows:
            ws.append([normalize_excel_cell(row.get(col)) for col in self.columns])

        for r1, c1, r2, c2, value in self.export_meta.get("merge_ranges", []):
            if c1 < c2:
                ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
            ws.cell(r1, c1).value = value

        header_font = Font(bold=True)
        center_align = Alignment(horizontal="center", vertical="center")
        for cell in ws[1]:
            cell.font = header_font
            cell.alignment = center_align
        for cell in ws[2]:
            cell.font = header_font
            cell.alignment = center_align

        ws.freeze_panes = "A3"
        for idx in range(1, len(self.columns) + 1):
            col = self.columns[idx - 1]
            width = 15
            if col == "fai":
                width = 28
            elif col in ("usl", "lsl"):
                width = 12
            elif "|" in col:
                width = 12
            ws.column_dimensions[get_excel_col(idx)].width = width

        wb.save(path)
        self.status_var.set("엑셀 저장 완료: %s" % path)
        messagebox.showinfo(APP_TITLE, "엑셀 저장 완료\n%s" % path)


def normalize_excel_cell(value):
    if value is None:
        return None
    return value


def add_labeled_entry(parent, label, var, row, col, width):
    ttk.Label(parent, text=label).grid(row=row, column=col, sticky="w", padx=(8, 4), pady=6)
    ent = ttk.Entry(parent, textvariable=var, width=width)
    ent.grid(row=row, column=col + 1, sticky="ew", padx=(0, 8), pady=6)
    return ent


def parse_tools(text):
    raw = re.split(r"[\s,;/]+", text.strip())
    out = []
    for item in raw:
        item = item.strip().upper()
        if not item:
            continue
        out.append(item)
    if not out:
        raise ValueError("툴 입력이 비어 있음")
    return out


def parse_cavities(text):
    text = text.strip()
    if not text:
        raise ValueError("캐비티 입력이 비어 있음")

    out = []
    for part in re.split(r"[\s,;/]+", text):
        part = part.strip()
        if not part:
            continue
        if "-" in part:
            a, b = part.split("-", 1)
            a_i = int(a)
            b_i = int(b)
            if a_i <= b_i:
                out.extend(list(range(a_i, b_i + 1)))
            else:
                out.extend(list(range(a_i, b_i - 1, -1)))
        else:
            out.append(int(part))

    if not out:
        raise ValueError("캐비티 해석 실패")
    return out


def parse_tc_text(text):
    raw = re.split(r"[,\n\r\t;]+", text)
    out = []
    seen = set()
    for item in raw:
        s = item.strip().upper().replace(" ", "")
        if not s:
            continue
        m = re.match(r"^([A-Z0-9]+)#([0-9]+)$", s)
        if not m:
            raise ValueError("Tool#Cavity 형식 오류: %s" % item)
        norm = "%s#%s" % (m.group(1), int(m.group(2)))
        if norm not in seen:
            out.append(norm)
            seen.add(norm)
    return out


def normalize_date(text):
    text = text.strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y%m%d"):
        try:
            return datetime.datetime.strptime(text, fmt).strftime("%Y-%m-%d")
        except Exception:
            pass
    raise ValueError("날짜 형식 오류: %s" % text)


def format_cell(value, empty=""):
    if value is None:
        return empty
    if isinstance(value, float):
        s = ("%.6f" % value).rstrip("0").rstrip(".")
        return s
    return str(value)


def get_excel_col(n):
    out = ""
    while n > 0:
        n, rem = divmod(n - 1, 26)
        out = chr(65 + rem) + out
    return out


if __name__ == "__main__":
    app = App()
    app.mainloop()
